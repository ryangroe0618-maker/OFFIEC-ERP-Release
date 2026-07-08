# -*- coding: utf-8 -*-

import os
import re
import tempfile
import time
from datetime import datetime
from io import BytesIO, StringIO

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from office_erp_paths import LIST_DIR as DESKTOP_LIST_DIR


KREAM_ORIGINAL_URL = (
    "https://docs.google.com/spreadsheets/d/e/"
    "2PACX-1vQSGUP7ZizELC9jJqGdxawP_HRsxo-fUZ_5BgMx8NUaEPTJYhH31iiH2-_NEE4Ff9UW3OATcxopoKEU/"
    "pub?gid=0&single=true&output=csv"
)
KREAM_SHIPMENT_URL = (
    "https://docs.google.com/spreadsheets/d/e/"
    "2PACX-1vQSGUP7ZizELC9jJqGdxawP_HRsxo-fUZ_5BgMx8NUaEPTJYhH31iiH2-_NEE4Ff9UW3OATcxopoKEU/"
    "pub?gid=226771413&single=true&output=csv"
)

LIST_DIR = str(DESKTOP_LIST_DIR)

OFFICE_PREFIX = "사무실 -"
COURIER_NAME = "CJ대한통운"

HEADER_ROW = 1
DATA_START_ROW = 2
START_COL = 1

MIN_COL_WIDTH = 8
MAX_COL_WIDTH = 60
COL_WIDTH_PADDING = 2

SESSION = requests.Session()
REQUEST_TIMEOUT = 30
RETRY_COUNT = 3


class KreamExportError(Exception):
    pass


def log(message):
    print(f"[KREAM 출고 등록] {message}")


def ensure_dir(path):
    os.makedirs(path, exist_ok=True)


def mmdd():
    return datetime.now().strftime("%m%d")


def safe_path(path):
    if not os.path.exists(path):
        return path

    base, ext = os.path.splitext(path)
    index = 1
    while True:
        new_path = f"{base}_{index}{ext}"
        if not os.path.exists(new_path):
            return new_path
        index += 1


def clean_text(value):
    if value is None:
        return ""
    text = str(value).replace("\ufeff", "").strip()
    if text.lower() in {"nan", "none", "nat"}:
        return ""
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def clean_df(df):
    if hasattr(df, "map"):
        return df.map(clean_text)
    return df.applymap(clean_text)


def normalize_order_no(value):
    return re.sub(r"\s+", "", clean_text(value))


def download_csv_df(url, label):
    last_error = None

    for attempt in range(1, RETRY_COUNT + 1):
        try:
            log(f"{label} 다운로드 시작 ({attempt}/{RETRY_COUNT})")
            response = SESSION.get(url, timeout=REQUEST_TIMEOUT)
            response.raise_for_status()
            content = response.content

            for encoding in ("utf-8-sig", "utf-8", "cp949"):
                try:
                    text = content.decode(encoding)
                    return pd.read_csv(StringIO(text), dtype=str, keep_default_na=False)
                except UnicodeDecodeError:
                    continue

            return pd.read_csv(BytesIO(content), dtype=str, keep_default_na=False)
        except Exception as exc:
            last_error = exc
            log(f"{label} 다운로드 실패 ({attempt}/{RETRY_COUNT}): {exc}")
            if attempt < RETRY_COUNT:
                time.sleep(1)

    raise KreamExportError(f"{label} CSV를 다운로드하지 못했습니다: {last_error}")


def require_columns(df, min_count, label):
    if df.shape[1] < min_count:
        raise KreamExportError(f"{label} 컬럼 수가 부족합니다. 필요: {min_count}개 / 실제: {df.shape[1]}개")


def load_original_df():
    df = download_csv_df(KREAM_ORIGINAL_URL, "KREAM 원본")
    require_columns(df, 13, "KREAM 원본")
    df = df.iloc[:, :13].copy()
    df.columns = [clean_text(col).split(".")[0] for col in df.columns]
    return clean_df(df)


def load_shipment_df():
    df = download_csv_df(KREAM_SHIPMENT_URL, "KREAM 출고 시트")
    require_columns(df, 11, "KREAM 출고 시트")

    shipment_df = df.iloc[:, [3, 4, 10]].copy()
    shipment_df.columns = ["주문번호", "운송장번호", "매장명"]
    shipment_df = clean_df(shipment_df)
    shipment_df = shipment_df[shipment_df["매장명"].str.startswith(OFFICE_PREFIX, na=False)].copy()

    if shipment_df.empty:
        raise KreamExportError(f'출고 시트에서 "{OFFICE_PREFIX}"로 시작하는 매장명이 없습니다.')

    shipment_df["주문번호키"] = shipment_df["주문번호"].map(normalize_order_no)
    shipment_df = shipment_df[shipment_df["주문번호키"] != ""].copy()
    if shipment_df.empty:
        raise KreamExportError("출고 시트에 사용할 주문번호가 없습니다.")

    return shipment_df


def build_tracking_lookup(shipment_df):
    tracking_by_order = {}
    duplicate_count = 0

    for _, row in shipment_df.iterrows():
        order_key = row["주문번호키"]
        tracking_no = clean_text(row["운송장번호"])
        if not order_key:
            continue
        if order_key in tracking_by_order:
            duplicate_count += 1
            if not tracking_by_order[order_key] and tracking_no:
                tracking_by_order[order_key] = tracking_no
            continue
        tracking_by_order[order_key] = tracking_no

    if not tracking_by_order:
        raise KreamExportError("출고 시트에서 주문번호 매칭 자료를 만들지 못했습니다.")

    if duplicate_count:
        log(f"중복 주문번호 {duplicate_count}건은 첫 번째 값을 기준으로 처리했습니다.")

    return tracking_by_order


def build_export_df(original_df, tracking_by_order):
    output_df = original_df.copy()
    order_keys = output_df.iloc[:, 0].map(normalize_order_no)
    matched_mask = order_keys.isin(tracking_by_order)
    output_df = output_df[matched_mask].copy()

    if output_df.empty:
        raise KreamExportError("원본 A열 주문번호와 출고 시트 주문번호가 매칭된 행이 없습니다.")

    matched_order_keys = order_keys[matched_mask].tolist()
    output_df.iloc[:, 10] = COURIER_NAME
    output_df.iloc[:, 11] = [tracking_by_order.get(order_key, "") for order_key in matched_order_keys]

    return output_df


def set_auto_width(ws, end_row, start_col, end_col):
    for col_idx in range(start_col, end_col + 1):
        max_len = 0
        for row_idx in range(HEADER_ROW, end_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            text = "" if value is None else str(value)
            text = re.sub(r"\s+", " ", text).strip()
            max_len = max(max_len, len(text))
        width = max(MIN_COL_WIDTH, min(max_len + COL_WIDTH_PADDING, MAX_COL_WIDTH))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def excel_value(value):
    text = clean_text(value)
    if text == "":
        return ""

    number_text = text.replace(",", "")
    if re.fullmatch(r"-?\d+", number_text):
        digits = number_text[1:] if number_text.startswith("-") else number_text
        if len(digits) <= 15 and (digits == "0" or not digits.startswith("0")):
            return int(number_text)

    if re.fullmatch(r"-?\d+\.\d+", number_text):
        integer_part = number_text.split(".", 1)[0].lstrip("-")
        if len(integer_part) <= 15 and (integer_part == "0" or not integer_part.startswith("0")):
            return float(number_text)

    return text


def save_excel(df, out_path):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "KREAM 출고"

    row_count = len(df)
    col_count = df.shape[1]

    for col_offset, header in enumerate(df.columns):
        worksheet.cell(row=HEADER_ROW, column=START_COL + col_offset, value=header)

    for row_offset in range(row_count):
        for col_offset in range(col_count):
            value = excel_value(df.iat[row_offset, col_offset])
            cell = worksheet.cell(
                row=DATA_START_ROW + row_offset,
                column=START_COL + col_offset,
                value=value,
            )
            if isinstance(value, int):
                cell.number_format = "0"
            elif isinstance(value, float):
                cell.number_format = "0.############"

    last_row = DATA_START_ROW + row_count - 1
    end_col = START_COL + col_count - 1
    set_auto_width(worksheet, last_row, START_COL, end_col)

    temp_dir = os.path.dirname(out_path) or "."
    fd, temp_path = tempfile.mkstemp(prefix=".kream_export_", suffix=".xlsx", dir=temp_dir)
    os.close(fd)

    try:
        workbook.save(temp_path)
        os.replace(temp_path, out_path)
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)


def main():
    ensure_dir(LIST_DIR)

    original_df = load_original_df()
    shipment_df = load_shipment_df()
    tracking_by_order = build_tracking_lookup(shipment_df)
    export_df = build_export_df(original_df, tracking_by_order)

    out_name = f"KREAM 출고 {mmdd()}.xlsx"
    out_path = safe_path(os.path.join(LIST_DIR, out_name))
    save_excel(export_df, out_path)

    log(f"저장 완료: {out_path}")
    log(f"출력 행 수: {len(export_df)}건")


if __name__ == "__main__":
    main()
