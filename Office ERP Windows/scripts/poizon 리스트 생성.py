# ot_ot_from_csv.py
# -*- coding: utf-8 -*-

import os
import re
from datetime import datetime
from io import StringIO, BytesIO
from typing import List, Tuple

import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font

from office_erp_paths import LIST_DIR as DESKTOP_LIST_DIR


# =========================
# 설정
# =========================
CSV_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vSzpo-DtuISMc_boM0XnjqnY-1hIlD2s_LMhzbvaRWdeNFBxdtO1Z0Fl94s4Dxo52wOwrBfisgDyQYt/pub?gid=1519011338&single=true&output=csv"

LIST_DIR = str(DESKTOP_LIST_DIR)

HEADER_ROW = 2      # B2부터 헤더
DATA_START_ROW = 3  # B3부터 데이터
START_COL = 2       # B열

MIN_COL_WIDTH = 10
MAX_COL_WIDTH = 70
COL_WIDTH_PADDING = 4

HEADER_FILL = PatternFill(start_color="F6D57A", end_color="F6D57A", fill_type="solid")

SESSION = requests.Session()


# =========================
# 유틸
# =========================
def ensure_dir(path):
    os.makedirs(path, exist_ok=True)

def mmdd():
    return datetime.now().strftime("%m%d")

def safe_path(path):
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    i = 1
    while True:
        new_path = f"{base}_{i}{ext}"
        if not os.path.exists(new_path):
            return new_path
        i += 1

def download_csv_df(url):
    r = SESSION.get(url, timeout=30)
    r.raise_for_status()
    content = r.content

    for enc in ("utf-8-sig", "utf-8", "cp949"):
        try:
            text = content.decode(enc)
            return pd.read_csv(StringIO(text), dtype=str, keep_default_na=False)
        except:
            continue

    return pd.read_csv(BytesIO(content), dtype=str, keep_default_na=False)

def excel_col_to_index(col_letter):
    col_letter = col_letter.strip().upper()
    n = 0
    for ch in col_letter:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n

def slice_df_by_excel_cols(df, start_col, end_col):
    s = excel_col_to_index(start_col) - 1
    e = excel_col_to_index(end_col)
    if df.shape[1] < e:
        raise ValueError("CSV 컬럼 수 부족")
    return df.iloc[:, s:e].copy()

def set_auto_width(ws, end_row, start_col, end_col):
    for c in range(start_col, end_col + 1):
        max_len = 0
        for r in range(HEADER_ROW, end_row + 1):
            v = ws.cell(row=r, column=c).value
            s = "" if v is None else str(v)
            s = re.sub(r"\s+", " ", s).strip()
            max_len = max(max_len, len(s))
        width = max(MIN_COL_WIDTH, min(max_len + COL_WIDTH_PADDING, MAX_COL_WIDTH))
        ws.column_dimensions[get_column_letter(c)].width = width


def to_excel_number(value):
    text = "" if value is None else str(value).strip()
    if text == "":
        return None
    text = text.replace(",", "")
    number = pd.to_numeric(text, errors="coerce")
    if pd.isna(number):
        return value
    number = float(number)
    if number.is_integer():
        return int(number)
    return number


# =========================
# 엑셀 생성
# =========================
def build_excel_from_df(df, out_path):
    wb = Workbook()
    ws = wb.active
    ws.sheet_view.showGridLines = False

    df.columns = [str(c).split(".")[0].strip() for c in df.columns]
    normalized_headers = [str(col).strip() for col in df.columns]
    quantity_col_indexes = [
        idx for idx, col in enumerate(normalized_headers)
        if col in {"수량", "사이즈"}
    ]

    n_rows = len(df)
    n_cols = df.shape[1]

    header_font = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 1️⃣ 헤더 B2부터
    for j in range(n_cols):
        cell = ws.cell(row=HEADER_ROW, column=START_COL + j, value=df.columns[j])
        cell.font = header_font
        cell.alignment = align_center

        # 앞 5열만 노란색
        if j < 5:
            cell.fill = HEADER_FILL

    # 2️⃣ 데이터 B3부터
    for i in range(n_rows):
        for j in range(n_cols):
            cell_value = df.iat[i, j]
            if j in quantity_col_indexes:
                cell_value = to_excel_number(cell_value)
            ws.cell(
                row=DATA_START_ROW + i,
                column=START_COL + j,
                value=cell_value
            )

    # 3️⃣ 테두리 (첫열 기준으로 값 있을 때만)
    last_row = DATA_START_ROW + n_rows - 1
    end_col = START_COL + n_cols - 1

    for r in range(HEADER_ROW, last_row + 1):

        first_val = ws.cell(row=r, column=START_COL).value

        # 첫열 값이 비어있으면 테두리 생략
        if first_val is None or str(first_val).strip() == "":
            continue

        for c in range(START_COL, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = align_center

    # 4️⃣ 열너비 자동 + 최소 10
    set_auto_width(ws, last_row, START_COL, end_col)

    wb.save(out_path)


# =========================
# 메인
# =========================
def main():
    ensure_dir(LIST_DIR)
    date_suffix = mmdd()
    df_all = download_csv_df(CSV_URL)

    jobs: List[Tuple[str, str, str]] = [
        ("노스페이스 OT", "A", "J"),
        ("휠라 OT", "L", "Q"),
        ("푸마 OT", "S", "V"),
    ]

    for name, s_col, e_col in jobs:
        part = slice_df_by_excel_cols(df_all, s_col, e_col)

        out_name = f"{name} {date_suffix}.xlsx"
        out_path = safe_path(os.path.join(LIST_DIR, out_name))

        build_excel_from_df(part, out_path)
        print(f"[OK] 저장 완료: {out_path}")

    print("[DONE] 전체 완료")


if __name__ == "__main__":
    main()
