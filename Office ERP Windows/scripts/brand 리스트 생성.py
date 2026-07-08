# -*- coding: utf-8 -*-

import os
import re
from datetime import datetime
from io import BytesIO, StringIO
from typing import List, Tuple

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from office_erp_paths import LIST_DIR as DESKTOP_LIST_DIR


CSV_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vSfBvby00YLSVYN-dPya7lNeGxtvDHfDDFiW0FwrhW3dHpIIYvupt2yW-t-QNZQhlRjJ98dHIWdNaMC/pub?gid=1726670761&single=true&output=csv"

LIST_DIR = str(DESKTOP_LIST_DIR)

HEADER_ROW = 2
DATA_START_ROW = 3
START_COL = 2

MIN_COL_WIDTH = 10
MAX_COL_WIDTH = 70
COL_WIDTH_PADDING = 4

HEADER_FILL = PatternFill(start_color="F6D57A", end_color="F6D57A", fill_type="solid")

SESSION = requests.Session()


def ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)


def mmdd() -> str:
    return datetime.now().strftime("%m%d")


def safe_path(path: str) -> str:
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    i = 1
    while True:
        new_path = f"{base}_{i}{ext}"
        if not os.path.exists(new_path):
            return new_path
        i += 1


def download_csv_df(url: str) -> pd.DataFrame:
    response = SESSION.get(url, timeout=30)
    response.raise_for_status()
    content = response.content

    for enc in ("utf-8-sig", "utf-8", "cp949"):
        try:
            text = content.decode(enc)
            return pd.read_csv(StringIO(text), dtype=str, keep_default_na=False)
        except Exception:
            continue

    return pd.read_csv(BytesIO(content), dtype=str, keep_default_na=False)


def excel_col_to_index(col_letter: str) -> int:
    col_letter = col_letter.strip().upper()
    n = 0
    for ch in col_letter:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def slice_df_by_excel_cols(df: pd.DataFrame, start_col: str, end_col: str) -> pd.DataFrame:
    start_idx = excel_col_to_index(start_col) - 1
    end_idx = excel_col_to_index(end_col)
    if df.shape[1] < end_idx:
        raise ValueError(f"CSV 컬럼 수 부족: {start_col}~{end_col}")
    return df.iloc[:, start_idx:end_idx].copy()


def trim_empty_edges(df: pd.DataFrame) -> pd.DataFrame:
    result = df.copy()
    while len(result.columns) > 0:
        first_col = result.iloc[:, 0].astype(str).str.strip()
        if first_col.replace({"nan": ""}).eq("").all():
            result = result.iloc[:, 1:]
        else:
            break
    while len(result.columns) > 0:
        last_col = result.iloc[:, -1].astype(str).str.strip()
        if last_col.replace({"nan": ""}).eq("").all():
            result = result.iloc[:, :-1]
        else:
            break
    return result.fillna("")


def trim_blank_outer_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df.copy()

    row_has_value = ~df.apply(lambda row: row.astype(str).str.strip().eq("").all(), axis=1)
    if not row_has_value.any():
        return df.iloc[0:0].copy()

    first_idx = row_has_value.idxmax()
    last_idx = row_has_value[::-1].idxmax()
    return df.loc[first_idx:last_idx].reset_index(drop=True)


def has_data_in_first_column(df: pd.DataFrame) -> bool:
    if df.empty:
        return False
    first_col = df.iloc[:, 0].fillna("").astype(str).str.strip()
    return first_col.ne("").any()


def set_auto_width(ws, end_row: int, start_col: int, end_col: int):
    for c in range(start_col, end_col + 1):
        max_len = 0
        for r in range(HEADER_ROW, end_row + 1):
            value = ws.cell(row=r, column=c).value
            text = "" if value is None else str(value)
            text = re.sub(r"\s+", " ", text).strip()
            max_len = max(max_len, len(text))
        width = max(MIN_COL_WIDTH, min(max_len + COL_WIDTH_PADDING, MAX_COL_WIDTH))
        ws.column_dimensions[get_column_letter(c)].width = width


def build_excel_from_df(df: pd.DataFrame, out_path: str):
    wb = Workbook()
    ws = wb.active
    ws.sheet_view.showGridLines = False

    df = df.copy()
    df.columns = [str(col).split(".")[0].strip() for col in df.columns]

    header_font = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for j, col_name in enumerate(df.columns):
        cell = ws.cell(row=HEADER_ROW, column=START_COL + j, value=col_name)
        cell.font = header_font
        cell.alignment = align_center
        if j < 4:
            cell.fill = HEADER_FILL

    for i in range(len(df)):
        for j in range(df.shape[1]):
            ws.cell(
                row=DATA_START_ROW + i,
                column=START_COL + j,
                value=df.iat[i, j],
            )

    last_row = DATA_START_ROW + len(df) - 1
    end_col = START_COL + df.shape[1] - 1

    if len(df) > 0:
        for r in range(HEADER_ROW, last_row + 1):
            first_val = ws.cell(row=r, column=START_COL).value
            if first_val is None or str(first_val).strip() == "":
                continue
            for c in range(START_COL, end_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                cell.alignment = align_center
        set_auto_width(ws, last_row, START_COL, end_col)
    else:
        set_auto_width(ws, HEADER_ROW, START_COL, end_col)

    wb.save(out_path)


def main():
    ensure_dir(LIST_DIR)
    date_suffix = mmdd()
    df_all = download_csv_df(CSV_URL)

    jobs: List[Tuple[str, str, str]] = [
        ("노스페이스 B업체", "A", "E"),
        ("휠라 B업체", "K", "O"),
        ("푸마 B업체", "Q", "T"),
    ]

    for name, start_col, end_col in jobs:
        part = slice_df_by_excel_cols(df_all, start_col, end_col)
        part = trim_empty_edges(part)
        part = trim_blank_outer_rows(part)
        if not has_data_in_first_column(part):
            print(f"건너뜀: {name} (첫 열 데이터 없음)")
            continue

        out_name = f"{name} {date_suffix}.xlsx"
        out_path = safe_path(os.path.join(LIST_DIR, out_name))

        build_excel_from_df(part, out_path)
        print(f"저장 완료: {out_path}")

    print("전체 완료")


if __name__ == "__main__":
    main()
