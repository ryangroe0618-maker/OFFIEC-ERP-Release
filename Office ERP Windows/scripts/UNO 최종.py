# -*- coding: utf-8 -*-

import os
import re
from datetime import datetime
from io import StringIO
import time

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from office_erp_paths import LIST_DIR as DESKTOP_LIST_DIR


SOURCE_CSV_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vRWhzq2rt_O-YnOwzWnE8L_d8--NBu-EMmDkxnnVy-oATAfCjDIX976b973bY2MiO8gYPTE5WCk-tVv/pub?gid=733480714&single=true&output=csv"
RETURN_SCAN_CSV_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vQhP9cP1QdWll3UPE-P-tUAmxFHCEXgQU_IKIDsftokFeyn5Y67OW2Zho5xYN4pwQKvcclbDS98bQum/pub?gid=202616820&single=true&output=csv"
WEB_APP_URL = "https://script.google.com/macros/s/AKfycbxzBOrRpU-zDYT_p3vf1d4oPf1Bv5QZ8cvXJW5HAqqN1p9IsnSERlHMnXi609iIampX/exec"
SPREADSHEET_ID = "1Vm5Nxs76ELKyk7QFK_9ohRT2eFpXlVUseEhY5qyCE2A"
TARGET_SHEET_NAME = "UNO"
LIST_DIR = str(DESKTOP_LIST_DIR)
EXCEL_BASENAME = "KASHION"
EXCEL_START_ROW = 2
EXCEL_START_COL = 2
RETURN_EXCEL_START_ROW = 2
RETURN_EXCEL_START_COL = 18
RETURN_EXCEL_HEADER = "取回单号"
RETURN_CHANNELS = {"RETURNTM", "RETURNJD"}

DOWNLOAD_TIMEOUT = (10, 60)
UPLOAD_TIMEOUT = (10, 120)
RETRIES = 3
RETRY_SLEEP_SEC = 2
MIN_COL_WIDTH = 12
MAX_COL_WIDTH = 60
COL_WIDTH_PADDING = 4
ROW_HEIGHT = 18
HEADER_FILL = PatternFill(start_color="F6D57A", end_color="F6D57A", fill_type="solid")
HIGHLIGHT_PINK_FILL = PatternFill(start_color="F4B6C2", end_color="F4B6C2", fill_type="solid")
HIGHLIGHT_PINK_FONT = Font(color="C62828")
HEADER_FONT = Font(bold=True)
class DataValidationError(ValueError):
    pass


def make_session():
    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0"})
    return session


def clean_text(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)


def today_label() -> str:
    now = datetime.now()
    return f"{now.month}월 {now.day}일"


def is_today_value(value) -> bool:
    text = clean_text(value)
    if text == "":
        return False

    now = datetime.now()
    today_candidates = {
        today_label(),
        now.strftime("%Y-%m-%d"),
        now.strftime("%Y/%m/%d"),
        now.strftime("%Y.%m.%d"),
        f"{now.month}/{now.day}",
        f"{now.month:02d}/{now.day:02d}",
        f"{now.month}.{now.day}",
        f"{now.month:02d}.{now.day:02d}",
    }
    if text in today_candidates:
        return True

    parsed = pd.to_datetime(text, errors="coerce")
    if pd.isna(parsed):
        return False
    return parsed.date() == now.date()


def normalize_channel(value) -> str:
    return clean_text(value).upper()


def extract_platform_from_order_no(order_no: str) -> str:
    prefix = clean_text(order_no)[:2].upper()
    if prefix == "LP":
        return "TM"
    if prefix == "JD":
        return "JD"
    return ""


def mmdd() -> str:
    return datetime.now().strftime("%m%d")


def safe_path(path: str) -> str:
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    i = 1
    while True:
        candidate = f"{base}_{i}{ext}"
        if not os.path.exists(candidate):
            return candidate
        i += 1


def fetch_csv_text(session: requests.Session, url: str) -> str:
    last_error = None

    for attempt in range(1, RETRIES + 1):
        try:
            response = session.get(url, timeout=DOWNLOAD_TIMEOUT)
            response.raise_for_status()
            response.encoding = "utf-8"
            return response.text
        except Exception as e:
            last_error = e
            if attempt < RETRIES:
                time.sleep(RETRY_SLEEP_SEC)

    raise last_error


def validate_apps_script_response(response_text: str) -> dict:
    text = (response_text or "").strip()
    try:
        response_json = requests.models.complexjson.loads(text)
    except Exception:
        response_json = None

    if isinstance(response_json, dict):
        if response_json.get("ok") is False:
            raise RuntimeError(
                response_json.get("error")
                or response_json.get("message")
                or text
            )
        return response_json

    lowered = text.lower()
    if "<html" in lowered or "<!doctype html" in lowered:
        if "doget" in lowered:
            raise RuntimeError(
                "Apps Script가 doGet 오류 HTML을 반환했습니다. "
                "웹앱 배포 코드에 doGet/doPost가 있는지 확인하고 새 배포해 주세요."
            )
        if "dopost" in lowered:
            raise RuntimeError(
                "Apps Script가 doPost 오류 HTML을 반환했습니다. "
                "웹앱 배포 코드에 doPost가 있는지 확인하고 새 배포해 주세요."
            )
        raise RuntimeError(f"Apps Script가 JSON 대신 HTML을 반환했습니다: {text[:300]}")

    raise RuntimeError(f"Apps Script 응답을 JSON으로 확인할 수 없습니다: {text[:300]}")


def load_source_df(session: requests.Session) -> pd.DataFrame:
    csv_text = fetch_csv_text(session, SOURCE_CSV_URL)
    df_raw = pd.read_csv(StringIO(csv_text), dtype=str, keep_default_na=False).fillna("")
    if df_raw.shape[1] < 20:
        raise DataValidationError(
            f"원본 시트 열 수가 부족합니다. A:T 기준 20열이 필요한데 실제 {df_raw.shape[1]}열입니다."
        )

    df = df_raw.iloc[:, :20].copy()
    df.columns = [clean_text(col) for col in df.columns]

    date_series = df.iloc[:, 0].apply(is_today_value)
    channel_series = df.iloc[:, 1].apply(normalize_channel).eq("KASHION")
    filtered = df[date_series & channel_series].copy()

    base_columns = ["订单号", "货号", "尺码", "数量", "快递", "单号", "备注"]
    output_columns = base_columns + base_columns

    if filtered.empty:
        return pd.DataFrame(columns=output_columns)

    picked = pd.DataFrame({
        "플랫폼": filtered.iloc[:, 2].apply(extract_platform_from_order_no),
        "订单号": filtered.iloc[:, 2].apply(clean_text),
        "货号": filtered.iloc[:, 6].apply(clean_text),
        "尺码": filtered.iloc[:, 7].apply(clean_text),
        "数量": filtered.iloc[:, 8].apply(clean_text),
        "快递": filtered.iloc[:, 18].apply(clean_text),
        "单号": filtered.iloc[:, 19].apply(clean_text),
    })
    picked["备注"] = ""
    picked = picked[picked["플랫폼"].isin(["TM", "JD"])].reset_index(drop=True)

    if picked.empty:
        return pd.DataFrame(columns=output_columns)

    tm_df = (
        picked[picked["플랫폼"].apply(clean_text).str.upper() == "TM"][base_columns]
        .sort_values(
            by=["快递", "单号", "订单号", "货号", "尺码"],
            ascending=[False, True, True, True, True],
            kind="stable",
        )
        .reset_index(drop=True)
    )
    jd_df = (
        picked[picked["플랫폼"].apply(clean_text).str.upper() == "JD"][base_columns]
        .sort_values(
            by=["快递", "单号", "订单号", "货号", "尺码"],
            ascending=[False, True, True, True, True],
            kind="stable",
        )
        .reset_index(drop=True)
    )

    max_len = max(len(tm_df), len(jd_df))
    tm_df = tm_df.reindex(range(max_len), fill_value="")
    jd_df = jd_df.reindex(range(max_len), fill_value="")

    result = pd.concat(
        [
            tm_df.reset_index(drop=True),
            jd_df.reset_index(drop=True),
        ],
        axis=1,
    )
    result.columns = output_columns
    return result


def load_return_tracking_numbers(session: requests.Session) -> list[str]:
    csv_text = fetch_csv_text(session, RETURN_SCAN_CSV_URL)
    df_raw = pd.read_csv(StringIO(csv_text), dtype=str, header=None, keep_default_na=False).fillna("")
    if df_raw.shape[0] < 3 or df_raw.shape[1] < 12:
        raise DataValidationError(
            f"출고 스캔 시트 구조가 부족합니다. 3행 머리글과 L열까지 필요, 실제 {df_raw.shape[0]}행 {df_raw.shape[1]}열입니다."
        )

    df = df_raw.iloc[2:, [4, 11]].copy()
    df.columns = [clean_text(col) for col in df.iloc[0].tolist()]
    df = df.iloc[1:].reset_index(drop=True)

    if "업체명" not in df.columns or "운송장" not in df.columns:
        raise DataValidationError("출고 스캔 시트 E/L열 3행 머리글이 '업체명', '운송장'인지 확인해주세요.")

    channel_series = df["업체명"].apply(lambda value: clean_text(value).upper())
    tracking_series = df["운송장"].apply(clean_text)
    return tracking_series[channel_series.isin(RETURN_CHANNELS) & tracking_series.ne("")].tolist()


def upload_to_google_sheet(df: pd.DataFrame):
    spacer_col = "\u00A0"
    upload_df = pd.concat(
        [
            df.iloc[:, :7].copy(),
            pd.DataFrame({spacer_col: ["\u00A0"] * len(df)}),
            df.iloc[:, 7:].copy(),
        ],
        axis=1,
    )

    payload = {
        "spreadsheetId": SPREADSHEET_ID,
        "sheetName": TARGET_SHEET_NAME,
        "values": [upload_df.columns.tolist()] + upload_df.fillna("").astype(str).values.tolist(),
        "clear": True,
    }

    last_error = None
    for attempt in range(1, RETRIES + 1):
        try:
            response = requests.post(WEB_APP_URL, json=payload, timeout=UPLOAD_TIMEOUT)
            response.raise_for_status()
            response_text = (response.text or "").strip()
            validate_apps_script_response(response_text)
            print(f"업로드 완료: {TARGET_SHEET_NAME} / {len(df)}행")
            print("응답:", response_text)
            return
        except Exception as e:
            last_error = e
            print(f"업로드 실패 ({attempt}/{RETRIES}): {e}")
            if attempt < RETRIES:
                time.sleep(RETRY_SLEEP_SEC)

    raise last_error


def set_auto_width(ws, start_row: int, end_row: int, start_col: int, end_col: int):
    for c in range(start_col, end_col + 1):
        max_len = 0
        for r in range(start_row, end_row + 1):
            value = ws.cell(row=r, column=c).value
            text = "" if value is None else str(value)
            text = re.sub(r"\s+", " ", text).strip()
            max_len = max(max_len, len(text))
        width = max(MIN_COL_WIDTH, min(max_len + COL_WIDTH_PADDING, MAX_COL_WIDTH))
        ws.column_dimensions[get_column_letter(c)].width = width


def save_to_excel(df: pd.DataFrame, return_tracking_numbers: list[str]):
    ensure_dir(LIST_DIR)

    wb = Workbook()
    ws = wb.active
    ws.title = TARGET_SHEET_NAME
    ws.sheet_view.showGridLines = False

    align_center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_row = EXCEL_START_ROW
    data_start_row = EXCEL_START_ROW + 1
    start_col = EXCEL_START_COL
    separator_after = start_col + 6  # left block 7 columns after B
    right_block_start = separator_after + 2

    excel_headers = (
        df.columns[:7].tolist()
        + [""]
        + df.columns[7:].tolist()
    )
    numeric_headers = {"数量"}

    for offset, col_name in enumerate(excel_headers):
        col_idx = start_col + offset
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.alignment = align_center
        if col_name != "":
            cell.fill = HEADER_FILL

    for row_offset, row in enumerate(df.fillna("").values.tolist()):
        row_idx = data_start_row + row_offset
        left_values = row[:7]
        right_values = row[7:]

        for value_offset, value in enumerate(left_values):
            col_idx = start_col + value_offset
            header_name = excel_headers[value_offset]
            cell_value = value
            if header_name in numeric_headers:
                try:
                    cell_value = int(float(clean_text(value))) if clean_text(value) != "" else None
                except Exception:
                    cell_value = clean_text(value)
            else:
                cell_value = clean_text(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.alignment = align_center

        for value_offset, value in enumerate(right_values):
            col_idx = right_block_start + value_offset
            header_name = excel_headers[8 + value_offset]
            cell_value = value
            if header_name in numeric_headers:
                try:
                    cell_value = int(float(clean_text(value))) if clean_text(value) != "" else None
                except Exception:
                    cell_value = clean_text(value)
            else:
                cell_value = clean_text(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.alignment = align_center

    ws.row_dimensions[header_row].height = ROW_HEIGHT
    for row_idx in range(data_start_row, data_start_row + len(df)):
        ws.row_dimensions[row_idx].height = ROW_HEIGHT

    if df.shape[1] > 0:
        last_row = max(data_start_row + len(df) - 1, header_row)
        left_order_col = start_col
        left_end_col = separator_after
        right_order_col = right_block_start
        right_end_col = right_block_start + len(df.columns[7:]) - 1

        for col_idx in range(start_col, left_end_col + 1):
            ws.cell(row=header_row, column=col_idx).border = border
        for col_idx in range(right_block_start, right_end_col + 1):
            ws.cell(row=header_row, column=col_idx).border = border

        for row_idx in range(data_start_row, last_row + 1):
            left_order_value = ws.cell(row=row_idx, column=left_order_col).value
            right_order_value = ws.cell(row=row_idx, column=right_order_col).value

            if left_order_value is not None and str(left_order_value).strip() != "":
                for col_idx in range(start_col, left_end_col + 1):
                    ws.cell(row=row_idx, column=col_idx).border = border

            if right_order_value is not None and str(right_order_value).strip() != "":
                for col_idx in range(right_block_start, right_end_col + 1):
                    ws.cell(row=row_idx, column=col_idx).border = border

        duplicate_counts = {}
        for row_idx in range(data_start_row, last_row + 1):
            for col_idx in (left_order_col, right_order_col):
                order_value = clean_text(ws.cell(row=row_idx, column=col_idx).value)
                if order_value == "":
                    continue
                duplicate_counts[order_value] = duplicate_counts.get(order_value, 0) + 1

        duplicate_values = {value for value, count in duplicate_counts.items() if count > 1}
        if duplicate_values:
            for row_idx in range(data_start_row, last_row + 1):
                for col_idx in (left_order_col, right_order_col):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    order_value = clean_text(cell.value)
                    if order_value in duplicate_values:
                        cell.fill = HIGHLIGHT_PINK_FILL
                        cell.font = HIGHLIGHT_PINK_FONT

        set_auto_width(ws, header_row, last_row, start_col, separator_after)
        set_auto_width(ws, header_row, last_row, right_block_start, right_end_col)
        ws.column_dimensions[get_column_letter(separator_after + 1)].width = 8

    return_header_cell = ws.cell(
        row=RETURN_EXCEL_START_ROW,
        column=RETURN_EXCEL_START_COL,
        value=RETURN_EXCEL_HEADER,
    )
    return_header_cell.font = HEADER_FONT
    return_header_cell.alignment = align_center
    return_header_cell.fill = HEADER_FILL
    return_header_cell.border = border

    for row_offset, tracking_no in enumerate(return_tracking_numbers, start=1):
        cell = ws.cell(
            row=RETURN_EXCEL_START_ROW + row_offset,
            column=RETURN_EXCEL_START_COL,
            value=clean_text(tracking_no),
        )
        cell.alignment = align_center
        cell.border = border

    return_last_row = max(RETURN_EXCEL_START_ROW + len(return_tracking_numbers), RETURN_EXCEL_START_ROW)
    set_auto_width(
        ws,
        RETURN_EXCEL_START_ROW,
        return_last_row,
        RETURN_EXCEL_START_COL,
        RETURN_EXCEL_START_COL,
    )

    out_name = f"{EXCEL_BASENAME} {mmdd()}.xlsx"
    out_path = safe_path(os.path.join(LIST_DIR, out_name))
    wb.save(out_path)
    print(f"엑셀 저장 완료: {out_path}")


def main():
    session = make_session()
    output_df = load_source_df(session)
    return_tracking_numbers = load_return_tracking_numbers(session)
    upload_to_google_sheet(output_df)
    save_to_excel(output_df, return_tracking_numbers)


if __name__ == "__main__":
    main()
