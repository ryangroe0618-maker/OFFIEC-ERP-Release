# -*- coding: utf-8 -*-

from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from io import StringIO
import os
import re
import time
from zoneinfo import ZoneInfo

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from office_erp_paths import LIST_DIR as DESKTOP_LIST_DIR


SOURCE_CSV_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vSfBvby00YLSVYN-dPya7lNeGxtvDHfDDFiW0FwrhW3dHpIIYvupt2yW-t-QNZQhlRjJ98dHIWdNaMC/pub?gid=594145141&single=true&output=csv"
STORE_URLS = {
    "스퀘어원": [
        "https://docs.google.com/spreadsheets/d/e/2PACX-1vTPSUW1W1iSIvGGrLkp1WHj6Dy_k4NQHv5xOZR4xviYMsZWUb6ZBQ4PqeI31RM_keSDaXeQsYyNLAav/pub?gid=1199416354&single=true&output=csv",
        "https://docs.google.com/spreadsheets/d/e/2PACX-1vTPSUW1W1iSIvGGrLkp1WHj6Dy_k4NQHv5xOZR4xviYMsZWUb6ZBQ4PqeI31RM_keSDaXeQsYyNLAav/pub?gid=1405262882&single=true&output=csv",
    ],
    "구월": [
        "https://docs.google.com/spreadsheets/d/e/2PACX-1vSHTZiYHkTrDlZ_pi1qxBsikvBAaMxtdEzwSYsWzk6sV1zk04SIYjflfnxMYRsmwevPovu4Mtnlx69M/pub?gid=1199416354&single=true&output=csv",
        "https://docs.google.com/spreadsheets/d/e/2PACX-1vSHTZiYHkTrDlZ_pi1qxBsikvBAaMxtdEzwSYsWzk6sV1zk04SIYjflfnxMYRsmwevPovu4Mtnlx69M/pub?gid=1405262882&single=true&output=csv",
    ],
    "부천": [
        "https://docs.google.com/spreadsheets/d/e/2PACX-1vQdYpr-dKLe-tguI2uOaYL9pjalY0jehboc1zb-B5XKbV8vAPQvtw1S4nu-TaxJULDsoKOTz8gz7A5y/pub?gid=1199416354&single=true&output=csv",
        "https://docs.google.com/spreadsheets/d/e/2PACX-1vQdYpr-dKLe-tguI2uOaYL9pjalY0jehboc1zb-B5XKbV8vAPQvtw1S4nu-TaxJULDsoKOTz8gz7A5y/pub?gid=1405262882&single=true&output=csv",
    ],
    "휠라 파주": [
        "https://docs.google.com/spreadsheets/d/e/2PACX-1vSA47SgFq9QQPg0D3AlBnpJX6q7Yx_Dh66E1ID9MlXTahJjL0FmFVtPgyTEtj4iVj7PvRkCUoCgbjkd/pub?gid=1226300508&single=true&output=csv",
    ],
    "푸마 여주": [
        "https://docs.google.com/spreadsheets/d/e/2PACX-1vSzkIBQ7UfqnUboNBWaQj6esNZzi_NSk0crAVPCljFog-YAnl1vSY6gqqTxH2CYosDoRL4q2PgMUhqL/pub?gid=1405262882&single=true&output=csv",
    ],
}
OFFICE_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vQhP9cP1QdWll3UPE-P-tUAmxFHCEXgQU_IKIDsftokFeyn5Y67OW2Zho5xYN4pwQKvcclbDS98bQum/pub?gid=202616820&single=true&output=csv"
WEB_APP_URL = "https://script.google.com/macros/s/AKfycbzjGo2ZxLTaaIyUttkcrSThwS6mZA6FB9nw47yJSghFScG-SgS1W7j_JZTykWoXr8kV/exec"
SPREADSHEET_ID = "1jHdxOK_Y9MrLyTZD1Fj62g21zCHG0LL9N_X72HvWz2k"
TARGET_SHEET_NAME = "취합"
LIST_DIR = str(DESKTOP_LIST_DIR)
EXCEL_BASENAME = "브랜더 출고 내역"

INTERNAL_COLUMNS = ["날짜", "업체명", "브랜드", "코드", "품번", "사이즈", "수량"]
STORE_QTY_COLUMNS = [
    "사무실",
    "스퀘어원",
    "구월",
    "부천",
    "휠라 파주",
    "푸마 여주",
]
OUTPUT_COLUMNS = [
    "날짜",
    "브랜드",
    "품번",
    "사이즈",
    "주문 수량",
    "출고 수량",
    "부족 수량",
] + STORE_QTY_COLUMNS
SUM_COLUMNS = ["주문 수량", "출고 수량", "부족 수량"] + STORE_QTY_COLUMNS
DETAIL_COLUMNS = ["품번", "컬러", "사이즈", "수량", "매장명"]
DETAIL_SUM_COLUMNS = ["수량"]
BRAND_BLOCK_COLUMNS = [
    "품번",
    "컬러",
    "사이즈",
    "수량",
    "매장명",
]
EXCEL_START_ROW = 2
EXCEL_START_COL = 2

DOWNLOAD_TIMEOUT = (10, 60)
UPLOAD_TIMEOUT = (10, 120)
RETRIES = 3
RETRY_SLEEP_SEC = 2
TODAY_TZ = ZoneInfo("Asia/Seoul")
MIN_COL_WIDTH = 12
MAX_COL_WIDTH = 30
COL_WIDTH_PADDING = 5
ROW_HEIGHT = 20
HEADER_FILL = PatternFill(start_color="F6D57A", end_color="F6D57A", fill_type="solid")
HEADER_FONT = Font(bold=True)
RED_FONT = Font(color="C62828")


class DataValidationError(ValueError):
    pass


def log(message: str):
    print(f"[BRAND-출고] {message}", flush=True)


def make_session() -> requests.Session:
    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0"})
    return session


def clean_text(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    text = text.replace('"', "").replace("\t", "").replace("\r", "").replace("\n", "")
    text = " ".join(text.split())
    if text.startswith("'"):
        text = text[1:]
    if text.endswith(".0"):
        try:
            number = float(text)
            if number.is_integer():
                return str(int(number))
        except Exception:
            pass
    return text


def normalize_date_text(value) -> str:
    return clean_text(value).replace(" ", "")


def today_sheet_text() -> str:
    now = datetime.now(TODAY_TZ)
    return f"{now.month}월{now.day}일"


def display_date_text() -> str:
    now = datetime.now(TODAY_TZ)
    return f"{now.month}월 {now.day}일"


def to_int_series(series: pd.Series) -> pd.Series:
    return (
        pd.to_numeric(
            series.fillna("").astype(str).str.replace(",", "", regex=False).str.strip(),
            errors="coerce",
        )
        .fillna(0)
        .astype(int)
    )


def make_code_key(code) -> str:
    return clean_text(code)


def ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)


def ymd() -> str:
    return datetime.now(TODAY_TZ).strftime("%m%d")


def safe_path(path: str) -> str:
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    i = 1
    while True:
        candidate = f"{base}_{i}{ext}"
        if not os.path.exists(candidate):
            return candidate
        i += 1


def fetch_csv_text(session: requests.Session, url: str) -> str:
    last_error = None
    for attempt in range(1, RETRIES + 1):
        try:
            response = session.get(url, timeout=DOWNLOAD_TIMEOUT)
            response.raise_for_status()
            response.encoding = "utf-8"
            return response.text
        except Exception as e:
            last_error = e
            log(f"CSV 다운로드 실패 ({attempt}/{RETRIES}) -> {e}")
            if attempt < RETRIES:
                time.sleep(RETRY_SLEEP_SEC)
    raise last_error


def fetch_csv_job(name: str, url: str) -> tuple[str, str]:
    return name, fetch_csv_text(make_session(), url)


def download_all_csv_texts() -> dict[str, str]:
    url_map = {"source": SOURCE_CSV_URL, "office": OFFICE_URL}
    for store_name, urls in STORE_URLS.items():
        for idx, url in enumerate(urls, start=1):
            url_map[f"store::{store_name}::{idx}"] = url

    results = {}
    with ThreadPoolExecutor(max_workers=min(len(url_map), 8)) as executor:
        future_map = {
            executor.submit(fetch_csv_job, name, url): name
            for name, url in url_map.items()
        }
        for future in as_completed(future_map):
            name, text = future.result()
            results[name] = text
    return results


def parse_csv_text(csv_text: str) -> pd.DataFrame:
    return pd.read_csv(
        StringIO(csv_text),
        dtype=str,
        header=0,
        keep_default_na=False,
        skip_blank_lines=False,
    ).fillna("")


def load_source_df(csv_text: str) -> pd.DataFrame:
    df_raw = parse_csv_text(csv_text)
    if df_raw.shape[1] < 7:
        raise DataValidationError(
            f"원본 출고 시트 열 개수가 부족합니다. 최소 G열까지 필요하지만 실제는 {df_raw.shape[1]}열입니다."
        )

    df = df_raw.iloc[:, [0, 1, 2, 3, 4, 5, 6]].copy()
    df.columns = INTERNAL_COLUMNS

    for col in INTERNAL_COLUMNS:
        df[col] = df[col].map(clean_text)

    df["날짜"] = df["날짜"].map(normalize_date_text)
    df = df.loc[df["날짜"].eq(today_sheet_text()) & df["코드"].ne("")].reset_index(drop=True)
    if df.empty:
        raise DataValidationError("원본 출고 시트에 오늘 날짜 기준 코드 데이터가 없습니다.")

    df["수량"] = to_int_series(df["수량"])
    df = (
        df.groupby(["날짜", "업체명", "브랜드", "코드", "품번", "사이즈"], dropna=False, as_index=False)["수량"]
        .sum()
        .reset_index(drop=True)
    )
    df["날짜"] = display_date_text()

    return df


def load_store_df(csv_text: str, store_name: str) -> pd.DataFrame:
    df_raw = parse_csv_text(csv_text)
    if df_raw.shape[1] < 7:
        raise DataValidationError(
            f"{store_name} 시트 열 개수가 부족합니다. B:G 열이 필요하지만 실제는 {df_raw.shape[1]}열입니다."
        )

    df = df_raw.iloc[:, 1:7].copy()
    df.columns = ["날짜", "출고처", "코드", "품번", "사이즈", "수량"]

    for col in ["날짜", "출고처", "코드", "품번", "사이즈", "수량"]:
        df[col] = df[col].map(clean_text)
    df["날짜"] = df["날짜"].map(normalize_date_text)
    df = df.loc[
        df["날짜"].eq(today_sheet_text())
        & df["출고처"].str.upper().eq("B")
        & df["코드"].ne("")
    ].reset_index(drop=True)
    df["업체명"] = "B"
    df["브랜드"] = ""
    df["수량"] = to_int_series(df["수량"])
    return df[INTERNAL_COLUMNS]


def load_office_df(csv_text: str) -> pd.DataFrame:
    df_raw = parse_csv_text(csv_text)
    required_indexes = [1, 4, 5, 6, 7, 9]
    if df_raw.shape[1] <= max(required_indexes):
        raise DataValidationError(
            f"사무실 시트 열 개수가 부족합니다. B,E,F,G,H,J 열이 필요하지만 실제는 {df_raw.shape[1]}열입니다."
        )

    df = df_raw.iloc[:, required_indexes].copy()
    df.columns = ["날짜", "업체명", "코드", "품번", "사이즈", "수량"]

    for col in ["날짜", "업체명", "코드", "품번", "사이즈", "수량"]:
        df[col] = df[col].map(clean_text)

    df["날짜"] = df["날짜"].map(normalize_date_text)
    df["업체명"] = df["업체명"].str.upper()
    df = df.loc[
        df["날짜"].eq(today_sheet_text())
        & df["업체명"].eq("BRAND")
        & df["코드"].ne("")
    ].reset_index(drop=True)
    df["브랜드"] = ""
    df["수량"] = to_int_series(df["수량"])
    return df[INTERNAL_COLUMNS]


def summarize_qty_by_code(df: pd.DataFrame) -> dict[str, int]:
    if df.empty:
        return {}
    grouped = df.groupby("코드", dropna=False)["수량"].sum()
    return {make_code_key(code): int(qty) for code, qty in grouped.items() if make_code_key(code)}


def summarize_store_with_fallback(dfs: list[pd.DataFrame]) -> dict[str, int]:
    first_lookup = summarize_qty_by_code(dfs[0]) if len(dfs) >= 1 else {}
    second_lookup = summarize_qty_by_code(dfs[1]) if len(dfs) >= 2 else {}
    result = {}
    for key in sorted(set(first_lookup) | set(second_lookup)):
        if key in first_lookup and first_lookup[key] > 0:
            result[key] = first_lookup[key]
        elif key in second_lookup:
            result[key] = second_lookup[key]
    return result


def attach_summary_columns(
    source_df: pd.DataFrame,
    office_lookup: dict[str, int],
    square_lookup: dict[str, int],
    guwol_lookup: dict[str, int],
    bucheon_lookup: dict[str, int],
    fila_lookup: dict[str, int],
    puma_lookup: dict[str, int],
) -> pd.DataFrame:
    result = source_df.copy()

    result["사무실"] = result["코드"].map(lambda key: office_lookup.get(make_code_key(key), 0))
    result["스퀘어원"] = result["코드"].map(lambda key: square_lookup.get(make_code_key(key), 0))
    result["구월"] = result["코드"].map(lambda key: guwol_lookup.get(make_code_key(key), 0))
    result["부천"] = result["코드"].map(lambda key: bucheon_lookup.get(make_code_key(key), 0))
    result["휠라 파주"] = result["코드"].map(lambda key: fila_lookup.get(make_code_key(key), 0))
    result["푸마 여주"] = result["코드"].map(lambda key: puma_lookup.get(make_code_key(key), 0))
    result["주문 수량"] = to_int_series(result["수량"])
    result["출고 수량"] = (
        result["사무실"]
        + result["스퀘어원"]
        + result["구월"]
        + result["부천"]
        + result["휠라 파주"]
        + result["푸마 여주"]
    )
    result["부족 수량"] = result["출고 수량"] - result["주문 수량"]

    for col in ["주문 수량", "출고 수량", "부족 수량"] + STORE_QTY_COLUMNS:
        result[col] = result[col].astype(int).astype(str).replace("0", "")

    result = result.sort_values(
        by=["브랜드", "품번", "사이즈", "코드"],
        ascending=[True, True, True, True],
        kind="stable",
    ).reset_index(drop=True)

    result = result.drop(columns=["업체명", "수량"], errors="ignore")

    for col in OUTPUT_COLUMNS:
        if col not in result.columns:
            result[col] = ""

    return result[OUTPUT_COLUMNS]


def build_total_row(df: pd.DataFrame) -> list[str]:
    total_row = []
    for col in OUTPUT_COLUMNS:
        if col == "날짜":
            total_row.append("합계")
        elif col in SUM_COLUMNS:
            total = to_int_series(df[col]).sum()
            total_row.append("" if total == 0 else str(int(total)))
        else:
            total_row.append("")
    return total_row


def build_total_row_for_columns(df: pd.DataFrame, columns: list[str]) -> list[str]:
    total_row = []
    for col in columns:
        if col == "날짜":
            total_row.append("합계")
        elif col in SUM_COLUMNS or col in DETAIL_SUM_COLUMNS:
            total = to_int_series(df[col]).sum()
            total_row.append("" if total == 0 else str(int(total)))
        else:
            total_row.append("")
    return total_row


def build_detail_df(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, row in df.iterrows():
        base = {
            "날짜": clean_text(row.get("날짜", "")),
            "브랜드": clean_text(row.get("브랜드", "")) or "브랜드 없음",
            "품번": clean_text(row.get("품번", "")),
            "컬러": "",
            "사이즈": clean_text(row.get("사이즈", "")),
        }

        for store_name in STORE_QTY_COLUMNS:
            qty = int(to_int_series(pd.Series([row.get(store_name, "")])).iloc[0])
            if qty <= 0:
                continue
            rows.append(
                {
                    **base,
                    "수량": str(qty),
                    "매장명": store_name,
                    "__block__": base["브랜드"],
                }
            )

        shortage_qty = int(to_int_series(pd.Series([row.get("부족 수량", "")])).iloc[0])
        if shortage_qty < 0:
            rows.append(
                {
                    **base,
                    "수량": str(abs(shortage_qty)),
                    "매장명": "재고 없음",
                    "__block__": "재고 없음",
                }
            )

    if not rows:
        return pd.DataFrame(columns=["브랜드", "__block__"] + DETAIL_COLUMNS)

    detail_df = pd.DataFrame(rows)
    detail_df = detail_df.sort_values(
        by=["__block__", "품번", "사이즈", "매장명"],
        ascending=[True, True, True, True],
        kind="stable",
    ).reset_index(drop=True)
    return detail_df


def build_brand_horizontal_values(df: pd.DataFrame) -> list[list[str]]:
    if df.empty:
        return [["데이터 없음"]]

    temp_df = build_detail_df(df)
    if temp_df.empty:
        return [["데이터 없음"]]

    normal_blocks = [
        block_name
        for block_name in temp_df["__block__"].drop_duplicates().tolist()
        if block_name != "재고 없음"
    ]
    blocks_order = normal_blocks + (["재고 없음"] if "재고 없음" in temp_df["__block__"].values else [])

    blocks = []
    for block_name in blocks_order:
        block_source_df = temp_df.loc[temp_df["__block__"].eq(block_name)].copy()
        if block_source_df.empty:
            continue
        block_df = block_source_df[BRAND_BLOCK_COLUMNS].fillna("").astype(str)
        title_row = [block_name] + [""] * (len(BRAND_BLOCK_COLUMNS) - 1)
        total_row = build_total_row_for_columns(block_df, BRAND_BLOCK_COLUMNS)
        header_row = BRAND_BLOCK_COLUMNS[:]
        blocks.append([title_row, total_row, header_row] + block_df.values.tolist())

    if not blocks:
        return [["데이터 없음"]]

    max_rows = max(len(block) for block in blocks)
    values = []
    for row_idx in range(max_rows):
        row_values = []
        for block_idx, block in enumerate(blocks):
            if block_idx > 0:
                row_values.append("")
            if row_idx < len(block):
                row_values.extend(block[row_idx])
            else:
                row_values.extend([""] * len(BRAND_BLOCK_COLUMNS))
        values.append(row_values)
    return values


def set_auto_width(ws, start_row: int, end_row: int, start_col: int, end_col: int):
    for col_idx in range(start_col, end_col + 1):
        max_len = 0
        for row_idx in range(start_row, end_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            text = "" if value is None else str(value)
            text = re.sub(r"\s+", " ", text).strip()
            max_len = max(max_len, len(text))
        width = max(MIN_COL_WIDTH, min(max_len + COL_WIDTH_PADDING, MAX_COL_WIDTH))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def save_to_excel(df: pd.DataFrame):
    ensure_dir(LIST_DIR)

    wb = Workbook()
    ws = wb.active
    ws.title = TARGET_SHEET_NAME
    ws.sheet_view.showGridLines = False

    align_center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    excel_rows = build_brand_horizontal_values(df)
    numeric_headers = set(DETAIL_SUM_COLUMNS)
    header_row_idx = EXCEL_START_ROW + 2
    header_values = excel_rows[2] if len(excel_rows) > 2 else []

    for row_offset, row_values in enumerate(excel_rows, start=EXCEL_START_ROW):
        for value_offset, value in enumerate(row_values, start=EXCEL_START_COL):
            header_pos = value_offset - EXCEL_START_COL
            header_name = header_values[header_pos] if header_pos < len(header_values) else ""
            is_separator_col = (
                len(BRAND_BLOCK_COLUMNS) > 0
                and header_pos % (len(BRAND_BLOCK_COLUMNS) + 1) == len(BRAND_BLOCK_COLUMNS)
            )
            if is_separator_col:
                continue
            cell_value = value
            if row_offset > EXCEL_START_ROW and header_name in numeric_headers:
                try:
                    cell_value = int(value) if clean_text(value) != "" else None
                except Exception:
                    cell_value = clean_text(value)
            else:
                cell_value = clean_text(value)

            cell = ws.cell(row=row_offset, column=value_offset, value=cell_value)
            cell.alignment = align_center
            if row_offset >= header_row_idx:
                cell.border = border

            if row_offset == EXCEL_START_ROW:
                cell.font = HEADER_FONT
            elif row_offset == header_row_idx:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT

            if header_name == "부족 수량" and row_offset != header_row_idx and clean_text(value) != "":
                cell.font = RED_FONT

    for row_idx in range(EXCEL_START_ROW, EXCEL_START_ROW + len(excel_rows)):
        ws.row_dimensions[row_idx].height = ROW_HEIGHT

    set_auto_width(
        ws,
        EXCEL_START_ROW,
        EXCEL_START_ROW + len(excel_rows) - 1,
        EXCEL_START_COL,
        EXCEL_START_COL + max(len(row) for row in excel_rows) - 1,
    )

    out_name = f"{EXCEL_BASENAME} {ymd()}.xlsx"
    out_path = safe_path(os.path.join(LIST_DIR, out_name))
    wb.save(out_path)
    log(f"엑셀 저장 완료: {out_path}")


def upload_to_google_sheet(df: pd.DataFrame):
    values = build_brand_horizontal_values(df)
    payload = {
        "spreadsheetId": SPREADSHEET_ID,
        "sheetName": TARGET_SHEET_NAME,
        "values": values,
        "append": False,
        "clear": True,
    }

    last_error = None
    log(f"업로드 시작 -> {TARGET_SHEET_NAME} ({len(df)}행)")
    for attempt in range(1, RETRIES + 1):
        try:
            response = requests.post(
                WEB_APP_URL,
                json=payload,
                headers={"User-Agent": "Mozilla/5.0"},
                timeout=UPLOAD_TIMEOUT,
            )
            response.raise_for_status()
            log(f"업로드 완료 -> {TARGET_SHEET_NAME}")
            log(f"응답: {(response.text or '').strip()}")
            return
        except Exception as e:
            last_error = e
            log(f"업로드 실패 ({attempt}/{RETRIES}) -> {e}")
            if attempt < RETRIES:
                time.sleep(RETRY_SLEEP_SEC)
    raise last_error


def main():
    log("출고 취합 업로드 시작")
    csv_texts = download_all_csv_texts()

    source_df = load_source_df(csv_texts["source"])
    office_df = load_office_df(csv_texts["office"])
    square_dfs = [load_store_df(csv_texts[f"store::스퀘어원::{idx}"], "스퀘어원") for idx in (1, 2)]
    guwol_dfs = [load_store_df(csv_texts[f"store::구월::{idx}"], "구월") for idx in (1, 2)]
    bucheon_dfs = [load_store_df(csv_texts[f"store::부천::{idx}"], "부천") for idx in (1, 2)]
    fila_df = load_store_df(csv_texts["store::휠라 파주::1"], "휠라 파주")
    puma_df = load_store_df(csv_texts["store::푸마 여주::1"], "푸마 여주")

    office_lookup = summarize_qty_by_code(office_df)
    square_lookup = summarize_store_with_fallback(square_dfs)
    guwol_lookup = summarize_store_with_fallback(guwol_dfs)
    bucheon_lookup = summarize_store_with_fallback(bucheon_dfs)
    fila_lookup = summarize_qty_by_code(fila_df)
    puma_lookup = summarize_qty_by_code(puma_df)

    result_df = attach_summary_columns(
        source_df,
        office_lookup=office_lookup,
        square_lookup=square_lookup,
        guwol_lookup=guwol_lookup,
        bucheon_lookup=bucheon_lookup,
        fila_lookup=fila_lookup,
        puma_lookup=puma_lookup,
    )
    upload_to_google_sheet(result_df)
    save_to_excel(result_df)
    log("작업 완료")


if __name__ == "__main__":
    main()
