# -*- coding: utf-8 -*-

from __future__ import annotations

from io import StringIO
from numbers import Number
from pathlib import Path
import re
import sys
import time
from datetime import date

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from office_erp_paths import LIST_DIR


PLATFORM_SUMMARY_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vRWhzq2rt_O-YnOwzWnE8L_d8--NBu-EMmDkxnnVy-oATAfCjDIX976b973bY2MiO8gYPTE5WCk-tVv/pub?gid=1407848473&single=true&output=csv"
BRAND_SUMMARY_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vRWhzq2rt_O-YnOwzWnE8L_d8--NBu-EMmDkxnnVy-oATAfCjDIX976b973bY2MiO8gYPTE5WCk-tVv/pub?gid=441477604&single=true&output=csv"
PLATFORM_SOURCE_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vRWhzq2rt_O-YnOwzWnE8L_d8--NBu-EMmDkxnnVy-oATAfCjDIX976b973bY2MiO8gYPTE5WCk-tVv/pub?gid=733480714&single=true&output=csv"
PLATFORM_RETURN_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vRWhzq2rt_O-YnOwzWnE8L_d8--NBu-EMmDkxnnVy-oATAfCjDIX976b973bY2MiO8gYPTE5WCk-tVv/pub?gid=1032857881&single=true&output=csv"
CURRENT_STOCK_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vQBC-ZAQixoxbta3FsLfG_f8qugvCGbjK2yrsKcLA6tJsi1ww5qDl-_21od9-55Lv0F9dmzlR5c1QIh/pub?gid=0&single=true&output=csv"

DOWNLOAD_TIMEOUT = (10, 60)
RETRIES = 3
RETRY_SLEEP_SEC = 2

PLATFORM_COLUMNS = ["날짜", "플랫폼", "주문건수", "판매 수량", "판매가", "수수료", "수입", "공급가", "마진", "마진율"]
BRAND_COLUMNS = ["날짜", "플랫폼", "브랜드", "수량", "판매가", "수입", "공급가", "마진", "마진율"]
NUMERIC_PLATFORM_COLUMNS = ["주문건수", "판매 수량", "판매가", "수수료", "수입", "공급가", "마진"]
NUMERIC_BRAND_COLUMNS = ["수량", "판매가", "수입", "공급가", "마진"]
RETURN_MONTH_COLUMNS = ["플랫폼", "반품 건수", "반품 수량", "반품 매출", "반품 정산", "반품 공급가", "반품 마진", "출고 수량", "반품율"]
NUMERIC_RETURN_MONTH_COLUMNS = ["반품 건수", "반품 수량", "반품 매출", "반품 정산", "반품 공급가", "반품 마진", "출고 수량"]
SOURCE_REQUIRED_COLUMNS = [
    "날짜",
    "플랫폼",
    "주문번호",
    "브랜드",
    "품번",
    "수량",
    "총 판매가",
    "총 수수료",
    "총 수입",
    "총 KRW",
    "총 공급가",
    "마진",
    "환율",
    "내역",
    "매장명",
]
RETURN_STATUS_VALUES = {"반품"}
TOP_BLOCK_HEADERS = ["순위", "품번", "판매 수량", "현재고", "매출", "마진", "마진율"]
TOP_BLOCKS_PER_ROW = 2
TOP_BLOCK_SPACER_COLS = 1
PLATFORM_SORT_ORDER = [
    "POIZON",
    "POIZON 보관",
    "KASHION (TM)",
    "KASHION (JD)",
    "KREAM",
    "KREAM 보관",
    "브랜더",
    "풀메이커",
]


def log(message: str):
    print(f"[온라인 월마감] {message}", flush=True)


def make_session() -> requests.Session:
    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0"})
    return session


def fetch_csv_text(session: requests.Session, url: str) -> str:
    last_error = None
    for attempt in range(1, RETRIES + 1):
        try:
            response = session.get(url, timeout=DOWNLOAD_TIMEOUT)
            response.raise_for_status()
            response.encoding = "utf-8"
            return response.text
        except Exception as exc:
            last_error = exc
            if attempt < RETRIES:
                time.sleep(RETRY_SLEEP_SEC)
    raise last_error


def read_csv_from_url(url: str) -> pd.DataFrame:
    return pd.read_csv(StringIO(fetch_csv_text(make_session(), url)), dtype=str).fillna("")


def read_required_source_csv(url: str, label: str) -> pd.DataFrame:
    df = read_csv_from_url(url)
    missing_columns = [col for col in SOURCE_REQUIRED_COLUMNS if col not in df.columns]
    if missing_columns:
        raise ValueError(f"{label} 필수 열이 없습니다: {missing_columns}")
    return df[SOURCE_REQUIRED_COLUMNS].copy()


def read_platform_source_df() -> pd.DataFrame:
    return read_required_source_csv(PLATFORM_SOURCE_URL, "플랫폼 원본 시트")


def read_platform_return_df() -> pd.DataFrame:
    return_df = read_required_source_csv(PLATFORM_RETURN_URL, "플랫폼 반품 시트")
    return_df["내역"] = return_df["내역"].apply(clean_text)
    return return_df[return_df["내역"].isin(RETURN_STATUS_VALUES)].copy()


def read_combined_platform_source_df() -> pd.DataFrame:
    source_df = read_platform_source_df()
    return_df = read_platform_return_df()
    if return_df.empty:
        log("플랫폼 반품 시트 계산 대상 없음 -> 내역이 '반품'인 행이 없습니다.")
        return source_df

    log(f"플랫폼 반품 시트 포함 -> 내역 '반품' {len(return_df)}행")
    return pd.concat([source_df, return_df], ignore_index=True)


def clean_text(value) -> str:
    if pd.isna(value):
        return ""
    return " ".join(str(value).strip().split())


def to_number_series(series: pd.Series) -> pd.Series:
    cleaned = (
        series.fillna("")
        .astype(str)
        .str.replace(",", "", regex=False)
        .str.replace('"', "", regex=False)
        .str.replace("%", "", regex=False)
        .str.strip()
    )
    return pd.to_numeric(cleaned, errors="coerce").fillna(0)


def format_int(value) -> int:
    return int(round(float(value)))


def safe_rate(numerator, denominator) -> float:
    denominator = float(denominator or 0)
    if denominator == 0:
        return 0.0
    return float(numerator or 0) / denominator


def display_label(value) -> str:
    text = clean_text(value)
    replacements = {
        "KASHION (TM)": "KASHION\n(TM)",
        "KASHION (JD)": "KASHION\n(JD)",
        "POIZON 보관": "POIZON\n보관",
        "KREAM 보관": "KREAM\n보관",
        "THE NORTH FACE (DC)": "THE NORTH\nFACE (DC)",
        "THE NORTH FACE": "THE NORTH\nFACE",
    }
    return replacements.get(text, text)


def platform_sort_key(value) -> tuple[int, str]:
    platform = clean_text(value)
    try:
        return (PLATFORM_SORT_ORDER.index(platform), platform)
    except ValueError:
        return (len(PLATFORM_SORT_ORDER), platform)


def parse_korean_date_key(value: str) -> tuple[int, int]:
    text = clean_text(value).replace(" ", "")
    try:
        if "월" in text and "일" in text:
            month = int(text.split("월", 1)[0])
            day = int(text.split("월", 1)[1].split("일", 1)[0])
            return (month, day)
        parsed_date = pd.to_datetime(text, errors="coerce")
        if not pd.isna(parsed_date):
            return (int(parsed_date.month), int(parsed_date.day))
    except Exception:
        pass
    return (-1, -1)


def parse_korean_month(value: str) -> int:
    text = clean_text(value).replace(" ", "")
    try:
        if "월" in text:
            return int(text.split("월", 1)[0])
        parsed_date = pd.to_datetime(text, errors="coerce")
        if not pd.isna(parsed_date):
            return int(parsed_date.month)
    except Exception:
        pass
    return -1


def parse_target_month(argv: list[str], platform_df: pd.DataFrame) -> int:
    if len(argv) >= 2:
        month_text = clean_text(argv[1]).replace("월", "")
        if month_text.isdigit() and 1 <= int(month_text) <= 12:
            return int(month_text)
        raise ValueError("월 지정은 1~12 숫자 또는 '4월' 형식으로 입력해주세요.")

    today = date.today()
    return 12 if today.month == 1 else today.month - 1


def normalize_best_platform(platform_value, order_no_value="") -> str:
    platform = clean_text(platform_value).upper()
    order_no = clean_text(order_no_value).upper()
    if platform == "KASHION":
        if order_no.startswith("LP"):
            return "KASHION (TM)"
        if order_no.startswith("JD"):
            return "KASHION (JD)"
        return "KASHION"
    if platform in {"TM", "天猫", "KASHION (TM)"}:
        return "KASHION (TM)"
    if platform in {"JD", "京东", "KASHION (JD)"}:
        return "KASHION (JD)"
    return clean_text(platform_value)


def select_margin_base_series(sales_series: pd.Series, krw_series: pd.Series) -> pd.Series:
    sales_num = to_number_series(sales_series)
    krw_num = to_number_series(krw_series)
    foreign_currency_mask = krw_num.gt(0) & sales_num.gt(0) & krw_num.gt(sales_num * 10)
    return sales_num.where(~foreign_currency_mask, krw_num)


def filter_valid_outbound_rows(df: pd.DataFrame) -> pd.DataFrame:
    output_df = df.copy()
    if "내역" in output_df.columns:
        output_df["내역"] = output_df["내역"].apply(clean_text)
        excluded_statuses = {"取消件", "调货中", "재고 없음", "재고없음", "已处理", "취소", "취소건"}
        output_df = output_df[~output_df["내역"].isin(excluded_statuses)].copy()
    if "매장명" in output_df.columns:
        output_df["매장명"] = output_df["매장명"].apply(clean_text)
        output_df = output_df[~output_df["매장명"].eq("재고없음")].copy()
    return output_df.reset_index(drop=True)


def normalize_platform_summary_df(df: pd.DataFrame) -> pd.DataFrame:
    missing_columns = [col for col in PLATFORM_COLUMNS if col not in df.columns]
    if missing_columns:
        raise ValueError(f"플랫폼 마감 시트 필수 열이 없습니다: {missing_columns}")

    output_df = df[PLATFORM_COLUMNS].copy()
    output_df["날짜"] = output_df["날짜"].apply(clean_text)
    output_df["플랫폼"] = output_df["플랫폼"].apply(clean_text)
    output_df = output_df.drop_duplicates(subset=["날짜", "플랫폼"], keep="last").reset_index(drop=True)
    for col in NUMERIC_PLATFORM_COLUMNS:
        output_df[col] = to_number_series(output_df[col])
    output_df["마진율"] = output_df["마진"].div(output_df["판매가"].replace(0, pd.NA)).fillna(0)
    return output_df


def normalize_brand_summary_df(df: pd.DataFrame) -> pd.DataFrame:
    renamed_df = df.rename(columns={"공금가": "공급가"}).copy()
    missing_columns = [col for col in BRAND_COLUMNS if col not in renamed_df.columns]
    if missing_columns:
        raise ValueError(f"브랜드 마감 시트 필수 열이 없습니다: {missing_columns}")

    output_df = renamed_df[BRAND_COLUMNS].copy()
    output_df["날짜"] = output_df["날짜"].apply(clean_text)
    output_df["플랫폼"] = output_df["플랫폼"].apply(clean_text)
    output_df["브랜드"] = output_df["브랜드"].apply(clean_text)
    output_df = output_df.drop_duplicates(subset=["날짜", "플랫폼", "브랜드"], keep="last").reset_index(drop=True)
    for col in NUMERIC_BRAND_COLUMNS:
        output_df[col] = to_number_series(output_df[col])
    output_df["마진율"] = output_df["마진"].div(output_df["판매가"].replace(0, pd.NA)).fillna(0)
    return output_df


def build_stock_lookup(df: pd.DataFrame) -> dict[str, int]:
    missing_columns = [col for col in ["품번", "현재고"] if col not in df.columns]
    if missing_columns:
        raise ValueError(f"현재고 시트 필수 열이 없습니다: {missing_columns}")
    output_df = df[["품번", "현재고"]].copy()
    output_df["품번"] = output_df["품번"].apply(clean_text)
    output_df = output_df[output_df["품번"] != ""].copy()
    output_df["현재고_num"] = to_number_series(output_df["현재고"])
    grouped_df = output_df.groupby("품번", dropna=False, sort=False)["현재고_num"].sum().reset_index()
    return {row["품번"]: format_int(row["현재고_num"]) for _, row in grouped_df.iterrows()}


def filter_month_df(df: pd.DataFrame, target_month: int) -> pd.DataFrame:
    return df[df["날짜"].apply(parse_korean_month).eq(target_month)].copy().reset_index(drop=True)


def build_source_sales_df(df: pd.DataFrame, target_month: int) -> pd.DataFrame:
    required_columns = [
        "날짜",
        "플랫폼",
        "주문번호",
        "브랜드",
        "품번",
        "수량",
        "총 판매가",
        "총 수수료",
        "총 수입",
        "환율",
        "총 KRW",
        "총 공급가",
        "마진",
        "내역",
        "매장명",
    ]
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"원본 플랫폼 시트 필수 열이 없습니다: {missing_columns}")

    output_df = df[required_columns].copy()
    output_df["날짜"] = output_df["날짜"].apply(clean_text)
    output_df = output_df[output_df["날짜"].apply(parse_korean_month).eq(target_month)].copy()
    output_df["주문번호"] = output_df["주문번호"].apply(clean_text)
    output_df["플랫폼"] = output_df.apply(lambda row: normalize_best_platform(row["플랫폼"], row["주문번호"]), axis=1)
    output_df["브랜드"] = output_df["브랜드"].apply(clean_text)
    output_df["품번"] = output_df["품번"].apply(clean_text)
    output_df = filter_valid_outbound_rows(output_df)
    output_df = output_df[output_df["날짜"].ne("") & output_df["플랫폼"].ne("")].copy()
    if output_df.empty:
        return output_df

    output_df["수량"] = to_number_series(output_df["수량"])
    sales_num = to_number_series(output_df["총 판매가"])
    fee_num = to_number_series(output_df["총 수수료"])
    income_num = to_number_series(output_df["총 수입"])
    krw_num = to_number_series(output_df["총 KRW"])
    exchange_num = to_number_series(output_df["환율"])
    supply_num = to_number_series(output_df["총 공급가"])
    margin_num = to_number_series(output_df["마진"])

    inferred_exchange = krw_num.div(income_num.replace(0, pd.NA)).fillna(0)
    exchange_rate = inferred_exchange.where(inferred_exchange.gt(0), exchange_num)
    foreign_currency_mask = krw_num.gt(0) & income_num.gt(0) & krw_num.gt(income_num * 10) & sales_num.gt(0)
    buyma_settlement_mask = output_df["플랫폼"].str.upper().eq("BUYMA") & krw_num.gt(0)

    output_df["매출"] = sales_num.where(~foreign_currency_mask, sales_num * exchange_rate)
    output_df["정산"] = income_num.where(~(foreign_currency_mask | buyma_settlement_mask), krw_num)
    output_df["수수료"] = fee_num.where(~foreign_currency_mask, fee_num * exchange_rate)
    output_df["공급가"] = supply_num
    output_df["마진"] = margin_num
    return_mask = output_df["내역"].isin(RETURN_STATUS_VALUES)
    for col in ["수량", "매출", "정산", "수수료", "공급가", "마진"]:
        output_df.loc[return_mask, col] = -output_df.loc[return_mask, col].abs()

    blank_order_mask = output_df["주문번호"].eq("")
    output_df["_주문키"] = output_df["주문번호"].where(~blank_order_mask, "ROW-" + output_df.index.astype(str))
    return output_df.reset_index(drop=True)


def count_orders(series: pd.Series) -> int:
    return int(series.nunique())


def build_platform_month_df(platform_df: pd.DataFrame) -> pd.DataFrame:
    platform_df = platform_df[platform_df["플랫폼"].ne("")].copy()
    grouped_df = (
        platform_df.groupby("플랫폼", dropna=False, sort=False)
        .agg(
            주문건수=("_주문키", count_orders),
            판매_수량=("수량", "sum"),
            매출=("매출", "sum"),
            수수료=("수수료", "sum"),
            정산=("정산", "sum"),
            공급가=("공급가", "sum"),
            마진=("마진", "sum"),
        )
        .reset_index()
        .rename(columns={"판매_수량": "판매 수량"})
    )
    total_sales = grouped_df["매출"].sum()
    total_margin = grouped_df["마진"].sum()
    grouped_df["마진율"] = grouped_df["마진"].div(grouped_df["매출"].replace(0, pd.NA)).fillna(0)
    grouped_df["매출 비중"] = grouped_df["매출"].apply(lambda value: safe_rate(value, total_sales))
    grouped_df["마진 비중"] = grouped_df["마진"].apply(lambda value: safe_rate(value, total_margin))
    for col in ["주문건수", "판매 수량", "매출", "수수료", "정산", "공급가", "마진"]:
        grouped_df[col] = grouped_df[col].apply(format_int)
    grouped_df["_platform_key"] = grouped_df["플랫폼"].apply(platform_sort_key)
    return grouped_df.sort_values(by=["_platform_key", "플랫폼"]).drop(columns=["_platform_key"]).reset_index(drop=True)


def build_brand_month_df(brand_df: pd.DataFrame) -> pd.DataFrame:
    brand_df = brand_df[brand_df["플랫폼"].ne("") & brand_df["브랜드"].ne("")].copy()
    grouped_df = (
        brand_df.groupby(["플랫폼", "브랜드"], dropna=False, sort=False)
        .agg(
            수량=("수량", "sum"),
            매출=("매출", "sum"),
            정산=("정산", "sum"),
            공급가=("공급가", "sum"),
            마진=("마진", "sum"),
        )
        .reset_index()
    )
    grouped_df["마진율"] = grouped_df["마진"].div(grouped_df["매출"].replace(0, pd.NA)).fillna(0)
    grouped_df["플랫폼 내 매출 비중"] = grouped_df.groupby("플랫폼")["매출"].transform(lambda values: values / values.sum() if values.sum() else 0)
    grouped_df["플랫폼 내 마진 비중"] = grouped_df.groupby("플랫폼")["마진"].transform(lambda values: values / values.sum() if values.sum() else 0)
    for col in ["수량", "매출", "정산", "공급가", "마진"]:
        grouped_df[col] = grouped_df[col].apply(format_int)
    grouped_df["_platform_key"] = grouped_df["플랫폼"].apply(platform_sort_key)
    return grouped_df.sort_values(by=["_platform_key", "매출", "마진", "브랜드"], ascending=[True, False, False, True]).drop(columns=["_platform_key"]).reset_index(drop=True)


def build_return_month_df(source_sales_df: pd.DataFrame) -> pd.DataFrame:
    if source_sales_df.empty:
        return pd.DataFrame(columns=RETURN_MONTH_COLUMNS)

    return_df = source_sales_df[
        source_sales_df["내역"].isin(RETURN_STATUS_VALUES)
        & source_sales_df["플랫폼"].ne("")
    ].copy()
    if return_df.empty:
        return pd.DataFrame(columns=RETURN_MONTH_COLUMNS)

    outbound_df = source_sales_df[
        ~source_sales_df["내역"].isin(RETURN_STATUS_VALUES)
        & source_sales_df["플랫폼"].ne("")
    ].copy()
    outbound_qty_df = (
        outbound_df.groupby("플랫폼", dropna=False, sort=False)
        .agg(출고_수량=("수량", "sum"))
        .reset_index()
    )

    grouped_df = (
        return_df.groupby("플랫폼", dropna=False, sort=False)
        .agg(
            반품_건수=("수량", "size"),
            반품_수량=("수량", lambda series: series.abs().sum()),
            반품_매출=("매출", lambda series: series.abs().sum()),
            반품_정산=("정산", lambda series: series.abs().sum()),
            반품_공급가=("공급가", lambda series: series.abs().sum()),
            반품_마진=("마진", lambda series: series.abs().sum()),
        )
        .reset_index()
        .merge(outbound_qty_df, on="플랫폼", how="left")
        .fillna({"출고_수량": 0})
        .rename(
            columns={
                "반품_건수": "반품 건수",
                "반품_수량": "반품 수량",
                "반품_매출": "반품 매출",
                "반품_정산": "반품 정산",
                "반품_공급가": "반품 공급가",
                "반품_마진": "반품 마진",
                "출고_수량": "출고 수량",
            }
        )
    )
    grouped_df["반품율"] = grouped_df["반품 수량"].div(grouped_df["출고 수량"].replace(0, pd.NA)).fillna(0)
    for col in NUMERIC_RETURN_MONTH_COLUMNS:
        grouped_df[col] = grouped_df[col].apply(format_int)
    grouped_df["_platform_key"] = grouped_df["플랫폼"].apply(platform_sort_key)
    return grouped_df.sort_values(by=["_platform_key", "플랫폼"]).drop(columns=["_platform_key"])[RETURN_MONTH_COLUMNS].reset_index(drop=True)


def build_daily_trend_df(platform_df: pd.DataFrame) -> pd.DataFrame:
    grouped_df = (
        platform_df.groupby("날짜", dropna=False, sort=False)
        .agg(
            주문건수=("_주문키", count_orders),
            판매_수량=("수량", "sum"),
            매출=("매출", "sum"),
            정산=("정산", "sum"),
            마진=("마진", "sum"),
        )
        .reset_index()
        .rename(columns={"판매_수량": "판매 수량"})
    )
    grouped_df["마진율"] = grouped_df["마진"].div(grouped_df["매출"].replace(0, pd.NA)).fillna(0)
    grouped_df["_date_key"] = grouped_df["날짜"].apply(parse_korean_date_key)
    for col in ["주문건수", "판매 수량", "매출", "정산", "마진"]:
        grouped_df[col] = grouped_df[col].apply(format_int)
    return grouped_df.sort_values("_date_key").drop(columns=["_date_key"]).reset_index(drop=True)


def previous_month(target_month: int) -> int:
    return 12 if target_month == 1 else target_month - 1


def build_month_summary(source_sales_df: pd.DataFrame) -> dict[str, float]:
    if source_sales_df.empty:
        return {}
    return {
        "매출": float(source_sales_df["매출"].sum()),
        "정산": float(source_sales_df["정산"].sum()),
        "마진": float(source_sales_df["마진"].sum()),
    }


def build_top10_grid_df(name_order: list[str], ranked_map: dict[str, pd.DataFrame], stock_lookup: dict[str, int], column_prefix: str) -> pd.DataFrame:
    if not name_order:
        return pd.DataFrame()

    block_size = len(TOP_BLOCK_HEADERS)
    row_block_width = (TOP_BLOCKS_PER_ROW * block_size) + ((TOP_BLOCKS_PER_ROW - 1) * TOP_BLOCK_SPACER_COLS)
    rows = []
    for chunk_start in range(0, len(name_order), TOP_BLOCKS_PER_ROW):
        chunk_names = name_order[chunk_start:chunk_start + TOP_BLOCKS_PER_ROW]
        title_row = []
        header_row = []
        for idx, block_name in enumerate(chunk_names):
            if idx > 0:
                title_row.extend([""] * TOP_BLOCK_SPACER_COLS)
                header_row.extend([""] * TOP_BLOCK_SPACER_COLS)
            title_row.extend([block_name] + [""] * (block_size - 1))
            header_row.extend(TOP_BLOCK_HEADERS)
        title_row.extend([""] * (row_block_width - len(title_row)))
        header_row.extend([""] * (row_block_width - len(header_row)))
        rows.append(title_row)
        rows.append(header_row)

        for rank_idx in range(10):
            row_values = []
            for idx, block_name in enumerate(chunk_names):
                if idx > 0:
                    row_values.extend([""] * TOP_BLOCK_SPACER_COLS)
                block_df = ranked_map[block_name]
                if rank_idx < len(block_df):
                    row = block_df.iloc[rank_idx]
                    row_values.extend(
                        [
                            rank_idx + 1,
                            row["품번"],
                            format_int(row["총_수량"]),
                            stock_lookup.get(clean_text(row["품번"]), 0),
                            format_int(row["총_매출"]),
                            format_int(row["총_마진"]),
                            float(row["마진율"]),
                        ]
                    )
                else:
                    row_values.extend(["", "", "", "", "", "", ""])
            row_values.extend([""] * (row_block_width - len(row_values)))
            rows.append(row_values)
        if chunk_start + TOP_BLOCKS_PER_ROW < len(name_order):
            rows.append([""] * row_block_width)

    return pd.DataFrame(rows, columns=[f"{column_prefix}_{idx}" for idx in range(1, row_block_width + 1)])


def build_source_top10_df(df: pd.DataFrame, stock_lookup: dict[str, int], target_month: int, group_column: str, column_prefix: str) -> pd.DataFrame:
    required_columns = ["날짜", group_column, "품번", "수량", "총 판매가", "총 수입", "총 KRW", "마진", "내역", "매장명"]
    if group_column == "플랫폼":
        required_columns.append("주문번호")
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"원본 플랫폼 시트 필수 열이 없습니다: {missing_columns}")

    output_df = df[required_columns].copy()
    output_df["날짜"] = output_df["날짜"].apply(clean_text)
    output_df = output_df[output_df["날짜"].apply(parse_korean_month).eq(target_month)].copy()
    if group_column == "플랫폼":
        output_df["플랫폼"] = output_df.apply(lambda row: normalize_best_platform(row["플랫폼"], row["주문번호"]), axis=1)
    else:
        output_df[group_column] = output_df[group_column].apply(clean_text)
    output_df["품번"] = output_df["품번"].apply(clean_text)
    output_df = filter_valid_outbound_rows(output_df)
    output_df = output_df[~output_df["내역"].isin(RETURN_STATUS_VALUES)].copy()
    output_df = output_df[output_df[group_column].ne("") & output_df["품번"].ne("")].copy()
    if output_df.empty:
        return pd.DataFrame()

    output_df["수량_num"] = to_number_series(output_df["수량"])
    sales_num = to_number_series(output_df["총 판매가"])
    income_num = to_number_series(output_df["총 수입"]) if "총 수입" in output_df.columns else pd.Series([0] * len(output_df), index=output_df.index)
    krw_num = to_number_series(output_df["총 KRW"])
    exchange_rate = krw_num.div(income_num.replace(0, pd.NA)).fillna(0)
    foreign_currency_mask = krw_num.gt(0) & income_num.gt(0) & krw_num.gt(income_num * 10) & sales_num.gt(0)
    output_df["매출_num"] = sales_num.where(~foreign_currency_mask, sales_num * exchange_rate)
    output_df["마진_num"] = to_number_series(output_df["마진"])
    output_df["마진율_분모_num"] = output_df["매출_num"]

    grouped_df = (
        output_df.groupby([group_column, "품번"], dropna=False, sort=False)
        .agg(
            총_수량=("수량_num", "sum"),
            총_매출=("매출_num", "sum"),
            총_마진=("마진_num", "sum"),
            총_마진율_분모=("마진율_분모_num", "sum"),
        )
        .reset_index()
    )
    if grouped_df.empty:
        return pd.DataFrame()

    name_order = list(dict.fromkeys(grouped_df[group_column].tolist()))
    ranked_map = {}
    for name in name_order:
        rank_df = grouped_df[grouped_df[group_column].eq(name)].copy()
        rank_df = rank_df.sort_values(by=["총_수량", "총_매출", "총_마진", "품번"], ascending=[False, False, False, True]).head(10).reset_index(drop=True)
        rank_df["마진율"] = rank_df["총_마진"].div(rank_df["총_마진율_분모"].replace(0, pd.NA)).fillna(0)
        ranked_map[name] = rank_df
    return build_top10_grid_df(name_order, ranked_map, stock_lookup, column_prefix)


def build_source_margin_top10_df(df: pd.DataFrame, stock_lookup: dict[str, int], target_month: int, group_column: str, column_prefix: str) -> pd.DataFrame:
    required_columns = ["날짜", group_column, "품번", "수량", "총 판매가", "총 수입", "총 KRW", "마진", "내역", "매장명"]
    if group_column == "플랫폼":
        required_columns.append("주문번호")
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"원본 플랫폼 시트 필수 열이 없습니다: {missing_columns}")

    output_df = df[required_columns].copy()
    output_df["날짜"] = output_df["날짜"].apply(clean_text)
    output_df = output_df[output_df["날짜"].apply(parse_korean_month).eq(target_month)].copy()
    if group_column == "플랫폼":
        output_df["플랫폼"] = output_df.apply(lambda row: normalize_best_platform(row["플랫폼"], row["주문번호"]), axis=1)
    else:
        output_df[group_column] = output_df[group_column].apply(clean_text)
    output_df["품번"] = output_df["품번"].apply(clean_text)
    output_df = filter_valid_outbound_rows(output_df)
    output_df = output_df[~output_df["내역"].isin(RETURN_STATUS_VALUES)].copy()
    output_df = output_df[output_df[group_column].ne("") & output_df["품번"].ne("")].copy()
    if output_df.empty:
        return pd.DataFrame()

    output_df["수량_num"] = to_number_series(output_df["수량"])
    sales_num = to_number_series(output_df["총 판매가"])
    income_num = to_number_series(output_df["총 수입"]) if "총 수입" in output_df.columns else pd.Series([0] * len(output_df), index=output_df.index)
    krw_num = to_number_series(output_df["총 KRW"])
    exchange_rate = krw_num.div(income_num.replace(0, pd.NA)).fillna(0)
    foreign_currency_mask = krw_num.gt(0) & income_num.gt(0) & krw_num.gt(income_num * 10) & sales_num.gt(0)
    output_df["매출_num"] = sales_num.where(~foreign_currency_mask, sales_num * exchange_rate)
    output_df["마진_num"] = to_number_series(output_df["마진"])
    output_df["마진율_분모_num"] = output_df["매출_num"]

    grouped_df = (
        output_df.groupby([group_column, "품번"], dropna=False, sort=False)
        .agg(
            총_수량=("수량_num", "sum"),
            총_매출=("매출_num", "sum"),
            총_마진=("마진_num", "sum"),
            총_마진율_분모=("마진율_분모_num", "sum"),
        )
        .reset_index()
    )
    if grouped_df.empty:
        return pd.DataFrame()

    name_order = list(dict.fromkeys(grouped_df[group_column].tolist()))
    ranked_map = {}
    for name in name_order:
        rank_df = grouped_df[grouped_df[group_column].eq(name)].copy()
        rank_df["마진율"] = rank_df["총_마진"].div(rank_df["총_마진율_분모"].replace(0, pd.NA)).fillna(0)
        rank_df = rank_df.sort_values(by=["총_마진", "총_매출", "총_수량", "마진율", "품번"], ascending=[False, False, False, False, True]).head(10).reset_index(drop=True)
        ranked_map[name] = rank_df
    return build_top10_grid_df(name_order, ranked_map, stock_lookup, column_prefix)


def build_dashboard_df(
    target_month: int,
    platform_month_df: pd.DataFrame,
    brand_month_df: pd.DataFrame,
    daily_trend_df: pd.DataFrame,
    previous_month_summary: dict[str, float] | None = None,
    return_month_df: pd.DataFrame | None = None,
) -> pd.DataFrame:
    total_qty = platform_month_df["판매 수량"].sum()
    total_sales = platform_month_df["매출"].sum()
    total_income = platform_month_df["정산"].sum()
    total_margin = platform_month_df["마진"].sum()
    total_margin_rate = safe_rate(total_margin, total_sales)
    previous_month_summary = previous_month_summary or {}

    def previous_rate(current_value: float, previous_key: str):
        previous_value = previous_month_summary.get(previous_key, 0)
        if not previous_value:
            return ""
        return safe_rate(current_value - previous_value, previous_value)

    sales_vs_previous = previous_rate(total_sales, "매출")
    income_vs_previous = previous_rate(total_income, "정산")
    margin_vs_previous = previous_rate(total_margin, "마진")

    rows = [
        [f"{target_month}월 월마감", "", "", "", "", "", "", ""],
        ["월간 KPI", "", "", "", "", "전월", "", ""],
        ["매출", "정산", "마진", "마진율", "판매 수량", "매출", "정산", "마진"],
        [total_sales, total_income, total_margin, total_margin_rate, total_qty, sales_vs_previous, income_vs_previous, margin_vs_previous],
        ["", "", "", "", "", "", "", ""],
        ["플랫폼 요약", "", "", "", "", "", "", ""],
        ["플랫폼", "매출", "정산", "마진", "마진율", "판매 수량", "매출 비중", "마진 비중"],
    ]
    for _, row in platform_month_df.iterrows():
        rows.append([row["플랫폼"], row["매출"], row["정산"], row["마진"], row["마진율"], row["판매 수량"], row["매출 비중"], row["마진 비중"]])

    return_month_df = return_month_df if return_month_df is not None else pd.DataFrame(columns=RETURN_MONTH_COLUMNS)
    rows.extend([
        ["", "", "", "", "", "", "", ""],
        ["반품 요약", "", "", "", "", "", "", ""],
        ["플랫폼", "반품 건수", "반품 수량", "반품 매출", "반품 정산", "반품 마진", "출고 수량", "반품율"],
    ])
    if return_month_df.empty:
        rows.append(["반품 데이터 없음", "", "", "", "", "", "", ""])
    else:
        for _, return_row in return_month_df.iterrows():
            rows.append(
                [
                    return_row["플랫폼"],
                    return_row["반품 건수"],
                    return_row["반품 수량"],
                    return_row["반품 매출"],
                    return_row["반품 정산"],
                    return_row["반품 마진"],
                    return_row["출고 수량"],
                    return_row["반품율"],
                ]
            )

    brand_total_df = (
        brand_month_df.groupby("브랜드", dropna=False, sort=False)
        .agg(
            수량=("수량", "sum"),
            매출=("매출", "sum"),
            정산=("정산", "sum"),
            마진=("마진", "sum"),
        )
        .reset_index()
    )
    brand_total_sales = brand_total_df["매출"].sum()
    brand_total_margin = brand_total_df["마진"].sum()
    brand_total_df["마진율"] = brand_total_df["마진"].div(brand_total_df["매출"].replace(0, pd.NA)).fillna(0)
    brand_total_df["매출 비중"] = brand_total_df["매출"].apply(lambda value: safe_rate(value, brand_total_sales))
    brand_total_df["마진 비중"] = brand_total_df["마진"].apply(lambda value: safe_rate(value, brand_total_margin))
    brand_total_df = brand_total_df.sort_values(by=["매출", "마진", "브랜드"], ascending=[False, False, True]).reset_index(drop=True)

    rows.extend([
        ["", "", "", "", "", "", "", ""],
        ["브랜드 전체 요약", "", "", "", "", "", "", ""],
        ["브랜드", "매출", "정산", "마진", "마진율", "수량", "매출 비중", "마진 비중"],
    ])
    for _, brand_row in brand_total_df.iterrows():
        rows.append(
            [
                brand_row["브랜드"],
                format_int(brand_row["매출"]),
                format_int(brand_row["정산"]),
                format_int(brand_row["마진"]),
                brand_row["마진율"],
                format_int(brand_row["수량"]),
                brand_row["매출 비중"],
                brand_row["마진 비중"],
            ]
        )

    rows.extend([
        ["", "", "", "", "", "", "", ""],
        ["플랫폼별 브랜드 상세", "", "", "", "", "", "", ""],
    ])

    platform_order = platform_month_df["플랫폼"].tolist()
    for platform_name in platform_order:
        platform_brand_df = brand_month_df[brand_month_df["플랫폼"].eq(platform_name)].copy()
        if platform_brand_df.empty:
            continue
        platform_brand_df = platform_brand_df.sort_values(by=["매출", "마진", "브랜드"], ascending=[False, False, True]).reset_index(drop=True)
        rows.append([f"{platform_name} 브랜드", "", "", "", "", "", "", ""])
        rows.append(["브랜드", "매출", "정산", "마진", "마진율", "수량", "매출 비중", "마진 비중"])
        for _, brand_row in platform_brand_df.iterrows():
            rows.append(
                [
                    brand_row["브랜드"],
                    brand_row["매출"],
                    brand_row["정산"],
                    brand_row["마진"],
                    brand_row["마진율"],
                    brand_row["수량"],
                    brand_row["플랫폼 내 매출 비중"],
                    brand_row["플랫폼 내 마진 비중"],
                ]
            )
        rows.append(["", "", "", "", "", "", "", ""])

    rows = [row + [""] * (8 - len(row)) for row in rows]
    return pd.DataFrame(rows, columns=[f"col_{idx}" for idx in range(1, 9)])


def build_presentation_dashboard_df(
    target_month: int,
    platform_month_df: pd.DataFrame,
    brand_month_df: pd.DataFrame,
    daily_trend_df: pd.DataFrame,
) -> pd.DataFrame:
    def make_top_with_other(df: pd.DataFrame, name_col: str, top_n: int = 5) -> pd.DataFrame:
        if df.empty:
            return pd.DataFrame(columns=[name_col, "매출", "마진", "마진율", "매출 비중", "마진 비중"])
        work_df = df[[name_col, "매출", "마진"]].copy()
        work_df = work_df.groupby(name_col, dropna=False, sort=False).agg(매출=("매출", "sum"), 마진=("마진", "sum")).reset_index()
        work_df = work_df.sort_values(by=["매출", "마진", name_col], ascending=[False, False, True]).reset_index(drop=True)
        top_df = work_df.head(top_n).copy()
        other_df = work_df.iloc[top_n:].copy()
        if not other_df.empty:
            top_df = pd.concat(
                [
                    top_df,
                    pd.DataFrame([{name_col: "기타", "매출": other_df["매출"].sum(), "마진": other_df["마진"].sum()}]),
                ],
                ignore_index=True,
            )
        total_sales_local = top_df["매출"].sum()
        total_margin_local = top_df["마진"].sum()
        top_df["마진율"] = top_df["마진"].div(top_df["매출"].replace(0, pd.NA)).fillna(0)
        top_df["매출 비중"] = top_df["매출"].apply(lambda value: safe_rate(value, total_sales_local))
        top_df["마진 비중"] = top_df["마진"].apply(lambda value: safe_rate(value, total_margin_local))
        return top_df[[name_col, "매출", "마진", "마진율", "매출 비중", "마진 비중"]]

    total_sales = int(platform_month_df["매출"].sum())
    total_income = int(platform_month_df["정산"].sum())
    total_margin = int(platform_month_df["마진"].sum())
    total_qty = int(platform_month_df["판매 수량"].sum())
    total_orders = int(platform_month_df["주문건수"].sum())
    margin_rate = safe_rate(total_margin, total_sales)
    average_order = int(round(safe_rate(total_sales, total_orders), 0))
    margin_per_order = int(round(safe_rate(total_margin, total_orders), 0))

    best_platform = platform_month_df.sort_values(by=["매출", "마진"], ascending=[False, False]).head(1)
    best_platform_name = best_platform.iloc[0]["플랫폼"] if not best_platform.empty else ""
    best_platform_share = best_platform.iloc[0]["매출 비중"] if not best_platform.empty else 0
    best_brand_df = (
        brand_month_df.groupby("브랜드", dropna=False, sort=False)
        .agg(수량=("수량", "sum"), 매출=("매출", "sum"), 정산=("정산", "sum"), 마진=("마진", "sum"))
        .reset_index()
    )
    best_brand_df["마진율"] = best_brand_df["마진"].div(best_brand_df["매출"].replace(0, pd.NA)).fillna(0)
    total_brand_sales = best_brand_df["매출"].sum()
    total_brand_margin = best_brand_df["마진"].sum()
    best_brand_df["매출 비중"] = best_brand_df["매출"].apply(lambda value: safe_rate(value, total_brand_sales))
    best_brand_df["마진 비중"] = best_brand_df["마진"].apply(lambda value: safe_rate(value, total_brand_margin))
    best_brand_df = best_brand_df.sort_values(by=["매출", "마진", "브랜드"], ascending=[False, False, True]).reset_index(drop=True)
    best_brand_name = best_brand_df.iloc[0]["브랜드"] if not best_brand_df.empty else ""
    best_brand_share = best_brand_df.iloc[0]["매출 비중"] if not best_brand_df.empty else 0

    best_day = daily_trend_df.sort_values(by=["매출", "마진"], ascending=[False, False]).head(1)
    best_day_text = best_day.iloc[0]["날짜"] if not best_day.empty else ""
    best_day_sales = int(best_day.iloc[0]["매출"]) if not best_day.empty else 0

    platform_top_df = make_top_with_other(platform_month_df, "플랫폼", 5)
    brand_top_df = make_top_with_other(best_brand_df, "브랜드", 5)
    low_margin_platform = platform_month_df[platform_month_df["매출"].gt(0)].sort_values(by=["마진율", "매출"], ascending=[True, False]).head(1)
    low_margin_text = ""
    if not low_margin_platform.empty:
        low_margin_text = f"{low_margin_platform.iloc[0]['플랫폼']} 마진율 {low_margin_platform.iloc[0]['마진율']:.2%}"

    rows = [
        [f"{target_month}월 온라인 월마감 리포트", "", "", "", "", "", "", "", "", "", "", ""],
        ["실제 출고 내역 기준 | 보고용 요약", "", "", "", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", "", "", ""],
        ["총 매출", "", "", "총 마진", "", "", "마진율", "", "", "판매 수량", "", ""],
        [total_sales, "", "", total_margin, "", "", margin_rate, "", "", total_qty, "", ""],
        ["월 전체 매출", "", "", "매출-비용 기준", "", "", "전체 마진/매출", "", "", "실제 출고 수량", "", ""],
        ["", "", "", "", "", "", "", "", "", "", "", ""],
        ["주문건수", "", "", "객단가", "", "", "주문당 마진", "", "", "정산", "", ""],
        [total_orders, "", "", average_order, "", "", margin_per_order, "", "", total_income, "", ""],
        ["월 주문 수", "", "", "매출/주문", "", "", "마진/주문", "", "", "총 정산금액", "", ""],
        ["", "", "", "", "", "", "", "", "", "", "", ""],
        ["이번 달 핵심 요약", "", "", "", "", "", "", "", "", "", "", ""],
        [f"1. 매출 1위 플랫폼은 {best_platform_name}이며 전체 매출의 {best_platform_share:.2%}를 차지합니다.", "", "", "", "", "", "", "", "", "", "", ""],
        [f"2. 매출 1위 브랜드는 {best_brand_name}이며 전체 브랜드 매출의 {best_brand_share:.2%}를 차지합니다.", "", "", "", "", "", "", "", "", "", "", ""],
        [f"3. 최고 매출일은 {best_day_text} / {format_int(best_day_sales)}원이며, 점검 포인트는 {low_margin_text}입니다.", "", "", "", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", "", "", ""],
        ["플랫폼 TOP 5 + 기타", "", "", "", "", "", "브랜드 TOP 5 + 기타", "", "", "", "", ""],
        ["플랫폼", "매출", "마진", "마진율", "매출 비중", "마진 비중", "브랜드", "매출", "마진", "마진율", "매출 비중", "마진 비중"],
    ]

    platform_rows = platform_top_df.reset_index(drop=True)
    brand_rows = brand_top_df.reset_index(drop=True)
    max_rows = max(len(platform_rows), len(brand_rows))
    for idx in range(max_rows):
        left_values = platform_rows.iloc[idx].tolist() if idx < len(platform_rows) else ["", "", "", "", "", ""]
        right_values = brand_rows.iloc[idx].tolist() if idx < len(brand_rows) else ["", "", "", "", "", ""]
        rows.append(left_values + right_values)

    return pd.DataFrame(rows, columns=[f"col_{idx}" for idx in range(1, 13)])


def add_total_row(df: pd.DataFrame, label_column: str) -> pd.DataFrame:
    if df.empty:
        return df
    output_df = df.copy()
    numeric_columns = output_df.select_dtypes(include="number").columns.tolist()
    total_row = {col: "" for col in output_df.columns}
    total_row[label_column] = "합계"
    for col in numeric_columns:
        if col in {"마진율", "매출 비중", "마진 비중", "플랫폼 내 매출 비중", "플랫폼 내 마진 비중"}:
            continue
        total_row[col] = output_df[col].sum()
    if "마진율" in output_df.columns and "마진" in output_df.columns:
        sales_col = "매출" if "매출" in output_df.columns else "판매가"
        total_row["마진율"] = safe_rate(total_row.get("마진", 0), total_row.get(sales_col, 0))
    return pd.concat([pd.DataFrame([total_row]), output_df], ignore_index=True)


def coerce_excel_value(cell):
    value = cell.value
    if not isinstance(value, str):
        return
    text = value.strip()
    if text == "":
        return
    if re.fullmatch(r"-?\d+(?:\.\d+)?%", text):
        cell.value = float(text.replace("%", "")) / 100
    elif re.fullmatch(r"-?\d[\d,]*", text):
        cell.value = int(text.replace(",", ""))
    elif re.fullmatch(r"-?\d[\d,]*\.\d+", text):
        cell.value = float(text.replace(",", ""))


def style_dashboard_sheet(ws):
    title_fill = PatternFill("solid", fgColor="12324A")
    section_fill = PatternFill("solid", fgColor="EEF5F8")
    sub_section_fill = PatternFill("solid", fgColor="F7FAFC")
    header_fill = PatternFill("solid", fgColor="DCEAF2")
    previous_section_fill = PatternFill("solid", fgColor="E8EDF6")
    previous_header_fill = PatternFill("solid", fgColor="D9E2F3")
    previous_value_fill = PatternFill("solid", fgColor="F4F6FB")
    kpi_fill = PatternFill("solid", fgColor="F4F8FB")
    thin = Side(style="thin", color="D7E2EA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = None
    widths = [22, 16, 16, 16, 13, 13, 13, 13]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for col in range(10, 34):
        ws.column_dimensions[get_column_letter(col)].width = 12

    percent_headers = {"마진율", "매출 비중", "마진 비중", "플랫폼 내 매출 비중", "플랫폼 내 마진 비중", "전월 매출 대비", "전월 정산 대비", "전월 마진 대비", "반품율"}
    integer_headers = {"매출", "정산", "마진", "판매 수량", "수량", "반품 건수", "반품 수량", "반품 매출", "반품 정산", "반품 공급가", "반품 마진", "출고 수량"}
    active_headers = {}
    section_titles = {"월간 KPI", "플랫폼 요약", "반품 요약", "브랜드 전체 요약", "플랫폼별 브랜드 상세"}

    for row in range(1, ws.max_row + 1):
        first_value = ws.cell(row=row, column=1).value
        is_blank = all(ws.cell(row=row, column=col).value in (None, "") for col in range(1, ws.max_column + 1))
        ws.row_dimensions[row].height = 22
        if is_blank:
            ws.row_dimensions[row].height = 10
            continue
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            coerce_excel_value(cell)
            cell.alignment = center
            cell.font = Font(size=11)

        if row == 1:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ws.max_column)
            cell = ws.cell(row=row, column=1)
            cell.fill = title_fill
            cell.font = Font(bold=True, size=16, color="FFFFFF")
            cell.border = border
            ws.row_dimensions[row].height = 32
            continue

        if first_value in section_titles:
            if row == 2:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
                ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=8)
                merge_blocks = [(1, 5, section_fill), (6, 8, previous_section_fill)]
            else:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ws.max_column)
                merge_blocks = [(1, ws.max_column, section_fill)]
            for start_col, end_col, fill in merge_blocks:
                cell = ws.cell(row=row, column=start_col)
                cell.fill = fill
                cell.font = Font(bold=True, size=12, color="17324D")
                cell.border = border
                cell.alignment = center
                for border_col in range(start_col + 1, end_col + 1):
                    ws.cell(row=row, column=border_col).border = border
                    ws.cell(row=row, column=border_col).fill = fill
            active_headers = {}
            continue

        if row == 3 and first_value == "매출":
            active_headers = {col: str(ws.cell(row=row, column=col).value or "").strip() for col in range(1, ws.max_column + 1)}
            ws.row_dimensions[row].height = 24
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value in (None, ""):
                    continue
                cell.fill = previous_header_fill if col >= 6 else header_fill
                cell.font = Font(bold=True, size=11, color="17324D")
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            continue

        if isinstance(first_value, str) and first_value.endswith(" 브랜드"):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ws.max_column)
            cell = ws.cell(row=row, column=1)
            cell.fill = sub_section_fill
            cell.font = Font(bold=True, size=11, color="17324D")
            cell.border = border
            cell.alignment = center
            active_headers = {}
            continue

        if first_value in {"플랫폼", "브랜드", "날짜"}:
            active_headers = {col: str(ws.cell(row=row, column=col).value or "").strip() for col in range(1, ws.max_column + 1)}
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value in (None, ""):
                    continue
                cell.fill = header_fill
                cell.font = Font(bold=True, size=11, color="17324D")
                cell.border = border
                cell.alignment = center
            continue

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if row == 4 and col in range(1, 9):
                cell.border = border
                cell.fill = previous_value_fill if col >= 6 else kpi_fill
                cell.font = Font(bold=True, size=11)
                header_name = active_headers.get(col, "")
                if col in {6, 7, 8}:
                    header_name = f"전월 {header_name} 대비"
                if header_name in percent_headers and isinstance(cell.value, Number):
                    cell.number_format = "0.00%"
                elif header_name in integer_headers and isinstance(cell.value, Number):
                    cell.number_format = "#,##0"
                continue
            if cell.value in (None, ""):
                continue
            cell.border = border
            header_name = active_headers.get(col, "")
            if col == 1:
                cell.alignment = center
            if header_name in percent_headers and isinstance(cell.value, Number):
                cell.number_format = "0.00%"
            elif header_name in integer_headers and isinstance(cell.value, Number):
                cell.number_format = "#,##0"


def style_presentation_dashboard_sheet(ws):
    dark_fill = PatternFill("solid", fgColor="0F2D3E")
    subtitle_fill = PatternFill("solid", fgColor="EDF4F7")
    card_label_fill = PatternFill("solid", fgColor="16384C")
    card_value_fill = PatternFill("solid", fgColor="F8FBFC")
    card_note_fill = PatternFill("solid", fgColor="F1F6F8")
    section_fill = PatternFill("solid", fgColor="E8F1F5")
    insight_fill = PatternFill("solid", fgColor="FBFCFD")
    header_fill = PatternFill("solid", fgColor="D7E6EE")
    thin = Side(style="thin", color="D1DDE5")
    hair = Side(style="hair", color="E7EEF3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    soft_border = Border(left=hair, right=hair, top=hair, bottom=hair)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = None

    widths = [18, 14, 12, 14, 12, 12, 18, 14, 12, 13, 12, 12]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    percent_headers = {"마진율", "매출 비중", "마진 비중"}
    money_headers = {"매출", "정산", "마진", "객단가", "주문당 마진"}
    integer_headers = {"판매 수량", "주문건수", "브랜드 수"}

    merge_ranges = [
        "A1:L1", "A2:L2",
        "A4:C4", "D4:F4", "G4:I4", "J4:L4",
        "A5:C5", "D5:F5", "G5:I5", "J5:L5",
        "A6:C6", "D6:F6", "G6:I6", "J6:L6",
        "A8:C8", "D8:F8", "G8:I8", "J8:L8",
        "A9:C9", "D9:F9", "G9:I9", "J9:L9",
        "A10:C10", "D10:F10", "G10:I10", "J10:L10",
        "A12:L12", "A13:L13", "A14:L14", "A15:L15",
        "A17:F17", "G17:L17",
    ]
    for cell_range in merge_ranges:
        ws.merge_cells(cell_range)

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 24
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            coerce_excel_value(cell)
            cell.alignment = center
            cell.font = Font(size=10, color="111827")
            if cell.value not in (None, ""):
                cell.border = soft_border

    ws.row_dimensions[1].height = 36
    ws["A1"].fill = dark_fill
    ws["A1"].font = Font(bold=True, size=19, color="FFFFFF")
    ws["A1"].alignment = center
    ws["A2"].fill = subtitle_fill
    ws["A2"].font = Font(size=11, color="536878")
    ws.row_dimensions[2].height = 26

    for row in [4, 8]:
        for start_col in [1, 4, 7, 10]:
            cell = ws.cell(row=row, column=start_col)
            cell.fill = card_label_fill
            cell.font = Font(bold=True, size=10, color="FFFFFF")
            cell.border = border
    for row in [5, 9]:
        ws.row_dimensions[row].height = 34
        for start_col in [1, 4, 7, 10]:
            cell = ws.cell(row=row, column=start_col)
            cell.fill = card_value_fill
            cell.font = Font(bold=True, size=15, color="111827")
            cell.border = border
    for row in [6, 10]:
        ws.row_dimensions[row].height = 22
        for start_col in [1, 4, 7, 10]:
            cell = ws.cell(row=row, column=start_col)
            cell.fill = card_note_fill
            cell.font = Font(size=9, color="607080")
            cell.border = border

    for row in [12, 17]:
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value not in (None, ""):
                cell.fill = section_fill
                cell.font = Font(bold=True, size=12, color="17324D")
                cell.border = border

    for row in [13, 14, 15]:
        ws.row_dimensions[row].height = 24
        ws.cell(row=row, column=1).fill = insight_fill
        ws.cell(row=row, column=1).font = Font(size=10, color="263746")
        ws.cell(row=row, column=1).alignment = left
        ws.cell(row=row, column=1).border = border

    for row in [18]:
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value not in (None, ""):
                cell.fill = header_fill
                cell.font = Font(bold=True, size=10, color="17324D")
                cell.border = border

    active_headers = {}
    for row in range(1, ws.max_row + 1):
        first_value = ws.cell(row=row, column=1).value
        if first_value in {"플랫폼", "브랜드"}:
            active_headers = {col: str(ws.cell(row=row, column=col).value or "").strip() for col in range(1, ws.max_column + 1)}
            continue
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            header_name = active_headers.get(col, "")
            if col in {1, 7} and isinstance(cell.value, str) and row > 18:
                cell.alignment = left
            elif header_name in money_headers and isinstance(cell.value, Number):
                cell.alignment = right
            if header_name in percent_headers and isinstance(cell.value, Number):
                cell.number_format = "0.00%"
            elif header_name in money_headers | integer_headers and isinstance(cell.value, Number):
                cell.number_format = "#,##0"

    for cell in ["A5", "D5", "J5", "A9", "D9", "G9", "J9"]:
        ws[cell].number_format = "#,##0"
    ws["G5"].number_format = "0.00%"


def add_presentation_dashboard_charts(ws):
    def chart_anchor(start_col: int, start_row: int, end_col: int, end_row: int) -> TwoCellAnchor:
        return TwoCellAnchor(
            _from=AnchorMarker(col=start_col - 1, row=start_row - 1),
            to=AnchorMarker(col=end_col - 1, row=end_row - 1),
        )

    platform_header_row = 18
    data_start = 19
    data_end = data_start
    while data_end <= ws.max_row and ws.cell(row=data_end, column=1).value not in (None, ""):
        data_end += 1
    data_end -= 1

    if data_end >= data_start:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "플랫폼 TOP"
        chart.legend.position = "b"
        chart.y_axis.majorGridlines = None
        for data_col in [2, 3]:
            data = Reference(ws, min_col=data_col, max_col=data_col, min_row=platform_header_row, max_row=data_end)
            chart.add_data(data, titles_from_data=True)
        cats = Reference(ws, min_col=1, min_row=data_start, max_row=data_end)
        chart.set_categories(cats)
        chart.anchor = chart_anchor(1, data_end + 3, 6, data_end + 16)
        ws.add_chart(chart)

        brand_chart = BarChart()
        brand_chart.type = "col"
        brand_chart.style = 10
        brand_chart.title = "브랜드 TOP"
        brand_chart.legend.position = "b"
        brand_chart.y_axis.majorGridlines = None
        for data_col in [8, 9]:
            data = Reference(ws, min_col=data_col, max_col=data_col, min_row=platform_header_row, max_row=data_end)
            brand_chart.add_data(data, titles_from_data=True)
        cats = Reference(ws, min_col=7, min_row=data_start, max_row=data_end)
        brand_chart.set_categories(cats)
        brand_chart.anchor = chart_anchor(7, data_end + 3, 12, data_end + 16)
        ws.add_chart(brand_chart)


def build_executive_dashboard_df(
    target_month: int,
    platform_month_df: pd.DataFrame,
    brand_month_df: pd.DataFrame,
    daily_trend_df: pd.DataFrame,
) -> pd.DataFrame:
    total_sales = int(platform_month_df["매출"].sum())
    total_income = int(platform_month_df["정산"].sum())
    total_margin = int(platform_month_df["마진"].sum())
    total_qty = int(platform_month_df["판매 수량"].sum())
    margin_rate = safe_rate(total_margin, total_sales)

    platform_df = platform_month_df[["플랫폼", "매출", "마진", "마진율", "매출 비중", "마진 비중"]].copy()
    platform_df = platform_df.sort_values(by=["매출", "마진", "플랫폼"], ascending=[False, False, True]).reset_index(drop=True)
    platform_df.insert(0, "순위", range(1, len(platform_df) + 1))

    brand_df = (
        brand_month_df.groupby("브랜드", dropna=False, sort=False)
        .agg(수량=("수량", "sum"), 매출=("매출", "sum"), 정산=("정산", "sum"), 마진=("마진", "sum"))
        .reset_index()
    )
    brand_total_sales = brand_df["매출"].sum()
    brand_total_margin = brand_df["마진"].sum()
    brand_df["마진율"] = brand_df["마진"].div(brand_df["매출"].replace(0, pd.NA)).fillna(0)
    brand_df["매출 비중"] = brand_df["매출"].apply(lambda value: safe_rate(value, brand_total_sales))
    brand_df["마진 비중"] = brand_df["마진"].apply(lambda value: safe_rate(value, brand_total_margin))
    brand_df = brand_df.sort_values(by=["매출", "마진", "브랜드"], ascending=[False, False, True]).reset_index(drop=True)
    brand_df.insert(0, "순위", range(1, len(brand_df) + 1))

    best_platform = platform_df.head(1)
    best_brand = brand_df.head(1)
    best_day = daily_trend_df.sort_values(by=["매출", "마진"], ascending=[False, False]).head(1)
    best_platform_text = ""
    if not best_platform.empty:
        best_platform_text = f"{best_platform.iloc[0]['플랫폼']} / 매출 비중 {best_platform.iloc[0]['매출 비중']:.2%}"
    best_brand_text = ""
    if not best_brand.empty:
        best_brand_text = f"{best_brand.iloc[0]['브랜드']} / 매출 비중 {best_brand.iloc[0]['매출 비중']:.2%}"
    best_day_text = ""
    if not best_day.empty:
        best_day_text = f"{best_day.iloc[0]['날짜']} / {format_int(best_day.iloc[0]['매출'])}원"

    rows = [
        [f"{target_month}월 온라인 월마감 Executive Report", "", "", "", "", "", "", ""],
        ["전체 플랫폼/브랜드 출력 | 실제 출고 기준 월간 성과", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", ""],
        ["매출", "", "마진", "", "마진율", "", "판매 수량", ""],
        [total_sales, "", total_margin, "", margin_rate, "", total_qty, ""],
        ["총 판매금액", "", "총 마진금액", "", "마진/매출", "", "실제 출고 수량", ""],
        ["", "", "", "", "", "", "", ""],
        ["핵심 체크", "", "", "", "", "", "", ""],
        [f"1. 플랫폼 매출 1위: {best_platform_text}", "", "", "", "", "", "", ""],
        [f"2. 브랜드 매출 1위: {best_brand_text}", "", "", "", "", "", "", ""],
        [f"3. 최고 매출일: {best_day_text}", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", ""],
    ]

    platform_cols = ["순위", "플랫폼", "매출", "마진", "마진율", "매출 비중", "마진 비중"]
    brand_cols = ["순위", "브랜드", "매출", "마진", "마진율", "매출 비중", "마진 비중"]
    rows.extend([["", "", "", "", "", "", "", ""] for _ in range(14)])
    rows.append(["플랫폼 전체 성과", "", "", "", "", "", "", ""])
    rows.append(["순위", "플랫폼", "매출", "마진", "마진율", "매출 비중", "마진 비중", ""])
    for _, row in platform_df.iterrows():
        rows.append(row[platform_cols].tolist() + [""])

    rows.extend([["", "", "", "", "", "", "", ""] for _ in range(3)])
    rows.extend([["", "", "", "", "", "", "", ""] for _ in range(14)])
    rows.append(["브랜드 전체 성과", "", "", "", "", "", "", ""])
    rows.append(["순위", "브랜드", "매출", "마진", "마진율", "매출 비중", "마진 비중", ""])
    for _, row in brand_df.iterrows():
        rows.append(row[brand_cols].tolist() + [""])

    return pd.DataFrame(rows, columns=[f"col_{idx}" for idx in range(1, 9)])


def style_executive_dashboard_sheet(ws):
    dark_fill = PatternFill("solid", fgColor="102A3A")
    subtitle_fill = PatternFill("solid", fgColor="EDF3F6")
    card_label_fill = PatternFill("solid", fgColor="174154")
    card_value_fill = PatternFill("solid", fgColor="F8FBFC")
    card_note_fill = PatternFill("solid", fgColor="F1F6F8")
    section_fill = PatternFill("solid", fgColor="E5EFF4")
    header_fill = PatternFill("solid", fgColor="D5E5EE")
    insight_fill = PatternFill("solid", fgColor="FBFCFD")
    thin = Side(style="thin", color="D3DEE6")
    hair = Side(style="hair", color="E8EEF3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    soft_border = Border(left=hair, right=hair, top=hair, bottom=hair)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = None
    widths = [8, 22, 16, 16, 12, 13, 13, 4]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    merge_ranges = [
        "A1:H1", "A2:H2",
        "A4:B4", "C4:D4", "E4:F4", "G4:H4",
        "A5:B5", "C5:D5", "E5:F5", "G5:H5",
        "A6:B6", "C6:D6", "E6:F6", "G6:H6",
        "A8:H8", "A9:H9", "A10:H10", "A11:H11",
    ]
    for cell_range in merge_ranges:
        ws.merge_cells(cell_range)

    percent_headers = {"마진율", "매출 비중", "마진 비중"}
    money_headers = {"매출", "마진", "정산", "객단가", "주문당 마진"}
    integer_headers = {"순위", "판매 수량"}
    active_headers = {}

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 24
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            coerce_excel_value(cell)
            cell.alignment = center
            cell.font = Font(size=10, color="111827")
            if cell.value not in (None, ""):
                cell.border = soft_border

    ws.row_dimensions[1].height = 36
    ws["A1"].fill = dark_fill
    ws["A1"].font = Font(bold=True, size=19, color="FFFFFF")
    ws["A2"].fill = subtitle_fill
    ws["A2"].font = Font(size=11, color="536878")

    for row in [4]:
        for start_col in [1, 3, 5, 7]:
            cell = ws.cell(row=row, column=start_col)
            cell.fill = card_label_fill
            cell.font = Font(bold=True, size=10, color="FFFFFF")
            cell.border = border
    for row in [5]:
        ws.row_dimensions[row].height = 34
        for start_col in [1, 3, 5, 7]:
            cell = ws.cell(row=row, column=start_col)
            cell.fill = card_value_fill
            cell.font = Font(bold=True, size=15, color="111827")
            cell.border = border
    for row in [6]:
        for start_col in [1, 3, 5, 7]:
            cell = ws.cell(row=row, column=start_col)
            cell.fill = card_note_fill
            cell.font = Font(size=9, color="607080")
            cell.border = border

    for row in [8]:
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value not in (None, ""):
                cell.fill = section_fill
                cell.font = Font(bold=True, size=12, color="17324D")
                cell.border = border

    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value in {"플랫폼 전체 성과", "브랜드 전체 성과"}:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            cell = ws.cell(row=row, column=1)
            cell.fill = section_fill
            cell.font = Font(bold=True, size=12, color="17324D")
            cell.border = border

    for row in [9, 10, 11]:
        cell = ws.cell(row=row, column=1)
        cell.fill = insight_fill
        cell.font = Font(size=10, color="263746")
        cell.alignment = left
        cell.border = border

    for row in range(1, ws.max_row + 1):
        first_value = ws.cell(row=row, column=1).value
        if first_value == "순위":
            active_headers = {col: str(ws.cell(row=row, column=col).value or "").strip() for col in range(1, ws.max_column + 1)}
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value not in (None, ""):
                    cell.fill = header_fill
                    cell.font = Font(bold=True, size=10, color="17324D")
                    cell.border = border
            continue

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            header_name = active_headers.get(col, "")
            if col == 2 and isinstance(cell.value, str):
                cell.alignment = left
            elif header_name in money_headers and isinstance(cell.value, Number):
                cell.alignment = right
            if header_name in percent_headers and isinstance(cell.value, Number):
                cell.number_format = "0.00%"
            elif header_name in money_headers | integer_headers and isinstance(cell.value, Number):
                cell.number_format = "#,##0"

    for cell in ["A5", "C5", "G5"]:
        ws[cell].number_format = "#,##0"
    ws["E5"].number_format = "0.00%"


def add_executive_dashboard_charts(ws):
    def chart_anchor(start_col: int, start_row: int, end_col: int, end_row: int) -> TwoCellAnchor:
        return TwoCellAnchor(
            _from=AnchorMarker(col=start_col - 1, row=start_row - 1),
            to=AnchorMarker(col=end_col - 1, row=end_row - 1),
        )

    def add_section_chart(section_title: str, chart_title: str):
        section_row = None
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == section_title:
                section_row = row
                break
        if not section_row:
            return

        header_row = section_row + 1
        data_start = section_row + 2
        data_end = data_start
        while data_end <= ws.max_row and ws.cell(row=data_end, column=2).value not in (None, ""):
            data_end += 1
        data_end -= 1
        if data_end < data_start:
            return

        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = chart_title
        chart.legend.position = "b"
        chart.y_axis.majorGridlines = None
        for data_col in [3, 4]:
            chart.add_data(Reference(ws, min_col=data_col, max_col=data_col, min_row=header_row, max_row=data_end), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=2, min_row=data_start, max_row=data_end))
        chart.anchor = chart_anchor(1, max(section_row - 14, 1), 8, section_row - 1)
        ws.add_chart(chart)

    add_section_chart("플랫폼 전체 성과", "플랫폼 전체 매출/마진")
    add_section_chart("브랜드 전체 성과", "브랜드 전체 매출/마진")


def style_table_sheet(ws):
    header_fill = PatternFill("solid", fgColor="D7E6F3")
    total_fill = PatternFill("solid", fgColor="EAF2F8")
    thin = Side(style="thin", color="D9E2EA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    percent_headers = {"마진율", "매출 비중", "마진 비중", "플랫폼 내 매출 비중", "플랫폼 내 마진 비중"}
    integer_headers = {"주문건수", "판매 수량", "수량", "매출", "판매가", "수수료", "정산", "수입", "공급가", "마진"}
    text_headers = {"플랫폼", "브랜드"}
    money_headers = {"매출", "판매가", "수수료", "정산", "수입", "공급가", "마진"}
    header_map = {col: str(ws.cell(row=1, column=col).value or "").strip() for col in range(1, ws.max_column + 1)}

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = None
    for column_cells in ws.columns:
        column_letter = column_cells[0].column_letter
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_letter].width = min(max(max_length + 4, 12), 24)

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 22
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            coerce_excel_value(cell)
            cell.alignment = center
            cell.border = border
            cell.font = Font(size=11)
            header_name = header_map.get(col, "")
            if row == 1:
                cell.fill = header_fill
                cell.font = Font(bold=True, size=11, color="17324D")
            elif row == 2 and ws.cell(row=row, column=1).value == "합계":
                cell.fill = total_fill
                cell.font = Font(bold=True, size=11)
            if row > 1 and header_name in text_headers:
                cell.alignment = left
            elif row > 1 and header_name in money_headers:
                cell.alignment = right
            if row > 1 and header_name in percent_headers and isinstance(cell.value, Number):
                cell.number_format = "0.00%"
            elif row > 1 and header_name in integer_headers and isinstance(cell.value, Number):
                cell.number_format = "#,##0"


def style_top_sheet(ws):
    header_fill = PatternFill("solid", fgColor="D7E6F3")
    section_fill = PatternFill("solid", fgColor="EAF2F8")
    thin = Side(style="thin", color="D9E2EA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = None

    if ws.max_column == 0:
        return

    block_size = len(TOP_BLOCK_HEADERS)
    stride = block_size + TOP_BLOCK_SPACER_COLS
    for col in range(1, ws.max_column + 1):
        mod = (col - 1) % stride
        width = 9 if mod in {0, block_size} else 14
        if mod == 1:
            width = 20
        ws.column_dimensions[get_column_letter(col)].width = width

    header_rows = set()
    title_rows = set()
    for row in range(1, ws.max_row + 1):
        values = [str(ws.cell(row=row, column=col).value or "").strip() for col in range(1, ws.max_column + 1)]
        if "순위" in values:
            header_rows.add(row)
            if row > 1:
                title_rows.add(row - 1)

    for title_row in title_rows:
        for start_col in range(1, ws.max_column + 1, stride):
            end_col = start_col + block_size - 1
            title_value = str(ws.cell(row=title_row, column=start_col).value or "").strip()
            if title_value == "" or end_col > ws.max_column:
                continue
            ws.merge_cells(start_row=title_row, start_column=start_col, end_row=title_row, end_column=end_col)
            cell = ws.cell(row=title_row, column=start_col)
            cell.fill = section_fill
            cell.font = Font(bold=True, size=12, color="17324D")
            cell.alignment = center
            cell.border = border

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 22
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            mod = (col - 1) % stride
            if mod == block_size or cell.value in (None, ""):
                continue
            cell.alignment = center
            cell.border = border
            cell.font = Font(size=11)
            if row in header_rows:
                cell.fill = header_fill
                cell.font = Font(bold=True, size=11, color="17324D")
                continue
            header_row = max((candidate for candidate in header_rows if candidate < row), default=None)
            header_name = str(ws.cell(row=header_row, column=col).value or "").strip() if header_row else ""
            if header_name in {"순위", "판매 수량", "현재고", "매출", "마진"} and isinstance(cell.value, Number):
                cell.number_format = "#,##0"
            elif header_name == "마진율" and isinstance(cell.value, Number):
                cell.number_format = "0.00%"


def add_dashboard_charts(ws):
    platform_header_row = None
    return_header_row = None
    brand_total_header_row = None
    platform_brand_sections = []
    for row in range(1, ws.max_row + 1):
        first = ws.cell(row=row, column=1).value
        if first == "플랫폼" and row > 1 and ws.cell(row=row - 1, column=1).value == "반품 요약":
            return_header_row = row
        elif first == "플랫폼" and platform_header_row is None:
            platform_header_row = row
        elif first == "브랜드" and brand_total_header_row is None:
            brand_total_header_row = row
        elif (
            isinstance(first, str)
            and first.endswith(" 브랜드")
            and row + 1 <= ws.max_row
            and ws.cell(row=row + 1, column=1).value == "브랜드"
        ):
            platform_brand_sections.append((row, row + 1, first))

    def chart_anchor(start_col: int, start_row: int, end_col: int, end_row: int) -> TwoCellAnchor:
        return TwoCellAnchor(
            _from=AnchorMarker(col=start_col - 1, row=start_row - 1),
            to=AnchorMarker(col=end_col - 1, row=end_row - 1),
        )

    def add_named_bar_chart(
        header_row: int | None,
        title: str,
        start_row: int,
        end_row_anchor: int | None,
        data_columns: list[int],
        anchor_start_col: int = 10,
        anchor_end_col: int = 20,
        show_data_labels: bool = True,
    ):
        if not header_row:
            return
        end_row = header_row + 1
        while end_row <= ws.max_row and ws.cell(row=end_row, column=1).value not in (None, ""):
            end_row += 1
        end_row -= 1
        if end_row <= header_row:
            return
        if all(ws.cell(row=row, column=data_col).value in (None, "") for row in range(header_row + 1, end_row + 1) for data_col in data_columns):
            return
        if end_row_anchor is None:
            end_row_anchor = end_row + 1

        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = title
        chart.legend = None if len(data_columns) == 1 else chart.legend
        if chart.legend:
            chart.legend.position = "r"
        chart.y_axis.title = None
        chart.x_axis.title = None
        chart.y_axis.tickLblPos = "nextTo"
        chart.x_axis.tickLblPos = "nextTo"
        chart.x_axis.tickLblSkip = 1
        chart.y_axis.delete = False
        chart.x_axis.delete = False
        if len(data_columns) == 1 and show_data_labels:
            chart.dLbls = DataLabelList()
            chart.dLbls.showCatName = False
            chart.dLbls.showVal = True
            chart.dLbls.showLegendKey = False
        for data_col in data_columns:
            data = Reference(ws, min_col=data_col, max_col=data_col, min_row=header_row, max_row=end_row)
            chart.add_data(data, titles_from_data=True)
        cats = Reference(ws, min_col=1, min_row=header_row + 1, max_row=end_row)
        chart.set_categories(cats)
        chart.anchor = chart_anchor(anchor_start_col, start_row, anchor_end_col, end_row_anchor)
        ws.add_chart(chart)

    if platform_header_row:
        add_named_bar_chart(platform_header_row, "플랫폼별 매출/마진", platform_header_row, None, [2, 4])

    if return_header_row:
        add_named_bar_chart(return_header_row, "플랫폼별 반품율", return_header_row, None, [8], show_data_labels=False)

    if brand_total_header_row:
        add_named_bar_chart(brand_total_header_row, "브랜드별 매출/마진", brand_total_header_row, None, [2, 4])

    for section_row, header_row, section_title in platform_brand_sections:
        chart_title = f"{section_title} 매출/마진"
        add_named_bar_chart(header_row, chart_title, header_row, None, [2, 4])


def save_report(
    target_month: int,
    dashboard_df: pd.DataFrame,
    best_top10_df: pd.DataFrame,
    brand_top10_df: pd.DataFrame,
    brand_margin_top10_df: pd.DataFrame,
) -> Path:
    output_dir = LIST_DIR
    output_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = output_dir / f"{target_month}월 월마감.xlsx"

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        dashboard_df.to_excel(writer, index=False, header=False, sheet_name="월 마감")
        best_top10_df.to_excel(writer, index=False, header=False, sheet_name="BEST TOP 10")
        brand_top10_df.to_excel(writer, index=False, header=False, sheet_name="BRAND TOP 10")
        brand_margin_top10_df.to_excel(writer, index=False, header=False, sheet_name="브랜드 마진 TOP 10")

    workbook = load_workbook(xlsx_path)
    style_dashboard_sheet(workbook["월 마감"])
    add_dashboard_charts(workbook["월 마감"])
    style_top_sheet(workbook["BEST TOP 10"])
    style_top_sheet(workbook["BRAND TOP 10"])
    style_top_sheet(workbook["브랜드 마진 TOP 10"])
    workbook.save(xlsx_path)
    return xlsx_path


def main():
    start_time = time.perf_counter()
    log("월마감 원본 플랫폼/반품 시트 다운로드 시작")
    source_raw_df = read_combined_platform_source_df()
    stock_raw_df = read_csv_from_url(CURRENT_STOCK_URL)

    target_month = parse_target_month(sys.argv, source_raw_df)
    source_sales_df = build_source_sales_df(source_raw_df, target_month)
    if source_sales_df.empty:
        raise ValueError(f"{target_month}월 실제 출고 데이터가 없습니다.")

    log(f"{target_month}월 데이터 집계 시작")
    previous_source_sales_df = build_source_sales_df(source_raw_df, previous_month(target_month))
    previous_month_summary = build_month_summary(previous_source_sales_df)
    platform_month_df = build_platform_month_df(source_sales_df)
    brand_month_df = build_brand_month_df(source_sales_df)
    return_month_df = build_return_month_df(source_sales_df)
    daily_trend_df = build_daily_trend_df(source_sales_df)
    dashboard_df = build_dashboard_df(target_month, platform_month_df, brand_month_df, daily_trend_df, previous_month_summary, return_month_df)
    stock_lookup = build_stock_lookup(stock_raw_df)
    best_top10_df = build_source_top10_df(source_raw_df, stock_lookup, target_month, "플랫폼", "best_top")
    brand_top10_df = build_source_top10_df(source_raw_df, stock_lookup, target_month, "브랜드", "brand_top")
    brand_margin_top10_df = build_source_margin_top10_df(source_raw_df, stock_lookup, target_month, "브랜드", "brand_margin_top")
    xlsx_path = save_report(target_month, dashboard_df, best_top10_df, brand_top10_df, brand_margin_top10_df)

    elapsed = time.perf_counter() - start_time
    log(f"엑셀 보고서 저장 완료 -> {xlsx_path}")
    log(f"완료 ({elapsed:.2f}초)")


if __name__ == "__main__":
    main()
