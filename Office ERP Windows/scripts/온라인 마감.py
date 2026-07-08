# -*- coding: utf-8 -*-

from __future__ import annotations

from copy import copy
from io import StringIO
from numbers import Number
from pathlib import Path
import re
import time
from datetime import datetime

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from office_erp_paths import LIST_DIR


PLATFORM_SOURCE_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vRWhzq2rt_O-YnOwzWnE8L_d8--NBu-EMmDkxnnVy-oATAfCjDIX976b973bY2MiO8gYPTE5WCk-tVv/pub?gid=733480714&single=true&output=csv"
PLATFORM_RETURN_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vRWhzq2rt_O-YnOwzWnE8L_d8--NBu-EMmDkxnnVy-oATAfCjDIX976b973bY2MiO8gYPTE5WCk-tVv/pub?gid=1032857881&single=true&output=csv"
CURRENT_STOCK_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vQBC-ZAQixoxbta3FsLfG_f8qugvCGbjK2yrsKcLA6tJsi1ww5qDl-_21od9-55Lv0F9dmzlR5c1QIh/pub?gid=0&single=true&output=csv"

DOWNLOAD_TIMEOUT = (10, 60)
RETRIES = 3
RETRY_SLEEP_SEC = 2

PLATFORM_COLUMNS = ["날짜", "플랫폼", "주문건수", "판매 수량", "판매가", "수수료", "수입", "공급가", "마진", "마진율"]
BRAND_COLUMNS = ["날짜", "플랫폼", "브랜드", "수량", "판매가", "수입", "공급가", "마진", "마진율"]
RETURN_PLATFORM_COLUMNS = ["날짜", "플랫폼", "반품 건수", "반품 수량", "반품 판매가", "반품 수입", "반품 공급가", "반품 마진", "출고 수량", "반품율"]
RETURN_BRAND_COLUMNS = ["날짜", "플랫폼", "브랜드", "반품 건수", "반품 수량", "반품 판매가", "반품 수입", "반품 공급가", "반품 마진", "출고 수량", "반품율"]
RETURN_BRAND_SKU_COLUMNS = ["날짜", "플랫폼", "브랜드", "품번", "반품 수량", "반품 판매가", "반품 수입", "반품 공급가", "반품 마진", "반품율"]
SOURCE_REQUIRED_COLUMNS = [
    "날짜",
    "플랫폼",
    "주문번호",
    "브랜드",
    "품번",
    "수량",
    "총 판매가",
    "총 수수료",
    "총 수입",
    "총 KRW",
    "총 공급가",
    "마진",
    "환율",
    "내역",
    "매장명",
]
RETURN_STATUS_VALUES = {"반품"}
REPORT_COLUMNS = [
    "날짜",
    "플랫폼",
    "판매 수량",
    "매출",
    "정산",
    "마진",
    "마진율",
    "브랜드 수",
    "상위 브랜드 1",
    "상위 브랜드 2",
    "상위 브랜드 3",
]
VERTICAL_REPORT_MAX_COLS = 8

NUMERIC_PLATFORM_COLUMNS = ["주문건수", "판매 수량", "판매가", "수수료", "수입", "공급가", "마진"]
NUMERIC_BRAND_COLUMNS = ["수량", "판매가", "수입", "공급가", "마진"]
NUMERIC_RETURN_PLATFORM_COLUMNS = ["반품 건수", "반품 수량", "반품 판매가", "반품 수입", "반품 공급가", "반품 마진", "출고 수량"]
NUMERIC_RETURN_BRAND_COLUMNS = ["반품 건수", "반품 수량", "반품 판매가", "반품 수입", "반품 공급가", "반품 마진", "출고 수량"]
BEST_TOP_BLOCK_HEADERS = ["순위", "품번", "판매 수량", "현재고", "마진율"]
TOP_BLOCKS_PER_ROW = 3
TOP_BLOCK_SPACER_COLS = 1


def log(message: str):
    print(f"[온라인 마감] {message}", flush=True)


def make_session() -> requests.Session:
    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0"})
    return session


def fetch_csv_text(session: requests.Session, url: str) -> str:
    last_error = None

    for attempt in range(1, RETRIES + 1):
        try:
            response = session.get(url, timeout=DOWNLOAD_TIMEOUT)
            response.raise_for_status()
            response.encoding = "utf-8"
            return response.text
        except Exception as e:
            last_error = e
            if attempt < RETRIES:
                time.sleep(RETRY_SLEEP_SEC)

    raise last_error


def read_csv_from_url(url: str) -> pd.DataFrame:
    csv_text = fetch_csv_text(make_session(), url)
    return pd.read_csv(StringIO(csv_text), dtype=str).fillna("")


def read_required_source_csv(url: str, label: str) -> pd.DataFrame:
    df = read_csv_from_url(url)
    missing_columns = [col for col in SOURCE_REQUIRED_COLUMNS if col not in df.columns]
    if missing_columns:
        raise ValueError(f"{label} 필수 열이 없습니다: {missing_columns}")
    return df[SOURCE_REQUIRED_COLUMNS].copy()


def read_platform_source_df() -> pd.DataFrame:
    return read_required_source_csv(PLATFORM_SOURCE_URL, "플랫폼 원본 시트")


def read_platform_return_df() -> pd.DataFrame:
    return_df = read_required_source_csv(PLATFORM_RETURN_URL, "플랫폼 반품 시트")
    return_df["내역"] = return_df["내역"].apply(clean_text)
    return return_df[return_df["내역"].isin(RETURN_STATUS_VALUES)].copy()


def read_combined_platform_source_df() -> pd.DataFrame:
    source_df = read_platform_source_df()
    return_df = read_platform_return_df()
    if return_df.empty:
        log("플랫폼 반품 시트 계산 대상 없음 -> 내역이 '반품'인 행이 없습니다.")
        return source_df

    log(f"플랫폼 반품 시트 포함 -> 내역 '반품' {len(return_df)}행")
    return pd.concat([source_df, return_df], ignore_index=True)


def clean_text(value) -> str:
    if pd.isna(value):
        return ""
    return " ".join(str(value).strip().split())


def normalize_best_platform(platform_value, order_no_value="") -> str:
    platform = clean_text(platform_value).upper()
    order_no = clean_text(order_no_value).upper()

    if platform == "KASHION":
        if order_no.startswith("LP"):
            return "KASHION (TM)"
        if order_no.startswith("JD"):
            return "KASHION (JD)"
        return "KASHION"

    if platform in {"TM", "天猫", "KASHION (TM)"}:
        return "KASHION (TM)"
    if platform in {"JD", "京东", "KASHION (JD)"}:
        return "KASHION (JD)"
    return clean_text(platform_value)


def to_number_series(series: pd.Series) -> pd.Series:
    cleaned = (
        series.fillna("")
        .astype(str)
        .str.replace(",", "", regex=False)
        .str.replace('"', "", regex=False)
        .str.replace("%", "", regex=False)
        .str.strip()
    )
    return pd.to_numeric(cleaned, errors="coerce").fillna(0)


def select_margin_base_series(sales_series: pd.Series, krw_series: pd.Series) -> pd.Series:
    sales_num = to_number_series(sales_series)
    krw_num = to_number_series(krw_series)
    foreign_currency_mask = krw_num.gt(0) & sales_num.gt(0) & krw_num.gt(sales_num * 10)
    return sales_num.where(~foreign_currency_mask, krw_num)


def format_int(value) -> int:
    return int(round(float(value)))


def format_percent(value) -> str:
    number = pd.to_numeric(value, errors="coerce")
    if pd.isna(number):
        return "0%"
    return f"{float(number) * 100:.2f}%"


def parse_korean_date_key(value: str) -> tuple[int, int]:
    text = clean_text(value).replace(" ", "")
    try:
        if "월" in text and "일" in text:
            month = int(text.split("월", 1)[0])
            day = int(text.split("월", 1)[1].split("일", 1)[0])
            return (month, day)
        parsed_date = pd.to_datetime(text, errors="coerce")
        if not pd.isna(parsed_date):
            return (int(parsed_date.month), int(parsed_date.day))
    except Exception:
        pass
    return (-1, -1)


def parse_korean_month(value: str) -> int:
    text = clean_text(value).replace(" ", "")
    try:
        if "월" in text:
            return int(text.split("월", 1)[0])
        parsed_date = pd.to_datetime(text, errors="coerce")
        if not pd.isna(parsed_date):
            return int(parsed_date.month)
    except Exception:
        pass
    return -1


def today_date_key() -> tuple[int, int]:
    now = datetime.now()
    return (now.month, now.day)


def today_date_text() -> str:
    now = datetime.now()
    return f"{now.month}월 {now.day}일"


def filter_today_date_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "날짜" not in df.columns:
        return df

    target_key = today_date_key()
    today_mask = df["날짜"].apply(lambda value: parse_korean_date_key(value) == target_key)
    return df[today_mask].copy().reset_index(drop=True)


def filter_valid_outbound_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df.copy()

    output_df = df.copy()
    if "내역" in output_df.columns:
        output_df["내역"] = output_df["내역"].apply(clean_text)
        excluded_statuses = {"取消件", "调货中", "재고 없음", "취소", "취소건"}
        output_df = output_df[~output_df["내역"].isin(excluded_statuses)].copy()
    if "매장명" in output_df.columns:
        output_df["매장명"] = output_df["매장명"].apply(clean_text)
        output_df = output_df[~output_df["매장명"].eq("재고없음")].copy()
    return output_df.reset_index(drop=True)


def normalize_source_df(df: pd.DataFrame) -> pd.DataFrame:
    missing_columns = [col for col in SOURCE_REQUIRED_COLUMNS if col not in df.columns]
    if missing_columns:
        raise ValueError(f"원본 플랫폼 시트 필수 열이 없습니다: {missing_columns}")

    output_df = df[SOURCE_REQUIRED_COLUMNS].copy()
    for col in ["날짜", "주문번호", "브랜드", "품번", "내역", "매장명"]:
        output_df[col] = output_df[col].apply(clean_text)
    output_df["플랫폼"] = output_df.apply(
        lambda row: normalize_best_platform(row["플랫폼"], row["주문번호"]),
        axis=1,
    )

    output_df = filter_valid_outbound_rows(output_df)
    output_df = output_df[output_df["날짜"].ne("") & output_df["플랫폼"].ne("")].copy()
    if output_df.empty:
        raise ValueError("원본 플랫폼 시트에서 유효한 출고 데이터가 없습니다.")

    output_df["수량_num"] = to_number_series(output_df["수량"])
    total_sales_num = to_number_series(output_df["총 판매가"])
    total_fee_num = to_number_series(output_df["총 수수료"])
    total_income_num = to_number_series(output_df["총 수입"])
    total_krw_num = to_number_series(output_df["총 KRW"])
    exchange_rate_num = to_number_series(output_df["환율"])
    has_exchange_rate = exchange_rate_num.gt(0)

    output_df["판매가_num"] = total_sales_num.where(~has_exchange_rate, total_sales_num * exchange_rate_num)
    output_df["수수료_num"] = total_fee_num.where(~has_exchange_rate, total_fee_num * exchange_rate_num)
    output_df["수입_num"] = total_krw_num.where(total_krw_num.gt(0), total_income_num.where(~has_exchange_rate, total_income_num * exchange_rate_num))
    output_df["공급가_num"] = to_number_series(output_df["총 공급가"])
    output_df["마진_num"] = to_number_series(output_df["마진"])
    return_mask = output_df["내역"].isin(RETURN_STATUS_VALUES)
    for col in ["수량_num", "판매가_num", "수수료_num", "수입_num", "공급가_num", "마진_num"]:
        output_df.loc[return_mask, col] = -output_df.loc[return_mask, col].abs()
    return output_df.reset_index(drop=True)


def build_platform_summary_from_source_df(source_df: pd.DataFrame) -> pd.DataFrame:
    grouped_df = (
        source_df.groupby(["날짜", "플랫폼"], dropna=False, sort=False)
        .agg(
            주문건수=("수량_num", "size"),
            판매_수량=("수량_num", "sum"),
            판매가=("판매가_num", "sum"),
            수수료=("수수료_num", "sum"),
            수입=("수입_num", "sum"),
            공급가=("공급가_num", "sum"),
            마진=("마진_num", "sum"),
        )
        .reset_index()
        .rename(columns={"판매_수량": "판매 수량"})
    )
    grouped_df["마진율"] = grouped_df["마진"].div(grouped_df["판매가"].replace(0, pd.NA)).fillna(0)

    for col in NUMERIC_PLATFORM_COLUMNS:
        grouped_df[col] = grouped_df[col].apply(format_int)

    return grouped_df[PLATFORM_COLUMNS].copy()


def build_brand_summary_from_source_df(source_df: pd.DataFrame) -> pd.DataFrame:
    grouped_df = (
        source_df.groupby(["날짜", "플랫폼", "브랜드"], dropna=False, sort=False)
        .agg(
            수량=("수량_num", "sum"),
            판매가=("판매가_num", "sum"),
            수입=("수입_num", "sum"),
            공급가=("공급가_num", "sum"),
            마진=("마진_num", "sum"),
        )
        .reset_index()
    )
    grouped_df["마진율"] = grouped_df["마진"].div(grouped_df["판매가"].replace(0, pd.NA)).fillna(0)

    for col in NUMERIC_BRAND_COLUMNS:
        grouped_df[col] = grouped_df[col].apply(format_int)

    return grouped_df[BRAND_COLUMNS].copy()


def build_return_platform_summary_df(source_df: pd.DataFrame) -> pd.DataFrame:
    return_df = source_df[source_df["내역"].isin(RETURN_STATUS_VALUES)].copy()
    if return_df.empty:
        return pd.DataFrame(columns=RETURN_PLATFORM_COLUMNS)

    outbound_df = source_df[~source_df["내역"].isin(RETURN_STATUS_VALUES)].copy()
    return_df["_month_key"] = return_df["날짜"].apply(parse_korean_month)
    outbound_df["_month_key"] = outbound_df["날짜"].apply(parse_korean_month)
    outbound_qty_df = (
        outbound_df.groupby(["_month_key", "플랫폼"], dropna=False, sort=False)
        .agg(출고_수량=("수량_num", "sum"))
        .reset_index()
    )

    grouped_df = (
        return_df.groupby(["날짜", "_month_key", "플랫폼"], dropna=False, sort=False)
        .agg(
            반품_건수=("수량_num", "size"),
            반품_수량=("수량_num", lambda series: series.abs().sum()),
            반품_판매가=("판매가_num", lambda series: series.abs().sum()),
            반품_수입=("수입_num", lambda series: series.abs().sum()),
            반품_공급가=("공급가_num", lambda series: series.abs().sum()),
            반품_마진=("마진_num", lambda series: series.abs().sum()),
        )
        .reset_index()
        .merge(outbound_qty_df, on=["_month_key", "플랫폼"], how="left")
        .fillna({"출고_수량": 0})
        .drop(columns=["_month_key"])
        .rename(
            columns={
                "반품_건수": "반품 건수",
                "반품_수량": "반품 수량",
                "반품_판매가": "반품 판매가",
                "반품_수입": "반품 수입",
                "반품_공급가": "반품 공급가",
                "반품_마진": "반품 마진",
                "출고_수량": "출고 수량",
            }
        )
    )
    grouped_df["반품율"] = grouped_df["반품 수량"].div(grouped_df["출고 수량"].replace(0, pd.NA)).fillna(0)

    for col in NUMERIC_RETURN_PLATFORM_COLUMNS:
        grouped_df[col] = grouped_df[col].apply(format_int)

    grouped_df["_date_key"] = grouped_df["날짜"].apply(parse_korean_date_key)
    grouped_df = grouped_df.sort_values(by=["_date_key", "플랫폼"], ascending=[False, True]).drop(columns=["_date_key"])
    return grouped_df[RETURN_PLATFORM_COLUMNS].copy()


def build_return_brand_summary_df(source_df: pd.DataFrame) -> pd.DataFrame:
    return_df = source_df[source_df["내역"].isin(RETURN_STATUS_VALUES)].copy()
    if return_df.empty:
        return pd.DataFrame(columns=RETURN_BRAND_COLUMNS)

    outbound_df = source_df[~source_df["내역"].isin(RETURN_STATUS_VALUES)].copy()
    return_df["_month_key"] = return_df["날짜"].apply(parse_korean_month)
    outbound_df["_month_key"] = outbound_df["날짜"].apply(parse_korean_month)
    outbound_qty_df = (
        outbound_df.groupby(["_month_key", "플랫폼", "브랜드"], dropna=False, sort=False)
        .agg(출고_수량=("수량_num", "sum"))
        .reset_index()
    )

    grouped_df = (
        return_df.groupby(["날짜", "_month_key", "플랫폼", "브랜드"], dropna=False, sort=False)
        .agg(
            반품_건수=("수량_num", "size"),
            반품_수량=("수량_num", lambda series: series.abs().sum()),
            반품_판매가=("판매가_num", lambda series: series.abs().sum()),
            반품_수입=("수입_num", lambda series: series.abs().sum()),
            반품_공급가=("공급가_num", lambda series: series.abs().sum()),
            반품_마진=("마진_num", lambda series: series.abs().sum()),
        )
        .reset_index()
        .merge(outbound_qty_df, on=["_month_key", "플랫폼", "브랜드"], how="left")
        .fillna({"출고_수량": 0})
        .drop(columns=["_month_key"])
        .rename(
            columns={
                "반품_건수": "반품 건수",
                "반품_수량": "반품 수량",
                "반품_판매가": "반품 판매가",
                "반품_수입": "반품 수입",
                "반품_공급가": "반품 공급가",
                "반품_마진": "반품 마진",
                "출고_수량": "출고 수량",
            }
        )
    )
    grouped_df["반품율"] = grouped_df["반품 수량"].div(grouped_df["출고 수량"].replace(0, pd.NA)).fillna(0)

    for col in NUMERIC_RETURN_BRAND_COLUMNS:
        grouped_df[col] = grouped_df[col].apply(format_int)

    grouped_df["_date_key"] = grouped_df["날짜"].apply(parse_korean_date_key)
    grouped_df = grouped_df.sort_values(
        by=["_date_key", "플랫폼", "반품 수량", "브랜드"],
        ascending=[False, True, False, True],
    ).drop(columns=["_date_key"])
    return grouped_df[RETURN_BRAND_COLUMNS].copy()


def build_return_brand_sku_df(source_df: pd.DataFrame, date_text: str) -> pd.DataFrame:
    return_df = source_df[
        source_df["내역"].isin(RETURN_STATUS_VALUES)
        & source_df["날짜"].eq(date_text)
        & source_df["브랜드"].ne("")
        & source_df["품번"].ne("")
    ].copy()
    if return_df.empty:
        return pd.DataFrame(columns=RETURN_BRAND_SKU_COLUMNS)

    grouped_df = (
        return_df.groupby(["날짜", "플랫폼", "브랜드", "품번"], dropna=False, sort=False)
        .agg(
            반품_수량=("수량_num", lambda series: series.abs().sum()),
            반품_판매가=("판매가_num", lambda series: series.abs().sum()),
            반품_수입=("수입_num", lambda series: series.abs().sum()),
            반품_공급가=("공급가_num", lambda series: series.abs().sum()),
            반품_마진=("마진_num", lambda series: series.abs().sum()),
        )
        .reset_index()
        .rename(
            columns={
                "반품_수량": "반품 수량",
                "반품_판매가": "반품 판매가",
                "반품_수입": "반품 수입",
                "반품_공급가": "반품 공급가",
                "반품_마진": "반품 마진",
            }
        )
    )
    outbound_df = source_df[
        ~source_df["내역"].isin(RETURN_STATUS_VALUES)
        & source_df["브랜드"].ne("")
        & source_df["품번"].ne("")
    ].copy()
    month_key = parse_korean_month(date_text)
    outbound_df = outbound_df[outbound_df["날짜"].apply(parse_korean_month).eq(month_key)].copy()
    outbound_qty_df = (
        outbound_df.groupby(["플랫폼", "브랜드", "품번"], dropna=False, sort=False)
        .agg(출고_수량=("수량_num", "sum"))
        .reset_index()
    )
    grouped_df = grouped_df.merge(outbound_qty_df, on=["플랫폼", "브랜드", "품번"], how="left").fillna({"출고_수량": 0})
    grouped_df["반품율"] = grouped_df["반품 수량"].div(grouped_df["출고_수량"].replace(0, pd.NA)).fillna(0)
    for col in ["반품 수량", "반품 판매가", "반품 수입", "반품 공급가", "반품 마진"]:
        grouped_df[col] = grouped_df[col].apply(format_int)
    grouped_df = grouped_df.sort_values(
        by=["브랜드", "반품 수량", "품번"],
        ascending=[True, False, True],
    ).reset_index(drop=True)
    return grouped_df[RETURN_BRAND_SKU_COLUMNS].copy()


def build_return_brand_sku_top10_df(source_df: pd.DataFrame, date_text: str, stock_lookup: dict[str, int]) -> pd.DataFrame:
    return_df = source_df[
        source_df["내역"].isin(RETURN_STATUS_VALUES)
        & source_df["날짜"].eq(date_text)
        & source_df["브랜드"].ne("")
        & source_df["품번"].ne("")
    ].copy()
    if return_df.empty:
        return pd.DataFrame()

    grouped_df = (
        return_df.groupby(["브랜드", "품번"], dropna=False, sort=False)
        .agg(
            총_수량=("수량_num", lambda series: series.abs().sum()),
            총_판매가=("판매가_num", lambda series: series.abs().sum()),
            총_마진=("마진_num", lambda series: series.abs().sum()),
        )
        .reset_index()
    )
    outbound_df = source_df[
        ~source_df["내역"].isin(RETURN_STATUS_VALUES)
        & source_df["브랜드"].ne("")
        & source_df["품번"].ne("")
    ].copy()
    month_key = parse_korean_month(date_text)
    outbound_df = outbound_df[outbound_df["날짜"].apply(parse_korean_month).eq(month_key)].copy()
    outbound_qty_df = (
        outbound_df.groupby(["브랜드", "품번"], dropna=False, sort=False)
        .agg(출고_수량=("수량_num", "sum"))
        .reset_index()
    )
    grouped_df = grouped_df.merge(outbound_qty_df, on=["브랜드", "품번"], how="left").fillna({"출고_수량": 0})
    grouped_df["반품율"] = grouped_df["총_수량"].div(grouped_df["출고_수량"].replace(0, pd.NA)).fillna(0)
    brand_names = list(dict.fromkeys(grouped_df["브랜드"].tolist()))

    ranked_map = {}
    for brand_name in brand_names:
        brand_df = grouped_df[grouped_df["브랜드"].eq(brand_name)].copy()
        ranked_map[brand_name] = brand_df.sort_values(
            by=["총_수량", "총_마진", "품번"],
            ascending=[False, False, True],
        ).head(10).reset_index(drop=True)

    return build_top10_grid_df(
        brand_names,
        ranked_map,
        stock_lookup,
        "return_brand_sku",
        quantity_header="반품 수량",
        rate_header="반품율",
        rate_column="반품율",
    )


def normalize_platform_summary_df(df: pd.DataFrame) -> pd.DataFrame:
    missing_columns = [col for col in PLATFORM_COLUMNS if col not in df.columns]
    if missing_columns:
        raise ValueError(f"플랫폼 마감 시트 필수 열이 없습니다: {missing_columns}")

    output_df = df[PLATFORM_COLUMNS].copy()
    output_df["날짜"] = output_df["날짜"].apply(clean_text)
    output_df["플랫폼"] = output_df["플랫폼"].apply(clean_text)
    output_df = output_df.drop_duplicates(subset=["날짜", "플랫폼"], keep="last").reset_index(drop=True)

    for col in NUMERIC_PLATFORM_COLUMNS:
        output_df[col] = to_number_series(output_df[col])

    output_df["마진율"] = output_df["마진"].div(output_df["판매가"].replace(0, pd.NA)).fillna(0)

    for col in NUMERIC_PLATFORM_COLUMNS:
        output_df[col] = output_df[col].apply(format_int)

    return output_df


def normalize_brand_summary_df(df: pd.DataFrame) -> pd.DataFrame:
    renamed_df = df.rename(columns={"공금가": "공급가"}).copy()
    missing_columns = [col for col in BRAND_COLUMNS if col not in renamed_df.columns]
    if missing_columns:
        raise ValueError(f"플랫폼 브랜드별 마감 시트 필수 열이 없습니다: {missing_columns}")

    output_df = renamed_df[BRAND_COLUMNS].copy()
    output_df["날짜"] = output_df["날짜"].apply(clean_text)
    output_df["플랫폼"] = output_df["플랫폼"].apply(clean_text)
    output_df["브랜드"] = output_df["브랜드"].apply(clean_text)
    output_df = output_df.drop_duplicates(subset=["날짜", "플랫폼", "브랜드"], keep="last").reset_index(drop=True)

    for col in NUMERIC_BRAND_COLUMNS:
        output_df[col] = to_number_series(output_df[col])

    output_df["마진율"] = output_df["마진"].div(output_df["판매가"].replace(0, pd.NA)).fillna(0)

    for col in NUMERIC_BRAND_COLUMNS:
        output_df[col] = output_df[col].apply(format_int)

    return output_df


def build_stock_lookup(df: pd.DataFrame) -> dict[str, int]:
    required_columns = ["품번", "현재고"]
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"현재고 시트 필수 열이 없습니다: {missing_columns}")

    output_df = df[required_columns].copy()
    output_df["품번"] = output_df["품번"].apply(clean_text)
    output_df = output_df[output_df["품번"] != ""].copy()
    output_df["현재고_num"] = to_number_series(output_df["현재고"])

    grouped_df = (
        output_df.groupby("품번", dropna=False, sort=False)["현재고_num"]
        .sum()
        .reset_index()
    )
    return {row["품번"]: format_int(row["현재고_num"]) for _, row in grouped_df.iterrows()}


def build_top10_grid_df(
    name_order: list[str],
    ranked_map: dict[str, pd.DataFrame],
    stock_lookup: dict[str, int],
    column_prefix: str,
    quantity_header: str = "판매 수량",
    rate_header: str = "마진율",
    rate_column: str = "마진율",
) -> pd.DataFrame:
    if not name_order:
        return pd.DataFrame()

    block_headers = ["순위", "품번", quantity_header, "현재고", rate_header]
    block_size = len(block_headers)
    row_block_width = (TOP_BLOCKS_PER_ROW * block_size) + ((TOP_BLOCKS_PER_ROW - 1) * TOP_BLOCK_SPACER_COLS)
    rows = []

    for chunk_start in range(0, len(name_order), TOP_BLOCKS_PER_ROW):
        chunk_names = name_order[chunk_start:chunk_start + TOP_BLOCKS_PER_ROW]
        title_row = []
        header_row = []

        for idx, block_name in enumerate(chunk_names):
            if idx > 0:
                title_row.extend([""] * TOP_BLOCK_SPACER_COLS)
                header_row.extend([""] * TOP_BLOCK_SPACER_COLS)
            title_row.extend([block_name] + [""] * (block_size - 1))
            header_row.extend(block_headers)

        title_row.extend([""] * (row_block_width - len(title_row)))
        header_row.extend([""] * (row_block_width - len(header_row)))
        rows.append(title_row)
        rows.append(header_row)

        for rank_idx in range(10):
            row_values = []
            for idx, block_name in enumerate(chunk_names):
                if idx > 0:
                    row_values.extend([""] * TOP_BLOCK_SPACER_COLS)
                block_df = ranked_map[block_name]
                if rank_idx < len(block_df):
                    row = block_df.iloc[rank_idx]
                    row_values.extend(
                        [
                            rank_idx + 1,
                            row["품번"],
                            format_int(row["총_수량"]),
                            stock_lookup.get(clean_text(row["품번"]), 0),
                            float(row[rate_column]),
                        ]
                    )
                else:
                    row_values.extend(["", "", "", "", ""])
            row_values.extend([""] * (row_block_width - len(row_values)))
            rows.append(row_values)

        if chunk_start + TOP_BLOCKS_PER_ROW < len(name_order):
            rows.append([""] * row_block_width)

    return pd.DataFrame(
        rows,
        columns=[f"{column_prefix}_{idx}" for idx in range(1, row_block_width + 1)],
    )


def build_best_top10_df(df: pd.DataFrame, stock_lookup: dict[str, int], report_date: str) -> pd.DataFrame:
    output_df = normalize_source_df(df)
    output_df = output_df[~output_df["내역"].isin(RETURN_STATUS_VALUES)].copy()
    output_df = output_df[output_df["날짜"].eq(report_date)].copy()
    output_df = output_df[output_df["품번"].ne("")].copy()

    if output_df.empty:
        return pd.DataFrame()

    grouped_df = (
        output_df.groupby(["날짜", "플랫폼", "품번"], dropna=False, sort=False)
        .agg(
            총_수량=("수량_num", "sum"),
            총_판매가=("판매가_num", "sum"),
            총_마진=("마진_num", "sum"),
        )
        .reset_index()
    )

    if grouped_df.empty:
        return pd.DataFrame()

    platform_names = list(dict.fromkeys(grouped_df["플랫폼"].tolist()))

    ranked_map = {}
    for platform_name in platform_names:
        platform_df = grouped_df[grouped_df["플랫폼"].eq(platform_name)].copy()
        platform_df = platform_df.sort_values(by=["총_수량", "총_마진", "품번"], ascending=[False, False, True]).head(10).reset_index(drop=True)
        platform_df["마진율"] = platform_df["총_마진"].div(platform_df["총_판매가"].replace(0, pd.NA)).fillna(0)
        ranked_map[platform_name] = platform_df
    return build_top10_grid_df(platform_names, ranked_map, stock_lookup, "best_top")


def build_brand_top10_df(df: pd.DataFrame, stock_lookup: dict[str, int], report_date: str) -> pd.DataFrame:
    output_df = normalize_source_df(df)
    output_df = output_df[~output_df["내역"].isin(RETURN_STATUS_VALUES)].copy()
    output_df = output_df[output_df["날짜"].eq(report_date)].copy()
    output_df = output_df[output_df["브랜드"].ne("") & output_df["품번"].ne("")].copy()

    if output_df.empty:
        return pd.DataFrame()

    grouped_df = (
        output_df.groupby(["날짜", "브랜드", "품번"], dropna=False, sort=False)
        .agg(
            총_수량=("수량_num", "sum"),
            총_판매가=("판매가_num", "sum"),
            총_마진=("마진_num", "sum"),
        )
        .reset_index()
    )

    if grouped_df.empty:
        return pd.DataFrame()

    brand_names = list(dict.fromkeys(grouped_df["브랜드"].tolist()))

    ranked_map = {}
    for brand_name in brand_names:
        brand_df = grouped_df[grouped_df["브랜드"].eq(brand_name)].copy()
        brand_df = brand_df.sort_values(by=["총_수량", "총_마진", "품번"], ascending=[False, False, True]).head(10).reset_index(drop=True)
        brand_df["마진율"] = brand_df["총_마진"].div(brand_df["총_판매가"].replace(0, pd.NA)).fillna(0)
        ranked_map[brand_name] = brand_df
    return build_top10_grid_df(brand_names, ranked_map, stock_lookup, "brand_top")


def build_brand_rank_lookup(brand_df: pd.DataFrame) -> pd.DataFrame:
    grouped_df = (
        brand_df.groupby(["날짜", "플랫폼"], dropna=False, sort=False)
        .agg(
            브랜드_수=("브랜드", "nunique"),
            브랜드목록=("브랜드", list),
            판매가목록=("판매가", list),
        )
        .reset_index()
    )

    top_brand_rows = []
    for row in grouped_df.itertuples(index=False):
        paired = list(zip(row.브랜드목록, row.판매가목록))
        paired.sort(key=lambda item: (-item[1], item[0]))
        top_brands = [brand for brand, _ in paired[:3]]
        while len(top_brands) < 3:
            top_brands.append("")

        top_brand_rows.append(
            {
                "날짜": row.날짜,
                "플랫폼": row.플랫폼,
                "브랜드 수": int(row.브랜드_수),
                "상위 브랜드 1": top_brands[0],
                "상위 브랜드 2": top_brands[1],
                "상위 브랜드 3": top_brands[2],
            }
        )

    return pd.DataFrame(top_brand_rows)


def build_daily_total_df(platform_df: pd.DataFrame) -> pd.DataFrame:
    total_df = (
        platform_df.groupby(["날짜"], dropna=False, sort=False)
        .agg(
            판매_수량=("판매 수량", "sum"),
            판매가=("판매가", "sum"),
            수입=("수입", "sum"),
            마진=("마진", "sum"),
        )
        .reset_index()
    )

    total_df["플랫폼"] = "합계"
    total_df["마진율"] = total_df["마진"].div(total_df["판매가"].replace(0, pd.NA)).fillna(0)
    total_df["브랜드 수"] = ""
    total_df["상위 브랜드 1"] = ""
    total_df["상위 브랜드 2"] = ""
    total_df["상위 브랜드 3"] = ""
    total_df = total_df.rename(columns={"판매_수량": "판매 수량"})

    for col in NUMERIC_PLATFORM_COLUMNS:
        total_df[col] = total_df[col].apply(format_int)

    return total_df[REPORT_COLUMNS]


def build_report_df(platform_df: pd.DataFrame, brand_df: pd.DataFrame) -> pd.DataFrame:
    brand_rank_df = build_brand_rank_lookup(brand_df)
    report_df = platform_df.merge(brand_rank_df, on=["날짜", "플랫폼"], how="left")

    for col in ["브랜드 수", "상위 브랜드 1", "상위 브랜드 2", "상위 브랜드 3"]:
        if col not in report_df.columns:
            report_df[col] = ""

    report_df["브랜드 수"] = to_number_series(report_df["브랜드 수"]).apply(format_int)
    report_df["상위 브랜드 1"] = report_df["상위 브랜드 1"].fillna("")
    report_df["상위 브랜드 2"] = report_df["상위 브랜드 2"].fillna("")
    report_df["상위 브랜드 3"] = report_df["상위 브랜드 3"].fillna("")

    report_df = report_df[REPORT_COLUMNS]
    total_df = build_daily_total_df(platform_df)
    report_df = pd.concat([report_df, total_df], ignore_index=True)

    report_df["_date_key"] = report_df["날짜"].apply(parse_korean_date_key)
    report_df["_platform_order"] = report_df["플랫폼"].apply(lambda value: 1 if value == "합계" else 0)
    report_df = report_df.sort_values(by=["_date_key", "_platform_order", "플랫폼"], ascending=[False, True, True]).reset_index(drop=True)
    report_df = report_df.drop(columns=["_date_key", "_platform_order"])

    return report_df


def pad_row(values: list) -> list:
    row = list(values)
    if len(row) < VERTICAL_REPORT_MAX_COLS:
        row.extend([""] * (VERTICAL_REPORT_MAX_COLS - len(row)))
    return row[:VERTICAL_REPORT_MAX_COLS]


def build_vertical_report_df(
    platform_df: pd.DataFrame,
    brand_df: pd.DataFrame,
    monthly_platform_df: pd.DataFrame | None = None,
    kpi_platform_df: pd.DataFrame | None = None,
    return_platform_df: pd.DataFrame | None = None,
    return_brand_df: pd.DataFrame | None = None,
    report_dates: list[str] | None = None,
) -> pd.DataFrame:
    rows = []
    monthly_source_df = monthly_platform_df.copy() if monthly_platform_df is not None else platform_df.copy()
    kpi_source_df = kpi_platform_df.copy() if kpi_platform_df is not None else platform_df.copy()
    return_platform_source_df = return_platform_df.copy() if return_platform_df is not None else pd.DataFrame(columns=RETURN_PLATFORM_COLUMNS)
    return_brand_source_df = return_brand_df.copy() if return_brand_df is not None else pd.DataFrame(columns=RETURN_BRAND_COLUMNS)
    if report_dates is None:
        report_date_values = set(platform_df["날짜"].dropna().unique())
        if kpi_source_df is not None and "날짜" in kpi_source_df.columns:
            report_date_values.update(kpi_source_df["날짜"].dropna().unique())
        if return_platform_source_df is not None and "날짜" in return_platform_source_df.columns:
            report_date_values.update(return_platform_source_df["날짜"].dropna().unique())
        sorted_dates = sorted(report_date_values, key=parse_korean_date_key, reverse=True)
    else:
        sorted_dates = report_dates

    for date_text in sorted_dates:
        date_platform_df = platform_df[platform_df["날짜"].eq(date_text)].copy()
        date_kpi_platform_df = kpi_source_df[kpi_source_df["날짜"].eq(date_text)].copy()
        date_brand_df = brand_df[brand_df["날짜"].eq(date_text)].copy()
        current_month = parse_korean_month(date_text)
        month_platform_df = monthly_source_df[monthly_source_df["날짜"].apply(parse_korean_month).eq(current_month)].copy()
        date_return_platform_df = return_platform_source_df[return_platform_source_df["날짜"].eq(date_text)].copy()
        date_return_brand_df = return_brand_source_df[return_brand_source_df["날짜"].eq(date_text)].copy()
        month_return_platform_df = return_platform_source_df[
            return_platform_source_df["날짜"].apply(parse_korean_month).eq(current_month)
        ].copy()
        month_return_platform_summary_df = pd.DataFrame(columns=RETURN_PLATFORM_COLUMNS)
        if not month_return_platform_df.empty:
            month_return_platform_summary_df = (
                month_return_platform_df.groupby(["플랫폼"], dropna=False, sort=False)
                .agg(
                    반품_판매가=("반품 판매가", "sum"),
                    반품_수입=("반품 수입", "sum"),
                    반품_마진=("반품 마진", "sum"),
                    반품_수량=("반품 수량", "sum"),
                    출고_수량=("출고 수량", "max"),
                )
                .reset_index()
                .rename(
                    columns={
                        "반품_판매가": "판매가",
                        "반품_수입": "수입",
                        "반품_마진": "마진",
                        "반품_수량": "판매 수량",
                        "출고_수량": "출고 수량",
                    }
                )
            )
            month_return_platform_summary_df["반품율"] = month_return_platform_summary_df["판매 수량"].div(
                month_return_platform_summary_df["출고 수량"].replace(0, pd.NA)
            ).fillna(0)
        month_platform_summary_df = (
            month_platform_df.groupby(["플랫폼"], dropna=False, sort=False)[["판매가", "수입", "마진", "판매 수량"]]
            .sum()
            .reset_index()
        )
        month_platform_summary_df["마진율"] = month_platform_summary_df.apply(
            lambda row: format_percent(row["마진"] / row["판매가"]) if row["판매가"] else "0%",
            axis=1,
        )
        total_sales = int(date_kpi_platform_df["판매가"].sum())
        total_income = int(date_kpi_platform_df["수입"].sum())
        total_margin = int(date_kpi_platform_df["마진"].sum())
        total_margin_rate = format_percent(total_margin / total_sales) if total_sales else "0%"
        total_qty = int(date_kpi_platform_df["판매 수량"].sum())
        month_total_sales = int(month_platform_df["판매가"].sum())
        month_total_income = int(month_platform_df["수입"].sum())
        month_total_margin = int(month_platform_df["마진"].sum())
        month_total_margin_rate = format_percent(month_total_margin / month_total_sales) if month_total_sales else "0%"

        rows.append(pad_row([f"{date_text} 일마감"]))
        rows.append(pad_row(["당일 KPI", "", "", "", "", "월 누적 KPI", "", ""]))
        rows.append(pad_row(["매출", "정산", "마진", "마진율", "판매 수량", "정산", "마진", "마진율"]))
        rows.append(pad_row([total_sales, total_income, total_margin, total_margin_rate, total_qty, month_total_income, month_total_margin, month_total_margin_rate]))
        rows.append(pad_row([]))
        rows.append(pad_row(["플랫폼 월 누적 요약"]))
        rows.append(pad_row(["플랫폼", "매출", "정산", "마진", "마진율", "판매 수량", "매출 비중", "마진 비중"]))

        for _, row in month_platform_summary_df.sort_values(by=["플랫폼"]).iterrows():
            month_sales_ratio = format_percent(row["판매가"] / month_total_sales) if month_total_sales else "0%"
            month_margin_ratio = format_percent(row["마진"] / month_total_margin) if month_total_margin else "0%"
            rows.append(
                pad_row(
                    [
                        row["플랫폼"],
                        row["판매가"],
                        row["수입"],
                        row["마진"],
                        row["마진율"],
                        row["판매 수량"],
                        month_sales_ratio,
                        month_margin_ratio,
                    ]
                )
            )
        rows.append(pad_row([]))

        if not month_return_platform_summary_df.empty:
            rows.append(pad_row(["플랫폼 반품 월 누적 요약"]))
            rows.append(pad_row(["플랫폼", "매출", "정산", "마진", "반품율", "판매 수량", "매출 비중", "마진 비중"]))
            for _, row in month_return_platform_summary_df.sort_values(by=["플랫폼"]).iterrows():
                return_sales_ratio = format_percent(row["판매가"] / month_total_sales) if month_total_sales else "0%"
                return_margin_ratio = format_percent(row["마진"] / month_total_margin) if month_total_margin else "0%"
                rows.append(
                    pad_row(
                        [
                            row["플랫폼"],
                            -row["판매가"],
                            -row["수입"],
                            -row["마진"],
                            format_percent(row["반품율"]),
                            -row["판매 수량"],
                            return_sales_ratio,
                            return_margin_ratio,
                        ]
                    )
                )
            rows.append(pad_row([]))

        if not date_platform_df.empty:
            rows.append(pad_row(["플랫폼 일별 요약"]))
            rows.append(pad_row(["플랫폼", "매출", "정산", "마진", "마진율", "판매 수량", "매출 비중", "마진 비중"]))

            for _, row in date_platform_df.sort_values(by=["플랫폼"]).iterrows():
                sales_ratio = format_percent(row["판매가"] / total_sales) if total_sales else "0%"
                margin_ratio = format_percent(row["마진"] / total_margin) if total_margin else "0%"
                rows.append(
                    pad_row(
                        [
                            row["플랫폼"],
                            row["판매가"],
                            row["수입"],
                            row["마진"],
                            row["마진율"],
                            row["판매 수량"],
                            sales_ratio,
                            margin_ratio,
                        ]
                    )
                )
            rows.append(pad_row([]))

        if not date_return_platform_df.empty:
            rows.append(pad_row(["플랫폼 반품 일별 요약"]))
            rows.append(pad_row(["플랫폼", "매출", "정산", "마진", "반품율", "판매 수량", "매출 비중", "마진 비중"]))
            for _, row in date_return_platform_df.sort_values(by=["플랫폼"]).iterrows():
                return_sales_ratio = format_percent(row["반품 판매가"] / total_sales) if total_sales else "0%"
                return_margin_ratio = format_percent(row["반품 마진"] / total_margin) if total_margin else "0%"
                rows.append(
                    pad_row(
                        [
                            row["플랫폼"],
                            -row["반품 판매가"],
                            -row["반품 수입"],
                            -row["반품 마진"],
                            format_percent(row["반품율"]),
                            -row["반품 수량"],
                            return_sales_ratio,
                            return_margin_ratio,
                        ]
                    )
                )
            rows.append(pad_row([]))

        for platform_name in sorted(date_platform_df["플랫폼"].dropna().unique()):
            platform_brand_df = date_brand_df[date_brand_df["플랫폼"].eq(platform_name)].copy()
            if platform_brand_df.empty:
                continue

            rows.append(pad_row([f"{platform_name} 브랜드 요약"]))
            rows.append(pad_row(["브랜드", "매출", "정산", "마진", "마진율", "수량", "매출 비중", "마진 비중"]))

            platform_brand_df = platform_brand_df.sort_values(by=["판매가", "브랜드"], ascending=[False, True])
            platform_total_sales = int(platform_brand_df["판매가"].sum())
            platform_total_income = int(platform_brand_df["수입"].sum())
            platform_total_margin = int(platform_brand_df["마진"].sum())
            platform_total_qty = int(platform_brand_df["수량"].sum())
            platform_total_margin_rate = format_percent(platform_total_margin / platform_total_sales) if platform_total_sales else "0%"
            for _, brand_row in platform_brand_df.iterrows():
                brand_sales_ratio = format_percent(brand_row["판매가"] / platform_total_sales) if platform_total_sales else "0%"
                brand_margin_ratio = format_percent(brand_row["마진"] / platform_total_margin) if platform_total_margin else "0%"
                rows.append(
                    pad_row(
                        [
                            brand_row["브랜드"],
                            brand_row["판매가"],
                            brand_row["수입"],
                            brand_row["마진"],
                            brand_row["마진율"],
                            brand_row["수량"],
                            brand_sales_ratio,
                            brand_margin_ratio,
                        ]
                    )
                )

            rows.append(pad_row([]))

        if not date_return_brand_df.empty:
            for platform_name in sorted(date_return_brand_df["플랫폼"].dropna().unique()):
                platform_return_brand_df = date_return_brand_df[date_return_brand_df["플랫폼"].eq(platform_name)].copy()
                if platform_return_brand_df.empty:
                    continue

                rows.append(pad_row([f"{platform_name} 반품 브랜드 요약"]))
                rows.append(pad_row(["브랜드", "매출", "정산", "마진", "반품율", "수량", "매출 비중", "마진 비중"]))
                return_brand_total_sales = int(platform_return_brand_df["반품 판매가"].sum())
                return_brand_total_margin = int(platform_return_brand_df["반품 마진"].sum())
                for _, row in platform_return_brand_df.sort_values(by=["반품 판매가", "브랜드"], ascending=[False, True]).iterrows():
                    brand_return_sales_ratio = format_percent(row["반품 판매가"] / return_brand_total_sales) if return_brand_total_sales else "0%"
                    brand_return_margin_ratio = format_percent(row["반품 마진"] / return_brand_total_margin) if return_brand_total_margin else "0%"
                    rows.append(
                        pad_row(
                            [
                                row["브랜드"],
                                -row["반품 판매가"],
                                -row["반품 수입"],
                                -row["반품 마진"],
                                format_percent(row["반품율"]),
                                -row["반품 수량"],
                                brand_return_sales_ratio,
                                brand_return_margin_ratio,
                            ]
                        )
                    )
                rows.append(pad_row([]))

        rows.append(pad_row([]))

    return pd.DataFrame(rows, columns=[f"col_{idx}" for idx in range(1, VERTICAL_REPORT_MAX_COLS + 1)])


def apply_return_rich_text(worksheet):
    for row in worksheet.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str) or "반품" not in cell.value or "반품율" in cell.value:
                continue

            text = cell.value
            rich_text = CellRichText()
            normal_font = InlineFont(rFont=cell.font.name, sz=cell.font.sz, b=cell.font.bold)
            red_font = InlineFont(rFont=cell.font.name, sz=cell.font.sz, b=cell.font.bold, color="FF0000")
            start = 0
            while True:
                idx = text.find("반품", start)
                if idx == -1:
                    if start < len(text):
                        rich_text.append(TextBlock(normal_font, text[start:]))
                    break
                if idx > start:
                    rich_text.append(TextBlock(normal_font, text[start:idx]))
                rich_text.append(TextBlock(red_font, "반품"))
                start = idx + 2
            cell.value = rich_text


def style_report_sheet(worksheet):
    base_font_size = 12
    header_font_size = 12
    section_font_size = 12
    title_font_size = 14
    title_fill = PatternFill("solid", fgColor="E8F1E9")
    section_fill = PatternFill("solid", fgColor="F3F6F4")
    header_fill = PatternFill("solid", fgColor="DCE6DC")
    monthly_fill = PatternFill("solid", fgColor="E6EFE7")
    thin_gray = Side(style="thin", color="D9D9D9")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    center = Alignment(horizontal="center", vertical="center")

    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A5"
    worksheet.column_dimensions["A"].width = 28
    worksheet.column_dimensions["B"].width = 16
    worksheet.column_dimensions["C"].width = 16
    worksheet.column_dimensions["D"].width = 14
    worksheet.column_dimensions["E"].width = 14
    worksheet.column_dimensions["F"].width = 14
    worksheet.column_dimensions["G"].width = 14
    worksheet.column_dimensions["H"].width = 14

    max_row = worksheet.max_row
    max_col = worksheet.max_column
    percent_headers = {"마진율", "총 마진율", "매출 비중", "마진 비중", "반품율"}
    integer_headers = {
        "매출", "정산", "정산 예정액", "수입", "마진", "판매 수량", "총 판매 수량", "총 판매가", "총 마진", "판매가", "수량",
        "반품 건수", "반품 수량", "반품 판매가", "반품 수입", "반품 마진", "출고 수량",
    }
    active_headers = {}

    def coerce_excel_value(cell):
        value = cell.value
        if not isinstance(value, str):
            return

        text = value.strip()
        if text == "":
            return

        if re.fullmatch(r"-?\d+(?:\.\d+)?%", text):
            cell.value = float(text.replace("%", "")) / 100
            cell.number_format = "0.00%"
            return

        if re.fullmatch(r"-?\d[\d,]*", text):
            cell.value = int(text.replace(",", ""))
            cell.number_format = "#,##0"
            return

        if re.fullmatch(r"-?\d[\d,]*\.\d+", text):
            cell.value = float(text.replace(",", ""))
            cell.number_format = "#,##0.00"
            return

    for row in range(1, max_row + 1):
        worksheet.row_dimensions[row].height = 22
        first_value = worksheet.cell(row=row, column=1).value
        second_value = worksheet.cell(row=row, column=2).value

        is_blank = all(
            (worksheet.cell(row=row, column=col).value in (None, ""))
            for col in range(1, max_col + 1)
        )
        if is_blank:
            worksheet.row_dimensions[row].height = 10
            continue

        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=row, column=col)
            coerce_excel_value(cell)
            cell.alignment = center
            cell.font = Font(size=base_font_size)

        if first_value and str(first_value).endswith("일마감"):
            worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
            title_cell = worksheet.cell(row=row, column=1)
            title_cell.fill = title_fill
            title_cell.font = Font(bold=True, size=title_font_size)
            title_cell.border = border
            worksheet.row_dimensions[row].height = 28
            continue

        if (
            isinstance(first_value, str)
            and (
                (str(first_value).startswith("플랫폼 ") and str(first_value).endswith("요약"))
                or str(first_value).endswith("브랜드 요약")
                or str(first_value).startswith("플랫폼 반품 ")
            )
        ):
            worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
            section_cell = worksheet.cell(row=row, column=1)
            section_cell.fill = section_fill
            section_cell.font = Font(bold=True, size=section_font_size)
            section_cell.border = border
            active_headers = {}
            continue

        if first_value == "당일 KPI":
            worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            worksheet.merge_cells(start_row=row, start_column=6, end_row=row, end_column=8)
            day_cell = worksheet.cell(row=row, column=1)
            month_cell = worksheet.cell(row=row, column=6)
            day_cell.fill = section_fill
            day_cell.font = Font(bold=True, size=header_font_size)
            day_cell.border = border
            day_cell.alignment = center
            month_cell.fill = monthly_fill
            month_cell.font = Font(bold=True, size=header_font_size)
            month_cell.border = border
            month_cell.alignment = center
            for col in range(2, 6):
                cell = worksheet.cell(row=row, column=col)
                cell.fill = section_fill
                cell.border = border
            for col in range(7, 9):
                cell = worksheet.cell(row=row, column=col)
                cell.fill = monthly_fill
                cell.border = border
            worksheet.row_dimensions[row].height = 20
            active_headers = {}
            continue

        if first_value == "매출":
            active_headers = {col: str(worksheet.cell(row=row, column=col).value or "").strip() for col in range(1, max_col + 1)}
            for col in range(1, 5 + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.fill = header_fill
                cell.font = Font(bold=True, size=header_font_size)
                cell.border = border
            for col in range(6, 8 + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.fill = monthly_fill
                cell.font = Font(bold=True, size=header_font_size)
                cell.border = border
            continue

        if row > 1 and worksheet.cell(row=row - 1, column=1).value == "매출":
            for col in range(1, 6):
                cell = worksheet.cell(row=row, column=col)
                cell.border = border
                header_name = active_headers.get(col, "")
                if header_name in integer_headers and isinstance(cell.value, Number):
                    cell.number_format = "#,##0"
                if header_name in percent_headers and isinstance(cell.value, Number):
                    cell.number_format = "0.00%"
            for col in range(6, 9):
                cell = worksheet.cell(row=row, column=col)
                cell.border = border
                header_name = active_headers.get(col, "")
                if header_name in integer_headers and isinstance(cell.value, Number):
                    cell.number_format = "#,##0"
                if header_name in percent_headers and isinstance(cell.value, Number):
                    cell.number_format = "0.00%"
            worksheet.row_dimensions[row].height = 24
            continue

        if first_value in {"플랫폼", "브랜드"}:
            active_headers = {col: str(worksheet.cell(row=row, column=col).value or "").strip() for col in range(1, max_col + 1)}
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=row, column=col)
                if cell.value in (None, ""):
                    continue
                cell.fill = header_fill
                cell.font = Font(bold=True, size=header_font_size)
                cell.border = border
            continue

        if second_value not in (None, ""):
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=row, column=col)
                if cell.value in (None, ""):
                    continue
                cell.border = border
                header_name = active_headers.get(col, "")
                if header_name in percent_headers and isinstance(cell.value, Number):
                    cell.number_format = "0.00%"
                elif header_name in integer_headers and isinstance(cell.value, Number):
                    cell.number_format = "#,##0"

    apply_return_rich_text(worksheet)


def style_summary_sheet(worksheet):
    base_font_size = 12
    header_font_size = 12
    header_fill = PatternFill("solid", fgColor="DCE6DC")
    thin_gray = Side(style="thin", color="D9D9D9")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    center = Alignment(horizontal="center", vertical="center")

    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"

    percent_headers = {"마진율", "매출 비중", "마진 비중", "반품율"}
    integer_headers = {"주문건수", "판매 수량", "수량", "매출", "판매가", "수수료", "정산", "정산 예정액", "수입", "공급가", "마진", "반품 건수", "반품 수량", "반품 판매가", "반품 수입", "반품 공급가", "반품 마진", "출고 수량"}
    header_map = {
        col: str(worksheet.cell(row=1, column=col).value or "").strip()
        for col in range(1, worksheet.max_column + 1)
    }

    for column_cells in worksheet.columns:
        column_letter = column_cells[0].column_letter
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 4, 12), 22)

    for row in range(1, worksheet.max_row + 1):
        worksheet.row_dimensions[row].height = 22
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.alignment = center
            cell.border = border
            cell.font = Font(size=base_font_size)

            if row == 1:
                cell.fill = header_fill
                cell.font = Font(bold=True, size=header_font_size)
                continue

            header_name = header_map.get(col, "")
            if isinstance(cell.value, str):
                text = cell.value.strip()
                if re.fullmatch(r"-?\d+(?:\.\d+)?%", text):
                    cell.value = float(text.replace("%", "")) / 100
                elif re.fullmatch(r"-?\d[\d,]*", text):
                    cell.value = int(text.replace(",", ""))
                elif re.fullmatch(r"-?\d[\d,]*\.\d+", text):
                    cell.value = float(text.replace(",", ""))

            if header_name in percent_headers and isinstance(cell.value, Number):
                cell.number_format = "0.00%"
            elif header_name in integer_headers and isinstance(cell.value, Number):
                cell.number_format = "#,##0"

    apply_return_rich_text(worksheet)


def style_best_top_sheet(worksheet):
    base_font_size = 12
    header_font_size = 12
    title_font_size = 12
    header_fill = PatternFill("solid", fgColor="DCE6DC")
    section_fill = PatternFill("solid", fgColor="F3F6F4")
    thin_gray = Side(style="thin", color="D9D9D9")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    center = Alignment(horizontal="center", vertical="center")

    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A3"

    if worksheet.max_column == 0:
        return

    block_size = len(BEST_TOP_BLOCK_HEADERS)
    stride = block_size + TOP_BLOCK_SPACER_COLS
    for col in range(1, worksheet.max_column + 1):
        mod = (col - 1) % stride
        if mod == block_size:
            worksheet.column_dimensions[worksheet.cell(row=1, column=col).column_letter].width = 9
        elif mod == 0:
            worksheet.column_dimensions[worksheet.cell(row=1, column=col).column_letter].width = 10
        elif mod == 1:
            worksheet.column_dimensions[worksheet.cell(row=1, column=col).column_letter].width = 20
        elif mod == 3:
            worksheet.column_dimensions[worksheet.cell(row=1, column=col).column_letter].width = 14
        elif mod == 4:
            worksheet.column_dimensions[worksheet.cell(row=1, column=col).column_letter].width = 14
        else:
            worksheet.column_dimensions[worksheet.cell(row=1, column=col).column_letter].width = 14

    header_rows = set()
    title_rows = set()
    for row in range(1, worksheet.max_row + 1):
        row_values = [str(worksheet.cell(row=row, column=col).value or "").strip() for col in range(1, worksheet.max_column + 1)]
        if "순위" in row_values:
            header_rows.add(row)
            if row > 1:
                title_rows.add(row - 1)

    for title_row in sorted(title_rows):
        for start_col in range(1, worksheet.max_column + 1, stride):
            end_col = start_col + block_size - 1
            if end_col > worksheet.max_column:
                break
            title_value = str(worksheet.cell(row=title_row, column=start_col).value or "").strip()
            if title_value == "":
                continue
            worksheet.merge_cells(start_row=title_row, start_column=start_col, end_row=title_row, end_column=end_col)
            title_cell = worksheet.cell(row=title_row, column=start_col)
            title_cell.fill = section_fill
            title_cell.font = Font(bold=True, size=title_font_size)
            title_cell.alignment = center
            title_cell.border = border
            for col in range(start_col + 1, end_col + 1):
                worksheet.cell(row=title_row, column=col).border = border

    for row in range(1, worksheet.max_row + 1):
        worksheet.row_dimensions[row].height = 22
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row, column=col)
            mod = (col - 1) % stride
            if mod == block_size:
                cell.border = Border()
                continue
            cell_value = cell.value
            if cell_value is None or (isinstance(cell_value, str) and cell_value.strip() == ""):
                cell.border = Border()
                continue
            cell.alignment = center
            cell.border = border
            cell.font = Font(size=base_font_size)

            if row in header_rows:
                cell.fill = header_fill
                cell.font = Font(bold=True, size=header_font_size)
                continue
            if row in title_rows:
                continue

            header_row = max((r for r in header_rows if r < row), default=None)
            header_name = str(worksheet.cell(row=header_row, column=col).value or "").strip() if header_row else ""
            if header_name in {"순위", "판매 수량", "반품 수량", "현재고"} and isinstance(cell.value, Number):
                cell.number_format = "#,##0"
            elif header_name in {"마진율", "반품율"} and isinstance(cell.value, Number):
                cell.number_format = "0.00%"


def save_report(
    vertical_report_df: pd.DataFrame,
    best_top10_df: pd.DataFrame,
    brand_top10_df: pd.DataFrame,
    return_brand_sku_df: pd.DataFrame,
):
    output_dir = LIST_DIR
    output_dir.mkdir(parents=True, exist_ok=True)
    today_stamp = time.strftime("%m%d")
    xlsx_path = output_dir / f"온라인 일마감 {today_stamp}.xlsx"
    report_sheet_name = "일 마감"

    try:
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            vertical_report_df.to_excel(writer, index=False, header=False, sheet_name=report_sheet_name)
            best_top10_df.to_excel(writer, index=False, header=False, sheet_name="BEST TOP 10")
            brand_top10_df.to_excel(writer, index=False, header=False, sheet_name="BRAND TOP 10")
            return_brand_sku_df.to_excel(writer, index=False, header=False, sheet_name="브랜드 반품 SKU")

        workbook = load_workbook(xlsx_path)
        style_report_sheet(workbook[report_sheet_name])
        style_best_top_sheet(workbook["BEST TOP 10"])
        style_best_top_sheet(workbook["BRAND TOP 10"])
        style_best_top_sheet(workbook["브랜드 반품 SKU"])
        workbook.save(xlsx_path)
        log(f"엑셀 보고서 저장 완료 -> {xlsx_path}")
        return
    except Exception as e:
        log(f"엑셀 저장 실패 -> {e}")

    report_csv_path = output_dir / f"온라인 일마감 {today_stamp}.csv"
    best_top_csv_path = output_dir / f"온라인 일마감 {today_stamp} BEST TOP 10.csv"
    brand_top_csv_path = output_dir / f"온라인 일마감 {today_stamp} BRAND TOP 10.csv"
    return_brand_sku_csv_path = output_dir / f"온라인 일마감 {today_stamp} 브랜드 반품 SKU.csv"

    vertical_report_df.to_csv(report_csv_path, index=False, header=False, encoding="utf-8-sig")
    best_top10_df.to_csv(best_top_csv_path, index=False, header=False, encoding="utf-8-sig")
    brand_top10_df.to_csv(brand_top_csv_path, index=False, header=False, encoding="utf-8-sig")
    return_brand_sku_df.to_csv(return_brand_sku_csv_path, index=False, encoding="utf-8-sig")
    log(f"CSV 보고서 저장 완료 -> {report_csv_path}")


def main():
    start_time = time.perf_counter()
    log("원본 플랫폼/반품 시트 다운로드 시작")
    source_raw_df = read_combined_platform_source_df()
    stock_raw_df = read_csv_from_url(CURRENT_STOCK_URL)

    log("원본 플랫폼/반품 시트 기준 마감 계산 시작")
    normalized_source_df = normalize_source_df(source_raw_df)
    outbound_source_df = normalized_source_df[~normalized_source_df["내역"].isin(RETURN_STATUS_VALUES)].copy()
    normalized_platform_df = build_platform_summary_from_source_df(outbound_source_df)
    normalized_brand_df = build_brand_summary_from_source_df(outbound_source_df)
    return_platform_df = build_return_platform_summary_df(normalized_source_df)
    return_brand_df = build_return_brand_summary_df(normalized_source_df)
    normalized_platform_with_return_df = build_platform_summary_from_source_df(normalized_source_df)
    today_platform_df = filter_today_date_df(normalized_platform_df)
    if today_platform_df.empty:
        today_platform_with_return_df = filter_today_date_df(normalized_platform_with_return_df)
        if today_platform_with_return_df.empty:
            raise ValueError(f"원본 플랫폼/반품 시트에서 오늘 날짜({today_date_text()}) 데이터를 찾을 수 없습니다.")
        log(f"오늘 날짜({today_date_text()}) 출고 데이터 없음 -> 반품 데이터 기준으로 마감 생성")
        report_date = str(today_platform_with_return_df.iloc[0]["날짜"])
    else:
        report_date = str(today_platform_df.iloc[0]["날짜"])
    platform_df = normalized_platform_df[normalized_platform_df["날짜"].eq(report_date)].copy().reset_index(drop=True)
    brand_df = normalized_brand_df[normalized_brand_df["날짜"].eq(report_date)].copy().reset_index(drop=True)
    kpi_platform_df = normalized_platform_with_return_df[
        normalized_platform_with_return_df["날짜"].eq(report_date)
    ].copy().reset_index(drop=True)
    vertical_report_df = build_vertical_report_df(
        platform_df,
        brand_df,
        normalized_platform_with_return_df,
        kpi_platform_df,
        return_platform_df,
        return_brand_df,
        [report_date],
    )
    stock_lookup = build_stock_lookup(stock_raw_df)
    best_top10_df = build_best_top10_df(source_raw_df, stock_lookup, report_date)
    brand_top10_df = build_brand_top10_df(source_raw_df, stock_lookup, report_date)
    return_brand_sku_df = build_return_brand_sku_top10_df(normalized_source_df, report_date, stock_lookup)
    save_report(vertical_report_df, best_top10_df, brand_top10_df, return_brand_sku_df)

    elapsed = time.perf_counter() - start_time
    log(f"완료 ({elapsed:.2f}초)")


if __name__ == "__main__":
    main()
