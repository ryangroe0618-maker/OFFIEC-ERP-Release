# -*- coding: utf-8 -*-
import os
import re
import shutil
import time
import zipfile
import tempfile
import traceback
from datetime import datetime
from io import StringIO, BytesIO
from typing import Optional, Dict

import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from pypdf import PdfReader, PdfWriter

from office_erp_paths import CLASSIFY_WAIT_DIR, KASHION_DIR, LIST_DIR, PDF_DIR, ROE_DIR

# -------------------------------------------------------
# [설정] 경로 및 구글 시트 URL
# -------------------------------------------------------
# 코드/가상환경이 있는 폴더(유지)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 대기 폴더: 바탕화면/Roe/구분 대기
WAIT_FOLDER = str(CLASSIFY_WAIT_DIR)

# KASHION 폴더: 바탕화면/Roe/KASHION
KASHION_ROOT = str(KASHION_DIR)

# 엑셀 저장 폴더: 바탕화면/Roe/LIST
EXCEL_TARGET_FOLDER = str(LIST_DIR)

# 통합 PDF 저장 폴더: 바탕화면/Roe/PDF
PDF_TARGET_FOLDER = str(PDF_DIR)

URLS = {
    "classify": "https://docs.google.com/spreadsheets/d/e/2PACX-1vTU9VO_90AH04WGk0CpddbLa5JHYZg4kiiG0UaHWMaGw82-hVcf5s-XcJtSGeErqMJYP22uy5jqdQ7a/pub?gid=223803252&single=true&output=csv",
    "rename":   "https://docs.google.com/spreadsheets/d/e/2PACX-1vTU9VO_90AH04WGk0CpddbLa5JHYZg4kiiG0UaHWMaGw82-hVcf5s-XcJtSGeErqMJYP22uy5jqdQ7a/pub?gid=1716604860&single=true&output=csv",
    "excel":    "https://docs.google.com/spreadsheets/d/e/2PACX-1vTU9VO_90AH04WGk0CpddbLa5JHYZg4kiiG0UaHWMaGw82-hVcf5s-XcJtSGeErqMJYP22uy5jqdQ7a/pub?gid=2020107466&single=true&output=csv",
}

CLASSIFY_SUBFOLDER = "KSJD"
EXCLUDE_FOLDER_NAME = "완료"

# ✅ 노스페이스 하위로 넣고 싶은 매장명(정확히 매칭되는 이름들)
NORTHFACE_SUBSTORES = {"구월", "스퀘어원", "부천"}

# ✅ 고정 범위 (0-index, end exclusive)
BRANDS_FIXED = [
    {"name": "노스페이스", "slice": (0, 11),  "startrow": 1, "startcol": 1, "header_fill_until": "수량"},
    {"name": "휠라",       "slice": (12, 19), "startrow": 1, "startcol": 1, "header_fill_until": "수량"},
    {"name": "푸마",       "slice": (20, 26), "startrow": 1, "startcol": 1, "header_fill_until": "수량"},
]

# -------------------------------------------------------
# 유틸
# -------------------------------------------------------
SESSION = requests.Session()
CSV_RETRY_COUNT = 3
CSV_RETRY_DELAY = 1.5

def ensure_folder(path: str) -> None:
    os.makedirs(path, exist_ok=True)

def clean_text(x) -> str:
    s = "" if x is None else str(x)
    s = s.replace('"', "").replace("\t", "").replace("\r", "").replace("\n", "")
    s = s.strip()
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    return s

def load_csv_from_url(url: str, label: str = "CSV") -> pd.DataFrame:
    last_error = None

    for attempt in range(1, CSV_RETRY_COUNT + 1):
        try:
            r = SESSION.get(url, timeout=30)
            r.raise_for_status()
            content = r.content

            for enc in ("utf-8-sig", "utf-8", "cp949"):
                try:
                    text = content.decode(enc)
                    return pd.read_csv(StringIO(text), dtype=str, keep_default_na=False)
                except Exception:
                    continue

            return pd.read_csv(BytesIO(content), dtype=str, keep_default_na=False)
        except requests.HTTPError as exc:
            last_error = exc
            status_code = exc.response.status_code if exc.response is not None else None
            if status_code not in {429, 500, 502, 503, 504} or attempt == CSV_RETRY_COUNT:
                break
            print(
                f"[WARN] {label} 다운로드 재시도 {attempt}/{CSV_RETRY_COUNT} "
                f"(HTTP {status_code})"
            )
            time.sleep(CSV_RETRY_DELAY)
        except requests.RequestException as exc:
            last_error = exc
            if attempt == CSV_RETRY_COUNT:
                break
            print(f"[WARN] {label} 다운로드 재시도 {attempt}/{CSV_RETRY_COUNT}")
            time.sleep(CSV_RETRY_DELAY)

    raise RuntimeError(f"{label} 다운로드 실패: {last_error}") from last_error

def safe_move(src_path: str, dst_dir: str) -> str:
    """
    dst_dir 안으로 이동. 같은 파일명이 있으면 _1, _2 붙여서 안전하게 이동.
    """
    ensure_folder(dst_dir)
    filename = os.path.basename(src_path)
    base, ext = os.path.splitext(filename)
    dst_path = os.path.join(dst_dir, filename)

    if not os.path.exists(dst_path):
        shutil.move(src_path, dst_path)
        return dst_path

    i = 1
    while True:
        candidate = os.path.join(dst_dir, f"{base}_{i}{ext}")
        if not os.path.exists(candidate):
            shutil.move(src_path, candidate)
            return candidate
        i += 1

def cleanup_empty_folders(base_dir: str) -> int:
    removed = 0
    for root, dirs, files in os.walk(base_dir, topdown=False):
        if os.path.abspath(root) == os.path.abspath(base_dir):
            continue

        real_files = [f for f in files if not f.startswith(".")]
        real_dirs = [d for d in dirs if not d.startswith(".")]

        if len(real_files) == 0 and len(real_dirs) == 0:
            try:
                os.rmdir(root)
                removed += 1
            except OSError:
                pass
    return removed

def save_workbook_atomic(wb, out_path: str) -> None:
    """
    엑셀 파일이 열려있거나(잠금) 저장 중 오류가 나면,
    원본을 망가뜨리지 않게 임시파일로 저장 후 교체.
    """
    out_dir = os.path.dirname(out_path)
    ensure_folder(out_dir)

    fd, tmp_path = tempfile.mkstemp(prefix="__tmp__", suffix=".xlsx", dir=out_dir)
    os.close(fd)
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, out_path)
    finally:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass

# -------------------------------------------------------
# ✅ (추가) 분류 전에 기존 파일을: 완료/날짜/매장/...(구조유지) 로 이동
# -------------------------------------------------------
def archive_existing_files_done_date_store() -> Dict[str, int]:
    """
    분류 시작 전에 KASHION_ROOT 아래에 이미 존재하는 파일들을 전부
    KASHION/완료/YYYY-MM-DD/매장명/상대경로... 로 이동한다. (폴더 구조 유지)

    - KASHION/완료 아래는 건드리지 않음
    - 숨김파일(.DS_Store 등) 제외
    """
    summary: Dict[str, int] = {}

    if not os.path.exists(KASHION_ROOT):
        return summary

    date_tag = datetime.now().strftime("%Y-%m-%d")
    done_root = os.path.join(KASHION_ROOT, "완료")
    done_date = os.path.join(done_root, date_tag)
    ensure_folder(done_date)

    for store_name in os.listdir(KASHION_ROOT):
        store_path = os.path.join(KASHION_ROOT, store_name)

        if store_name.startswith("."):
            continue
        if store_name == "완료":
            continue
        if not os.path.isdir(store_path):
            continue

        moved = 0

        for root, dirs, files in os.walk(store_path):
            dirs[:] = [d for d in dirs if d != "완료" and not d.startswith(".")]

            for file in files:
                if file.startswith("."):
                    continue

                src_path = os.path.join(root, file)
                if not os.path.isfile(src_path):
                    continue

                rel_dir = os.path.relpath(root, store_path)
                rel_dir = "" if rel_dir == "." else rel_dir

                dst_dir = os.path.join(done_date, store_name, rel_dir)
                safe_move(src_path, dst_dir)
                moved += 1

        if moved > 0:
            summary[store_name] = moved

    return summary

# -------------------------------------------------------
# 엑셀 서식
# -------------------------------------------------------
def format_excel_full_table(
    out_path: str,
    startrow: int,
    startcol: int,
    df: pd.DataFrame,
    header_fill_until: Optional[str],
) -> None:
    wb = load_workbook(out_path)
    ws = wb.active
    ws.sheet_view.showGridLines = False

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align = Alignment(horizontal="center", vertical="center")

    header_fill = PatternFill(start_color="F6D57A", end_color="F6D57A", fill_type="solid")
    header_font = Font(bold=True)

    header_row = startrow + 1
    table_min_col = startcol + 1
    table_max_col = startcol + df.shape[1]
    first_col = table_min_col

    last_data_row = header_row
    first_series = df.iloc[:, 0].astype(str).tolist()
    for i, val in enumerate(first_series):
        if val.strip() != "":
            last_data_row = header_row + i + 1

    fill_end_idx = None
    if header_fill_until:
        cols = [str(c).strip() for c in df.columns]
        if header_fill_until in cols:
            fill_end_idx = cols.index(header_fill_until)

    for i, _colname in enumerate(df.columns):
        c = table_min_col + i
        cell = ws.cell(row=header_row, column=c)

        if fill_end_idx is not None and i <= fill_end_idx:
            cell.fill = header_fill

        cell.font = header_font
        cell.alignment = align
        cell.border = border

    for r in range(header_row + 1, last_data_row + 1):
        first_val = ws.cell(row=r, column=first_col).value
        if first_val is None or str(first_val).strip() == "":
            continue

        for c in range(table_min_col, table_max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = align
            cell.border = border

    for i in range(df.shape[1]):
        col_name = str(df.columns[i])
        col_values = df.iloc[:, i].astype(str)

        max_len = max(col_values.map(len).max(), len(col_name))
        col_letter = ws.cell(row=header_row, column=table_min_col + i).column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    save_workbook_atomic(wb, out_path)

# -------------------------------------------------------
# 0) 폴더 준비
# -------------------------------------------------------
def setup_directories():
    ensure_folder(ROE_DIR)
    ensure_folder(WAIT_FOLDER)

    ensure_folder(KASHION_ROOT)
    ensure_folder(os.path.join(KASHION_ROOT, "완료"))

    # ✅ LIST 폴더만 생성 (LIST/완료는 사용 안 함)
    ensure_folder(EXCEL_TARGET_FOLDER)
    ensure_folder(PDF_TARGET_FOLDER)

# -------------------------------------------------------
# 1) ZIP 해제
# -------------------------------------------------------
def process_zips():
    if not os.path.exists(WAIT_FOLDER):
        raise FileNotFoundError(f"대기 폴더 없음: {WAIT_FOLDER}")

    for root, _, files in os.walk(WAIT_FOLDER):
        for file in files:
            if file.startswith("."):
                continue
            if file.lower().endswith(".zip"):
                zip_path = os.path.join(root, file)
                try:
                    with zipfile.ZipFile(zip_path, "r") as zip_ref:
                        zip_ref.extractall(WAIT_FOLDER)
                    os.remove(zip_path)
                    print("[OK] ZIP 해제:", file)
                except zipfile.BadZipFile:
                    print("[WARN] ZIP 손상/불가:", file)

# -------------------------------------------------------
# 2) 분류
# -------------------------------------------------------
def classify_files() -> int:
    try:
        df = load_csv_from_url(URLS["classify"], "분류 CSV")
        if df.shape[1] < 2:
            print("[WARN] 분류 CSV 컬럼이 2개 미만입니다.")
            return 0

        id_col = df.columns[0]
        store_col = df.columns[1]

        df[id_col] = df[id_col].apply(clean_text)
        df[store_col] = df[store_col].apply(clean_text)
        df = df[(df[id_col] != "") & (df[store_col] != "")]
        df = df[df[id_col] != "주문번호"]

        mapping = dict(zip(df[id_col], df[store_col]))
        print(f"[OK] 매핑 개수: {len(mapping)}")
        keys_sorted = sorted(mapping.keys(), key=len, reverse=True)

        moved = 0
        skipped = 0

        for root, _, files in os.walk(WAIT_FOLDER):
            for file in files:
                if file.startswith("."):
                    continue

                src_path = os.path.join(root, file)
                if not os.path.isfile(src_path):
                    continue

                f_name = clean_text(file)
                matched_store = None

                for tID in keys_sorted:
                    if tID and (tID in f_name):
                        matched_store = mapping[tID]
                        break

                if not matched_store:
                    skipped += 1
                    continue

                # ✅ 구월/스퀘어원/부천은 노스페이스 아래로
                if matched_store in NORTHFACE_SUBSTORES:
                    target_dir = os.path.join(KASHION_ROOT, "노스페이스", matched_store, CLASSIFY_SUBFOLDER)
                else:
                    target_dir = os.path.join(KASHION_ROOT, matched_store, CLASSIFY_SUBFOLDER)

                safe_move(src_path, target_dir)
                moved += 1

        print(f"[OK] 분류 완료: 이동 {moved} / 스킵 {skipped}")

        removed = cleanup_empty_folders(WAIT_FOLDER)
        print(f"[OK] 구분대기 빈 폴더 삭제: {removed}개")

        return moved

    except Exception as exc:
        print(f"[FAIL] 분류 실패: {exc}")
        print("분류 중 오류(상세):")
        traceback.print_exc()
        return 0

# -------------------------------------------------------
# 3) 리네임 (완료 폴더 제외)
# -------------------------------------------------------
def rename_pdfs() -> int:
    try:
        df = load_csv_from_url(URLS["rename"], "리네임 CSV")
        if df.shape[1] < 2:
            print("[WARN] 리네임 CSV 컬럼이 2개 미만입니다.")
            return 0

        old_col = df.columns[0]
        new_col = df.columns[1]

        df[old_col] = df[old_col].apply(clean_text)
        df[new_col] = df[new_col].apply(clean_text)
        df = df[(df[old_col] != "") & (df[new_col] != "")]
        df = df[df[old_col] != "주문번호"]

        rename_map = dict(zip(df[old_col], df[new_col]))
        print(f"[OK] 리네임 매핑 개수: {len(rename_map)}")
        keys_sorted = sorted(rename_map.keys(), key=len, reverse=True)

        count = 0
        exclude_abs = os.path.join(KASHION_ROOT, EXCLUDE_FOLDER_NAME)

        for root, dirs, files in os.walk(KASHION_ROOT):
            if os.path.abspath(root).startswith(os.path.abspath(exclude_abs)):
                continue
            dirs[:] = [d for d in dirs if d != EXCLUDE_FOLDER_NAME]

            for file in files:
                if not file.lower().endswith(".pdf"):
                    continue

                f_name = clean_text(file)
                matched_old = None

                for old_id in keys_sorted:
                    if len(old_id) < 5 or old_id == "주문번호":
                        continue
                    if old_id in f_name:
                        matched_old = old_id
                        break

                if not matched_old:
                    continue

                old_path = os.path.join(root, file)
                new_base = rename_map[matched_old]
                new_path = os.path.join(root, f"{new_base}.pdf")

                if os.path.exists(new_path):
                    i = 1
                    while True:
                        cand = os.path.join(root, f"{new_base}_{i}.pdf")
                        if not os.path.exists(cand):
                            new_path = cand
                            break
                        i += 1

                os.rename(old_path, new_path)
                count += 1

        return count

    except Exception as exc:
        print(f"[FAIL] 이름 변경 실패: {exc}")
        print("이름 변경 중 오류(상세):")
        traceback.print_exc()
        return 0

# -------------------------------------------------------
# 4) 매장별 PDF 통합 (기존 개별 PDF 유지)
# -------------------------------------------------------
def make_merged_pdf_name(pdf_dir: str) -> str:
    rel_path = os.path.relpath(pdf_dir, KASHION_ROOT)
    parts = [p for p in rel_path.split(os.sep) if p and p != CLASSIFY_SUBFOLDER]
    if not parts:
        parts = [os.path.basename(pdf_dir)]
    safe_name = " ".join(parts)
    date_suffix = datetime.now().strftime("%m%d")
    return f"{safe_name} {date_suffix}.pdf"


def merge_pdfs_with_pypdf(pdf_paths: list[str], out_path: str) -> bool:
    writer = PdfWriter()

    for pdf_path in pdf_paths:
        reader = PdfReader(pdf_path)
        for page in reader.pages:
            writer.add_page(page)

    with open(out_path, "wb") as out_file:
        writer.write(out_file)

    return True


def merge_store_pdfs() -> int:
    """
    KASHION_ROOT 아래 완료 폴더를 제외하고, PDF가 들어있는 각 분류 폴더마다
    파일명 오름차순으로 하나의 통합 PDF를 만든다. 원본 PDF는 삭제하지 않는다.
    """
    if not os.path.exists(KASHION_ROOT):
        return 0

    ensure_folder(PDF_TARGET_FOLDER)
    merged_count = 0
    exclude_abs = os.path.abspath(os.path.join(KASHION_ROOT, EXCLUDE_FOLDER_NAME))

    for root, dirs, files in os.walk(KASHION_ROOT):
        root_abs = os.path.abspath(root)
        if root_abs.startswith(exclude_abs):
            continue
        dirs[:] = [d for d in dirs if d != EXCLUDE_FOLDER_NAME and not d.startswith(".")]

        pdf_files = [
            f for f in files
            if f.lower().endswith(".pdf")
            and not f.startswith(".")
        ]
        if len(pdf_files) < 1:
            continue

        pdf_files = sorted(pdf_files, key=lambda name: clean_text(name).lower())
        pdf_paths = [os.path.join(root, f) for f in pdf_files]
        out_path = os.path.join(PDF_TARGET_FOLDER, make_merged_pdf_name(root))

        try:
            merge_pdfs_with_pypdf(pdf_paths, out_path)
            merged_count += 1
            print(f"[OK] PDF 통합: {os.path.basename(out_path)} ({len(pdf_paths)}개)")
        except Exception as exc:
            print(f"[FAIL] PDF 통합 실패: {root} -> {exc}")
            traceback.print_exc()

    return merged_count

# -------------------------------------------------------
# 5) 엑셀 저장 (✅ LIST 폴더 바로 저장)
# -------------------------------------------------------
def update_excel_fixed() -> bool:
    date_suffix = datetime.now().strftime("%m%d")

    # ✅ LIST 폴더에 바로 저장
    ensure_folder(EXCEL_TARGET_FOLDER)

    try:
        raw_df = load_csv_from_url(URLS["excel"], "엑셀 CSV")
        raw_df.columns = [re.sub(r"\.\d+$", "", str(c)).strip() for c in raw_df.columns]

        for b in BRANDS_FIXED:
            s, e = b["slice"]
            if raw_df.shape[1] < e:
                raise ValueError(f"{b['name']}: CSV 컬럼 부족 (필요 {e}열, 현재 {raw_df.shape[1]}열)")

            brand_df = raw_df.iloc[:, s:e].copy()

            out_name = f"{b['name']} KS JD {date_suffix}.xlsx"
            out_path = os.path.join(EXCEL_TARGET_FOLDER, out_name)

            brand_df.to_excel(
                out_path,
                index=False,
                startrow=b["startrow"],
                startcol=b["startcol"],
                engine="openpyxl",
            )

            format_excel_full_table(
                out_path,
                startrow=b["startrow"],
                startcol=b["startcol"],
                df=brand_df,
                header_fill_until=b["header_fill_until"],
            )

            print("[OK] 엑셀 저장:", b["name"], "->", out_name)

        return True

    except PermissionError:
        print("[FAIL] 엑셀 저장 실패: 파일이 열려있거나(잠금) 권한이 없습니다.")
        print("[INFO] Excel에서 해당 파일을 닫고 다시 실행하세요.")
        traceback.print_exc()
        return False
    except Exception as exc:
        print(f"[FAIL] 엑셀 업데이트 실패: {exc}")
        print("[FAIL] 엑셀 업데이트 중 오류(상세):")
        traceback.print_exc()
        return False

# -------------------------------------------------------
# 실행
# -------------------------------------------------------
if __name__ == "__main__":
    print("[START] KASHION 통합 자동화 시작...")

    setup_directories()
    process_zips()

    archived_map = archive_existing_files_done_date_store()
    if archived_map:
        print("[INFO] 기존 파일 완료/날짜/매장으로 이동:")
        for store, cnt in archived_map.items():
            print(f" - {store}: {cnt}건")
    else:
        print("[INFO] 기존 파일 이동: 없음")

    c_count = classify_files()
    r_count = rename_pdfs()
    m_count = merge_store_pdfs()
    excel_ok = update_excel_fixed()

    print("-" * 30)
    print(f"[OK] 분류 완료: {c_count}건")
    print(f"[OK] 이름 변경: {r_count}건")
    print(f"[OK] PDF 통합: {m_count}건")
    print(f"[OK] 엑셀 업데이트: {'성공' if excel_ok else '실패'}")
    print("[DONE] 완료")
