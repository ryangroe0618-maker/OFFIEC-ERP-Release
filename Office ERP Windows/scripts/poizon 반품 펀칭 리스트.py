# -*- coding: utf-8 -*-

import os
import re
from datetime import datetime
from io import BytesIO, StringIO

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from office_erp_paths import LIST_DIR as DESKTOP_LIST_DIR


CSV_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vQFuLbNobNHN9skvSn0CJemLbJmplfqcUiLYr_KTlChLSijczOy_INmeYmaMtimT3LyY8YT8FfU2ws8/pub?gid=828126855&single=true&output=csv"

LIST_DIR = str(DESKTOP_LIST_DIR)

OUTPUT_BASENAME = "노스페이스 원거래 반품"
START_COL = 2
HEADER_ROW = 2
DATA_START_ROW = 3
SOURCE_START_COL = "A"
SOURCE_END_COL = "F"

MIN_COL_WIDTH = 10
MAX_COL_WIDTH = 70
COL_WIDTH_PADDING = 4
STORE_NAME_MIN_COL_WIDTH = 12

HEADER_FILL = PatternFill(start_color="F6D57A", end_color="F6D57A", fill_type="solid")
SESSION = requests.Session()
SESSION.headers.update({"User-Agent": "Mozilla/5.0"})


def ensure_dir(path):
    os.makedirs(path, exist_ok=True)


def mmdd():
    return datetime.now().strftime("%m%d")


def safe_path(path):
    if not os.path.exists(path):
        return path

    base, ext = os.path.splitext(path)
    index = 1
    while True:
        new_path = f"{base}_{index}{ext}"
        if not os.path.exists(new_path):
            return new_path
        index += 1


def download_csv_df(url):
    response = SESSION.get(url, timeout=30)
    response.raise_for_status()
    content = response.content

    for encoding in ("utf-8-sig", "utf-8", "cp949"):
        try:
            text = content.decode(encoding)
            return pd.read_csv(StringIO(text), dtype=str, keep_default_na=False)
        except Exception:
            continue

    return pd.read_csv(BytesIO(content), dtype=str, keep_default_na=False)


def excel_col_to_index(col_letter):
    col_letter = col_letter.strip().upper()
    index = 0
    for char in col_letter:
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index


def slice_df_by_excel_cols(df, start_col, end_col):
    start_index = excel_col_to_index(start_col) - 1
    end_index = excel_col_to_index(end_col)
    if df.shape[1] < end_index:
        raise ValueError(f"CSV 컬럼 수가 부족합니다. 필요 범위: {start_col}~{end_col}, 실제 컬럼 수: {df.shape[1]}")
    return df.iloc[:, start_index:end_index].copy()


def normalize_headers(df):
    df.columns = [str(column).split(".")[0].strip() for column in df.columns]
    return df


def to_excel_number(value):
    text = "" if value is None else str(value).strip()
    if text == "":
        return None

    text = text.replace(",", "")
    number = pd.to_numeric(text, errors="coerce")
    if pd.isna(number):
        return value

    number = float(number)
    if number.is_integer():
        return int(number)
    return number


def set_auto_width(ws, end_row, start_col, end_col):
    for col in range(start_col, end_col + 1):
        max_len = 0
        for row in range(HEADER_ROW, end_row + 1):
            value = ws.cell(row=row, column=col).value
            text = "" if value is None else str(value)
            text = re.sub(r"\s+", " ", text).strip()
            max_len = max(max_len, len(text))

        width = max(MIN_COL_WIDTH, min(max_len + COL_WIDTH_PADDING, MAX_COL_WIDTH))
        header_name = str(ws.cell(row=HEADER_ROW, column=col).value or "").strip()
        if header_name == "매장명":
            width = max(width, STORE_NAME_MIN_COL_WIDTH)
        ws.column_dimensions[get_column_letter(col)].width = width


def build_excel_from_df(df, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "반품 펀칭"
    ws.sheet_view.showGridLines = False

    df = normalize_headers(df)
    headers = [str(column).strip() for column in df.columns]
    number_col_indexes = {
        index
        for index, header in enumerate(headers)
        if header in {"수량"}
    }

    n_rows = len(df)
    n_cols = df.shape[1]

    header_font = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_index in range(n_cols):
        cell = ws.cell(row=HEADER_ROW, column=START_COL + col_index, value=df.columns[col_index])
        cell.font = header_font
        cell.alignment = align_center
        cell.fill = HEADER_FILL
        cell.border = border

    for row_index in range(n_rows):
        for col_index in range(n_cols):
            cell_value = df.iat[row_index, col_index]
            if col_index in number_col_indexes:
                cell_value = to_excel_number(cell_value)
            cell = ws.cell(row=DATA_START_ROW + row_index, column=START_COL + col_index, value=cell_value)
            cell.alignment = align_center

    last_row = DATA_START_ROW + n_rows - 1
    end_col = START_COL + n_cols - 1

    for row in range(HEADER_ROW, last_row + 1):
        first_value = ws.cell(row=row, column=START_COL).value
        if first_value is None or str(first_value).strip() == "":
            continue

        for col in range(START_COL, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = align_center

    set_auto_width(ws, last_row, START_COL, end_col)
    wb.save(out_path)


def main():
    ensure_dir(LIST_DIR)
    df_all = download_csv_df(CSV_URL)
    output_df = slice_df_by_excel_cols(df_all, SOURCE_START_COL, SOURCE_END_COL)

    out_name = f"{OUTPUT_BASENAME} {mmdd()}.xlsx"
    out_path = safe_path(os.path.join(LIST_DIR, out_name))

    build_excel_from_df(output_df, out_path)
    print(f"저장 완료: {out_path}")


if __name__ == "__main__":
    main()
