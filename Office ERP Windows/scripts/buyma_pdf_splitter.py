# -*- coding: utf-8 -*-

from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime
from io import StringIO
from pathlib import Path
import re
import sys
import tempfile
import os
import time

import fitz
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import pandas as pd
import requests


CSV_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vTZBOxAyzArARCW9RzCuiPwO2he7rx99JhEaD7l91mhY6ZiZ_urK_Z2DJo9gD53KybDwcdwYPVY-_ro/pub?output=csv"
EXCEL_CSV_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vTZBOxAyzArARCW9RzCuiPwO2he7rx99JhEaD7l91mhY6ZiZ_urK_Z2DJo9gD53KybDwcdwYPVY-_ro/pub?gid=1012969196&single=true&output=csv"
def get_desktop_dir() -> Path:
    candidates = [
        Path(os.environ.get("USERPROFILE", "")) / "Desktop",
        Path(os.environ.get("OneDrive", "")) / "Desktop",
        Path(os.environ.get("OneDriveConsumer", "")) / "Desktop",
        Path.home() / "Desktop",
    ]
    for candidate in candidates:
        if str(candidate) and candidate.exists():
            return candidate
    return Path.home() / "Desktop"


DESKTOP_DIR = get_desktop_dir()
PDF_SOURCE_DIR_CANDIDATES = [
    DESKTOP_DIR / "buyma pdf" / "buyma 구분 대기",
    DESKTOP_DIR / "buyma 구분 대기",
]
OUTPUT_DIR = DESKTOP_DIR / "PDF"
EXCEL_OUTPUT_DIR = DESKTOP_DIR / "LIST"

ORDER_COLUMN_INDEX = 3  # D열
STORE_COLUMN_INDEX = 10  # K열
DOWNLOAD_TIMEOUT = (10, 60)
DOWNLOAD_RETRIES = 3
DOWNLOAD_RETRY_SLEEP_SEC = 2
BUYMA_ORDER_PATTERN = re.compile(r"\bK\d{8,20}\b")
INVALID_FILENAME_CHARS = re.compile(r'[\\/:*?"<>|]')
BRANDS_FIXED = [
    {"name": "노스페이스", "slice": (0, 11), "startrow": 1, "startcol": 1, "header_fill_until": "수량"},
    {"name": "휠라", "slice": (12, 19), "startrow": 1, "startcol": 1, "header_fill_until": "수량"},
    {"name": "푸마", "slice": (20, 26), "startrow": 1, "startcol": 1, "header_fill_until": "수량"},
]


class PdfSplitError(RuntimeError):
    pass


def log(message: str) -> None:
    print(f"[BUYMA PDF 분배] {message}", flush=True)


def clean_text(value) -> str:
    if pd.isna(value):
        return ""
    return " ".join(str(value).strip().split())


def normalize_order_no(value) -> str:
    return clean_text(value).upper()


def normalize_store_name(value: str) -> str:
    store_name = clean_text(value)
    if store_name.startswith("사무실"):
        return "사무실"
    return store_name or "미매칭"


def safe_filename(name: str) -> str:
    cleaned = INVALID_FILENAME_CHARS.sub("_", clean_text(name))
    return cleaned.strip(". ") or "미매칭"


def download_text(url: str, label: str) -> str:
    log(f"{label} 다운로드 시작")
    last_error = None
    for attempt in range(1, DOWNLOAD_RETRIES + 1):
        try:
            response = requests.get(url, timeout=DOWNLOAD_TIMEOUT)
            response.raise_for_status()
            response.encoding = "utf-8"
            if not response.text.strip():
                raise PdfSplitError(f"{label} 응답이 비어 있습니다.")
            return response.text
        except Exception as exc:
            last_error = exc
            if attempt < DOWNLOAD_RETRIES:
                log(f"{label} 다운로드 재시도 ({attempt}/{DOWNLOAD_RETRIES}): {exc}")
                time.sleep(DOWNLOAD_RETRY_SLEEP_SEC)
    raise PdfSplitError(f"{label} 다운로드 실패: {last_error}") from last_error


def clean_excel_header(value) -> str:
    return re.sub(r"\.\d+$", "", str(value)).strip()


def load_order_store_lookup() -> dict[str, str]:
    csv_text = download_text(CSV_URL, "출고 리스트 CSV")
    try:
        df = pd.read_csv(
            StringIO(csv_text),
            header=0,
            dtype=str,
            keep_default_na=False,
            usecols=[ORDER_COLUMN_INDEX, STORE_COLUMN_INDEX],
        ).fillna("")
    except ValueError as exc:
        raise PdfSplitError("출고 리스트 CSV에서 D열/K열을 읽을 수 없습니다.") from exc
    df.columns = ["주문번호", "매장명"]

    lookup: dict[str, str] = {}
    duplicate_count = 0
    for _, row in df.iterrows():
        order_no = normalize_order_no(row["주문번호"])
        store_name = normalize_store_name(row["매장명"])
        if not order_no:
            continue
        if order_no in lookup:
            duplicate_count += 1
            continue
        lookup[order_no] = store_name

    log(f"출고 리스트 로드 완료: {len(lookup)}건")
    if duplicate_count:
        log(f"중복 주문번호 {duplicate_count}건은 첫 번째 매장명 기준으로 처리")
    if not lookup:
        raise PdfSplitError("출고 리스트에서 주문번호/매장명 데이터를 찾지 못했습니다.")
    return lookup


def save_workbook_atomic(workbook, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        prefix="__tmp__",
        suffix=".xlsx",
        dir=output_path.parent,
        delete=False,
    ) as temp_file:
        temp_path = Path(temp_file.name)

    try:
        workbook.save(temp_path)
        temp_path.replace(output_path)
    finally:
        if temp_path.exists():
            temp_path.unlink()


def format_excel_full_table(
    output_path: Path,
    startrow: int,
    startcol: int,
    df: pd.DataFrame,
    header_fill_until: str | None,
) -> None:
    workbook = load_workbook(output_path)
    worksheet = workbook.active
    worksheet.sheet_view.showGridLines = False

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="F6D57A", end_color="F6D57A", fill_type="solid")
    header_font = Font(bold=True)

    header_row = startrow + 1
    table_min_col = startcol + 1
    table_max_col = startcol + df.shape[1]
    first_col = table_min_col

    last_data_row = header_row
    first_series = df.iloc[:, 0].astype(str).tolist()
    for index, value in enumerate(first_series):
        if value.strip():
            last_data_row = header_row + index + 1

    fill_end_index = None
    if header_fill_until:
        columns = [str(column).strip() for column in df.columns]
        if header_fill_until in columns:
            fill_end_index = columns.index(header_fill_until)

    for index, _column_name in enumerate(df.columns):
        cell = worksheet.cell(row=header_row, column=table_min_col + index)
        if fill_end_index is not None and index <= fill_end_index:
            cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align
        cell.border = border

    for row_index in range(header_row + 1, last_data_row + 1):
        first_value = worksheet.cell(row=row_index, column=first_col).value
        if first_value is None or not str(first_value).strip():
            continue
        for col_index in range(table_min_col, table_max_col + 1):
            cell = worksheet.cell(row=row_index, column=col_index)
            cell.alignment = align
            cell.border = border

    for index in range(df.shape[1]):
        column_name = str(df.columns[index])
        column_values = df.iloc[:, index].astype(str)
        max_len = max(column_values.map(len).max(), len(column_name))
        column_letter = worksheet.cell(row=header_row, column=table_min_col + index).column_letter
        worksheet.column_dimensions[column_letter].width = max(max_len + 4, 12)

    save_workbook_atomic(workbook, output_path)


def create_excel_files() -> None:
    date_suffix = datetime.now().strftime("%m%d")
    EXCEL_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    log(f"엑셀 저장 폴더: {EXCEL_OUTPUT_DIR}")

    csv_text = download_text(EXCEL_CSV_URL, "엑셀 출고 리스트 CSV")
    raw_df = pd.read_csv(StringIO(csv_text), dtype=str, keep_default_na=False).fillna("")
    raw_df.columns = [clean_excel_header(column) for column in raw_df.columns]

    for brand_info in BRANDS_FIXED:
        start, end = brand_info["slice"]
        brand_name = brand_info["name"]
        if raw_df.shape[1] < end:
            raise PdfSplitError(f"{brand_name}: CSV 컬럼 부족 (필요 {end}열, 현재 {raw_df.shape[1]}열)")

        brand_df = raw_df.iloc[:, start:end].copy()
        output_name = f"{brand_name} RB {date_suffix}.xlsx"
        output_path = EXCEL_OUTPUT_DIR / output_name

        brand_df.to_excel(
            output_path,
            index=False,
            startrow=brand_info["startrow"],
            startcol=brand_info["startcol"],
            engine="openpyxl",
        )
        format_excel_full_table(
            output_path,
            startrow=brand_info["startrow"],
            startcol=brand_info["startcol"],
            df=brand_df,
            header_fill_until=brand_info["header_fill_until"],
        )
        log(f"엑셀 저장 완료: {output_name}")


def resolve_pdf_source_dir() -> Path:
    existing_dirs = [source_dir for source_dir in PDF_SOURCE_DIR_CANDIDATES if source_dir.exists()]
    for source_dir in existing_dirs:
        if any(source_dir.glob("*.pdf")):
            return source_dir
    if existing_dirs:
        return existing_dirs[0]

    source_dir = PDF_SOURCE_DIR_CANDIDATES[-1]
    source_dir.mkdir(parents=True, exist_ok=True)
    log(f"PDF 원본 폴더를 생성했습니다: {source_dir}")
    return source_dir


def ensure_base_folders() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    EXCEL_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if not any(source_dir.exists() for source_dir in PDF_SOURCE_DIR_CANDIDATES):
        PDF_SOURCE_DIR_CANDIDATES[-1].mkdir(parents=True, exist_ok=True)


def describe_source_candidates() -> str:
    return "\n".join(f"- {path}" for path in PDF_SOURCE_DIR_CANDIDATES)


def find_pdf_files() -> tuple[Path, list[Path]]:
    source_dir = resolve_pdf_source_dir()
    pdf_files = sorted(source_dir.glob("*.pdf"))
    if not pdf_files:
        raise PdfSplitError(f"PDF 원본 폴더에 PDF 파일이 없습니다:\n{describe_source_candidates()}")
    log(f"PDF 원본 폴더: {source_dir}")
    log(f"PDF 파일 발견: {len(pdf_files)}개")
    return source_dir, pdf_files


def delete_source_pdfs(pdf_files: list[Path]) -> None:
    deleted = 0
    failed = 0
    for pdf_path in pdf_files:
        try:
            pdf_path.unlink()
            deleted += 1
            log(f"원본 PDF 삭제: {pdf_path.name}")
        except Exception as exc:
            failed += 1
            log(f"원본 PDF 삭제 실패: {pdf_path.name} / {exc}")
    log(f"원본 PDF 정리 완료: 삭제 {deleted}개, 실패 {failed}개")


def extract_order_no_from_page(page: fitz.Page) -> str:
    text = page.get_text("text") or ""
    matches = BUYMA_ORDER_PATTERN.findall(text.upper())
    if not matches:
        return ""
    return matches[0]


def split_pdf_pages(pdf_files: list[Path], lookup: dict[str, str]) -> tuple[dict[str, fitz.Document], Counter, list[Path]]:
    output_docs: dict[str, fitz.Document] = defaultdict(fitz.open)
    page_counts: Counter = Counter()
    processed_pdf_files: list[Path] = []
    failed_files = 0

    for pdf_path in pdf_files:
        log(f"PDF 읽기: {pdf_path.name}")
        try:
            added_pages = 0
            with fitz.open(pdf_path) as source_doc:
                if source_doc.page_count == 0:
                    log(f"빈 PDF 건너뜀: {pdf_path.name}")
                    continue
                for page_index in range(source_doc.page_count):
                    page = source_doc.load_page(page_index)
                    order_no = extract_order_no_from_page(page)
                    store_name = lookup.get(order_no, "미매칭") if order_no else "미매칭"
                    output_docs[store_name].insert_pdf(
                        source_doc,
                        from_page=page_index,
                        to_page=page_index,
                    )
                    page_counts[store_name] += 1
                    added_pages += 1

                    display_order_no = order_no or "주문번호 없음"
                    log(f"{display_order_no} → {store_name}")
            if added_pages:
                processed_pdf_files.append(pdf_path)
        except Exception as exc:
            failed_files += 1
            log(f"PDF 읽기 실패: {pdf_path.name} / {exc}")

    if not page_counts:
        close_output_docs(output_docs)
        raise PdfSplitError("분배할 PDF 페이지가 없습니다.")
    if failed_files:
        log(f"읽기 실패 PDF {failed_files}개는 원본 삭제 대상에서 제외합니다.")

    return dict(output_docs), page_counts, processed_pdf_files


def close_output_docs(output_docs: dict[str, fitz.Document]) -> None:
    for doc in output_docs.values():
        try:
            doc.close()
        except Exception:
            pass


def save_output_pdfs(output_docs: dict[str, fitz.Document]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    log(f"결과 저장 폴더: {OUTPUT_DIR}")
    date_suffix = datetime.now().strftime("%m%d")

    for store_name, doc in sorted(output_docs.items()):
        if doc.page_count == 0:
            doc.close()
            continue
        output_path = OUTPUT_DIR / f"{safe_filename(store_name)} RB {date_suffix}.pdf"
        temp_output_path = output_path.with_name(f"__tmp__{output_path.name}")
        if temp_output_path.exists():
            temp_output_path.unlink()
        page_count = doc.page_count
        doc.save(temp_output_path)
        doc.close()
        if output_path.exists():
            output_path.unlink()
        temp_output_path.replace(output_path)
        log(f"저장 완료: {output_path.name} ({store_name}, {page_count}페이지)")


def print_summary(page_counts: Counter) -> None:
    log("최종 페이지 수")
    for store_name, count in sorted(page_counts.items()):
        log(f"- {store_name}: {count}페이지")


def main() -> int:
    output_docs = {}
    try:
        ensure_base_folders()
        create_excel_files()
        lookup = load_order_store_lookup()
        _source_dir, pdf_files = find_pdf_files()
        output_docs, page_counts, processed_pdf_files = split_pdf_pages(pdf_files, lookup)
        save_output_pdfs(output_docs)
        delete_source_pdfs(processed_pdf_files)
        print_summary(page_counts)
        log("완료")
        return 0
    except Exception as exc:
        close_output_docs(output_docs)
        log(f"실패: {exc}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
