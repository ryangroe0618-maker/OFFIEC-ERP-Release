# -*- coding: utf-8 -*-

import os
import re
from datetime import datetime
from io import StringIO, BytesIO

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter

from office_erp_paths import LIST_DIR as DESKTOP_LIST_DIR


CSV_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vQSGUP7ZizELC9jJqGdxawP_HRsxo-fUZ_5BgMx8NUaEPTJYhH31iiH2-_NEE4Ff9UW3OATcxopoKEU/pub?gid=686878635&single=true&output=csv"

LIST_DIR = str(DESKTOP_LIST_DIR)

HEADER_ROW = 2
DATA_START_ROW = 3
START_COL = 2

MIN_COL_WIDTH = 10
MAX_COL_WIDTH = 70
COL_WIDTH_PADDING = 4

HEADER_FILL = PatternFill(start_color="F6D57A", end_color="F6D57A", fill_type="solid")

SESSION = requests.Session()


def ensure_dir(path):
    os.makedirs(path, exist_ok=True)


def mmdd():
    return datetime.now().strftime("%m%d")


def safe_path(path):
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    index = 1
    while True:
        new_path = f"{base}_{index}{ext}"
        if not os.path.exists(new_path):
            return new_path
        index += 1


def download_csv_df(url):
    response = SESSION.get(url, timeout=30)
    response.raise_for_status()
    content = response.content

    for encoding in ("utf-8-sig", "utf-8", "cp949"):
        try:
            text = content.decode(encoding)
            return pd.read_csv(StringIO(text), dtype=str, keep_default_na=False)
        except Exception:
            continue

    return pd.read_csv(BytesIO(content), dtype=str, keep_default_na=False)


def excel_col_to_index(col_letter):
    col_letter = col_letter.strip().upper()
    number = 0
    for ch in col_letter:
        number = number * 26 + (ord(ch) - ord("A") + 1)
    return number


def slice_df_by_excel_cols(df, start_col, end_col):
    start_index = excel_col_to_index(start_col) - 1
    end_index = excel_col_to_index(end_col)
    if df.shape[1] < end_index:
        raise ValueError("CSV 컬럼 수 부족")
    return df.iloc[:, start_index:end_index].copy()


def set_auto_width(ws, end_row, start_col, end_col):
    for col_idx in range(start_col, end_col + 1):
        max_len = 0
        for row_idx in range(HEADER_ROW, end_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            text = "" if value is None else str(value)
            text = re.sub(r"\s+", " ", text).strip()
            max_len = max(max_len, len(text))
        width = max(MIN_COL_WIDTH, min(max_len + COL_WIDTH_PADDING, MAX_COL_WIDTH))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def build_excel_from_df(df, out_path):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.sheet_view.showGridLines = False

    df.columns = [str(col).split(".")[0].strip() for col in df.columns]

    row_count = len(df)
    col_count = df.shape[1]

    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_offset in range(col_count):
        cell = worksheet.cell(row=HEADER_ROW, column=START_COL + col_offset, value=df.columns[col_offset])
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = HEADER_FILL

    for row_offset in range(row_count):
        for col_offset in range(col_count):
            worksheet.cell(
                row=DATA_START_ROW + row_offset,
                column=START_COL + col_offset,
                value=df.iat[row_offset, col_offset],
            )

    last_row = DATA_START_ROW + row_count - 1
    end_col = START_COL + col_count - 1

    for row_idx in range(HEADER_ROW, last_row + 1):
        first_value = worksheet.cell(row=row_idx, column=START_COL).value
        if first_value is None or str(first_value).strip() == "":
            continue
        for col_idx in range(START_COL, end_col + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = center_align

    set_auto_width(worksheet, last_row, START_COL, end_col)
    workbook.save(out_path)


def main():
    ensure_dir(LIST_DIR)
    date_suffix = mmdd()
    df_all = download_csv_df(CSV_URL)
    part_df = slice_df_by_excel_cols(df_all, "A", "D")

    out_name = f"푸마 KR {date_suffix}.xlsx"
    out_path = safe_path(os.path.join(LIST_DIR, out_name))

    build_excel_from_df(part_df, out_path)
    print(f"저장 완료: {out_path}")


if __name__ == "__main__":
    main()
